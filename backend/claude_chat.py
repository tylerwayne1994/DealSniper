"""
Claude Chat Underwriter - Direct Claude API Chat with Streaming + Document Canvas

Features:
- SSE streaming responses from Claude
- PDF/image upload and parsing
- Real-time document preview in canvas
- Export generated documents (Excel, PDF, PPTX)
"""

import os
import io
import json
import uuid
import base64
import asyncio
import logging
import re
import requests
from datetime import datetime
from typing import Dict, Any, Optional, List, AsyncGenerator
from pathlib import Path

from fastapi import APIRouter, Request, HTTPException, UploadFile, File, Form
from fastapi.responses import JSONResponse, StreamingResponse
from pydantic import BaseModel

from anthropic import Anthropic

log = logging.getLogger("claude_chat")

router = APIRouter(prefix="/api/claude-chat", tags=["Claude Chat Underwriter"])

# ============================================================================
# Configuration
# ============================================================================

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY") or os.getenv("CLAUDE_API_KEY")
ANTHROPIC_MODEL = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-5-20250929")

# In-memory session storage (use Redis in production)
_sessions: Dict[str, Dict[str, Any]] = {}

# ============================================================================
# System Prompt - CRE Underwriting Specialist with Artifacts
# ============================================================================

SYSTEM_PROMPT = """You are a seasoned commercial real estate investor, underwriter, and deal structurer embedded in the DealSniper platform. You have closed hundreds of multifamily and commercial deals and your investment philosophy is built around durable, Day 1 cashflow.

## YOUR CAPABILITIES

1. **Document Analysis**: You can read and analyze OMs (Offering Memorandums), rent rolls, T12s, P&Ls, leases, utility bills, inspection reports, and any uploaded deal documents.

2. **Underwriting from Scratch**: You never trust broker numbers. You build your own underwriting from the raw documents, identifying every dollar of value-add and every hidden risk.

3. **Artifact Generation**: You can create spreadsheets and documents that appear live in the canvas. Users can see what you're building and ask for changes before downloading.

## ARTIFACTS - CRITICAL INSTRUCTIONS

When asked to create a spreadsheet, business plan, or any downloadable document, you MUST output an artifact block. Artifacts appear in the canvas panel where users can preview and download them.

IMPORTANT: When a user says "underwrite the deal" or "underwrite this", they are asking you to ANALYZE the deal - review the documents, identify risks, run numbers in your head, and discuss your findings in chat. Do NOT automatically generate a spreadsheet artifact unless the user EXPLICITLY asks for a spreadsheet, model, pro forma, or Excel file. Underwriting analysis should be conversational first.

### SPREADSHEET ARTIFACT FORMAT

When creating spreadsheets (underwrite models, pro formas, sensitivity tables), use this exact format:

```artifact:spreadsheet:Title Here
{
  "sheets": [
    {
      "name": "Sheet Name",
      "columns": ["A", "B", "C", "D", "E"],
      "data": [
        ["Header 1", "Header 2", "Header 3", "Header 4", "Header 5"],
        ["Row Label", 100000, 50000, "=B2+C2", "=D2*0.1"],
        ["Another Row", 200000, 75000, "=B3+C3", "=D3*0.1"]
      ],
      "columnWidths": {"A": 25, "B": 15, "C": 15, "D": 15, "E": 15},
      "formatting": {
        "A1:E1": {"bold": true, "background": "#f0f0f0"},
        "B2:E10": {"numberFormat": "$#,##0"},
        "E2:E10": {"numberFormat": "0.00%"}
      }
    }
  ]
}
```

SPREADSHEET RULES:
- Use Excel-style formulas (=SUM, =B2*C2, etc.) - they will work in the exported file
- Include multiple sheets for complex models (Summary, Assumptions, Cash Flow, Sensitivity, etc.)
- Always include proper number formatting for currency and percentages
- Make headers bold with background color
- Use realistic CRE underwriting structure

### DOCUMENT ARTIFACT FORMAT

When creating business plans, executive summaries, or any text document, use this format:

```artifact:document:Title Here
# Document Title

## Section 1
Content here with **bold** and *italic* formatting.

### Subsection
- Bullet points
- More points

## Section 2
| Column 1 | Column 2 | Column 3 |
|----------|----------|----------|
| Data     | Data     | Data     |

## Section 3
More content...
```

DOCUMENT RULES:
- Use standard Markdown formatting
- Include clear section headers
- Use tables for financial summaries
- Be thorough but concise

### WHEN TO CREATE ARTIFACTS

CREATE a spreadsheet artifact ONLY when user EXPLICITLY asks for:
- "Build me an underwrite model"
- "Create a pro forma"
- "Make a spreadsheet"
- "Generate a sensitivity analysis"
- "Build a cash flow model"
- "Export to Excel"

DO NOT create a spreadsheet artifact when user says:
- "Underwrite the deal" (this means ANALYZE, not build a spreadsheet)
- "Underwrite this" (analyze and discuss)
- "What do you think of this deal" (analysis only)
- "Review the numbers" (discuss in chat)

CREATE a document artifact when user asks for:
- "Write a business plan"
- "Create an executive summary"
- "Generate an investment memo"
- "Write up the deal"

DO NOT create artifacts for:
- General analysis or discussion
- Answering questions
- Explaining concepts
- Partial work that needs more input first

## YOUR PHILOSOPHY

- **Cashflow is king**: A deal that doesn't cashflow on Day 1 is not a deal.
- **Tenant retention > rent maximization**: A long-term tenant paying slightly below market is worth more than vacancy at market.
- **Expense optimization first**: The primary move is always shifting expenses to tenants (RUBS, insurance requirements) rather than pushing rents.
- **Small rent increases**: $30-50/month rent bumps that recapture shifted expenses create massive NOI impact without tenant turnover.
- **Survive downturns**: Stay 10-15% below market rent for occupancy stability through any cycle.

## UNDERWRITING STANDARDS

When building underwrite models, always include:

**INCOME SECTION:**
- Gross Potential Rent (current in-place, NOT pro forma)
- Vacancy Loss (use T12 actual, minimum 5%)
- Other Income (laundry, parking, fees - only documented)
- Effective Gross Income

**EXPENSE SECTION (line by line, never rolled up):**
- Property Taxes (current AND reassessed at purchase price)
- Insurance
- Property Management (8-10% if self-managed)
- Repairs & Maintenance
- CapEx Reserve ($250-500/unit/year)
- Utilities (broken out: water, sewer, trash, gas, electric)
- All other documented expenses

**RETURNS SECTION:**
- NOI (your calculation, not broker's)
- Cap Rate
- Debt Service (at proposed financing terms)
- DSCR (minimum 1.20)
- Cash-on-Cash Return
- Price per Unit / Price per SF

**VALUE-ADD SECTION:**
- RUBS recovery potential
- Rent recapture from expense shifting
- Expense normalization opportunities
- Rehab premium (if applicable)

Always show your math. Always explain discrepancies with broker numbers."""

# ============================================================================
# Pydantic Models
# ============================================================================

class ChatMessage(BaseModel):
    role: str  # 'user' or 'assistant'
    content: str

class ChatRequest(BaseModel):
    session_id: str
    message: str
    conversation_history: List[ChatMessage] = []

class UploadResponse(BaseModel):
    success: bool
    session_id: str
    file_id: str
    filename: str
    file_type: str
    preview_available: bool
    extracted_text: Optional[str] = None

# ============================================================================
# Helper Functions
# ============================================================================

def get_anthropic_client():
    """Get Anthropic client, raising error if not configured."""
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="Anthropic API not configured")
    return Anthropic(api_key=ANTHROPIC_API_KEY)


def get_session(session_id: str) -> Dict[str, Any]:
    """Get or create a session."""
    if session_id not in _sessions:
        _sessions[session_id] = {
            "id": session_id,
            "created_at": datetime.utcnow().isoformat(),
            "files": [],  # Uploaded files with their content
            "conversation": [],
            "generated_docs": {
                "underwrite_model": None,
                "business_plan": None,
                "pitch_deck": None
            }
        }
    return _sessions[session_id]


def _extract_pdf_text_fast(file_bytes: bytes) -> str:
    """Cheap, local (no LLM call) text extraction for re-attaching a deal's
    already-uploaded documents (OM/T12/rent roll/etc.) to a chat session so
    they can be referenced again later (e.g. business plan generation)
    without asking the user to re-upload them. Good enough for real-text-
    layer PDFs; scanned/image-only pages just won't extract much, which is
    an acceptable tradeoff for avoiding an extra Claude vision call on every
    single deal load."""
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(file_bytes))
        pages = []
        for page in reader.pages[:40]:  # cap runaway page counts
            try:
                pages.append(page.extract_text() or "")
            except Exception:
                continue
        return "\n\n".join(pages).strip()
    except Exception as e:
        log.warning(f"[Claude Chat] Fast PDF text extraction failed: {e}")
        return ""


def _attach_deal_vault_documents(session: Dict[str, Any], deal_id: str) -> int:
    """Pulls this deal's already-saved documents (Deal Room 'Documents' tab /
    deal_documents table — the OM, T12, rent roll, etc. the user uploaded
    while underwriting) into the chat session's file context, so generators
    like the business plan can actually reference/cite the real source
    documents instead of only the pre-extracted JSON fields. Skips files
    already attached (by filename) and caps how many/how large to keep
    this fast. Returns the number of documents attached."""
    if not deal_id:
        return 0
    try:
        from token_manager import get_supabase
        sb = get_supabase()
        res = (
            sb.table("deal_documents")
            .select("file_name, file_type, public_url, storage_path, file_size")
            .eq("deal_id", deal_id)
            .order("uploaded_at", desc=True)
            .limit(8)
            .execute()
        )
        rows = res.data or []
    except Exception as e:
        log.warning(f"[Claude Chat] Could not look up deal_documents for {deal_id}: {e}")
        return 0

    already = {f["filename"] for f in session["files"]}
    attached = 0
    for row in rows:
        filename = row.get("file_name") or "document"
        if filename in already:
            continue
        url = row.get("public_url")
        file_type = row.get("file_type") or ""
        if not url or (row.get("file_size") or 0) > 25_000_000:
            continue
        if "pdf" not in file_type.lower() and not filename.lower().endswith(".pdf"):
            continue  # text extraction below only handles PDFs for now
        try:
            resp = requests.get(url, timeout=15)
            resp.raise_for_status()
            file_bytes = resp.content
        except Exception as e:
            log.warning(f"[Claude Chat] Failed to download deal document {filename}: {e}")
            continue

        extracted_text = _extract_pdf_text_fast(file_bytes)
        session["files"].append({
            "file_id": str(uuid.uuid4())[:8],
            "filename": filename,
            "file_type": "application/pdf",
            "size": len(file_bytes),
            "uploaded_at": datetime.utcnow().isoformat(),
            "extracted_text": extracted_text,
            "parsed_data": None,
            "base64_data": base64.standard_b64encode(file_bytes).decode("utf-8"),
            "source": "deal_vault",
        })
        attached += 1
    return attached


async def stream_claude_response(
    messages: List[Dict[str, Any]],
    system: str,
    files_context: List[Dict[str, Any]] = None
) -> AsyncGenerator[str, None]:
    """Stream response from Claude using SSE format."""
    client = get_anthropic_client()
    
    # Build the first user message with file context if files are uploaded
    if files_context and messages:
        # Prepend file context to the conversation
        file_summary = "\n\n---\n**UPLOADED DOCUMENTS:**\n"
        for f in files_context:
            file_summary += f"\n**{f['filename']}** ({f['file_type']})\n"
            if f.get('extracted_text'):
                # Truncate very long documents
                text = f['extracted_text'][:50000] if len(f.get('extracted_text', '')) > 50000 else f.get('extracted_text', '')
                file_summary += f"```\n{text}\n```\n"
        file_summary += "\n---\n"
        
        # Add to system prompt
        system = system + file_summary

    try:
        with client.messages.stream(
            model=ANTHROPIC_MODEL,
            max_tokens=8192,
            system=system,
            messages=messages
        ) as stream:
            for text in stream.text_stream:
                # SSE format
                yield f"data: {json.dumps({'type': 'text', 'content': text})}\n\n"

        # Signal completion
        yield f"data: {json.dumps({'type': 'done'})}\n\n"

    except Exception as e:
        log.exception(f"[Claude Chat] Streaming error: {e}")
        yield f"data: {json.dumps({'type': 'error', 'content': str(e)})}\n\n"


async def parse_document_with_claude(file_bytes: bytes, file_type: str, filename: str) -> Dict[str, Any]:
    """Parse uploaded document using Claude's native PDF/vision capabilities."""
    client = get_anthropic_client()
    
    # Determine media type
    if file_type == "application/pdf":
        media_type = "application/pdf"
    elif file_type in ["image/png", "image/jpeg", "image/jpg", "image/webp"]:
        media_type = file_type
    else:
        # For other file types, try to read as text
        try:
            text_content = file_bytes.decode('utf-8')
            return {"extracted_text": text_content, "parsed_data": None}
        except:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_type}")
    
    # Encode file
    file_b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
    
    # Build message content
    if media_type == "application/pdf":
        content = [
            {
                "type": "document",
                "source": {
                    "type": "base64",
                    "media_type": "application/pdf",
                    "data": file_b64
                }
            },
            {
                "type": "text",
                "text": """Extract ALL text and data from this document. If it's a real estate document (OM, rent roll, T12, etc.), also identify and structure the key financial data.

Return your response in this format:
1. First, provide the full extracted text content
2. If applicable, provide a JSON summary of key data points at the end

Be thorough - extract every number, every line item, every detail visible in the document."""
            }
        ]
    else:
        # Image
        content = [
            {
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": media_type,
                    "data": file_b64
                }
            },
            {
                "type": "text",
                "text": "Extract all text and data from this image. If it contains financial or property information, structure the key data points."
            }
        ]
    
    log.info(f"[Claude Chat] Parsing {filename} ({file_type})...")
    
    response = client.messages.create(
        model=ANTHROPIC_MODEL,
        max_tokens=16000,
        messages=[{"role": "user", "content": content}]
    )
    
    extracted_text = response.content[0].text
    
    log.info(f"[Claude Chat] Parsed {filename}: {len(extracted_text)} chars extracted")
    return {"extracted_text": extracted_text, "parsed_data": None}


# ============================================================================
# API Endpoints
# ============================================================================

@router.post("/session")
async def create_session():
    """Create a new chat session."""
    session_id = str(uuid.uuid4())[:8]
    session = get_session(session_id)
    return JSONResponse(content={
        "success": True,
        "session_id": session_id,
        "created_at": session["created_at"]
    })


@router.post("/upload")
async def upload_document(
    file: UploadFile = File(...),
    session_id: str = Form(...)
):
    """
    Upload a document (PDF, image, or text file) to the session.
    The document will be parsed and made available for chat.
    """
    try:
        session = get_session(session_id)
        
        # Read file
        file_bytes = await file.read()
        file_type = file.content_type or "application/octet-stream"
        filename = file.filename
        
        log.info(f"[Claude Chat] Upload: {filename} ({len(file_bytes)} bytes, {file_type})")
        
        # Generate file ID
        file_id = str(uuid.uuid4())[:8]
        
        # Parse document
        parsed = await parse_document_with_claude(file_bytes, file_type, filename)
        
        # Store file info in session
        file_record = {
            "file_id": file_id,
            "filename": filename,
            "file_type": file_type,
            "size": len(file_bytes),
            "uploaded_at": datetime.utcnow().isoformat(),
            "extracted_text": parsed.get("extracted_text"),
            "parsed_data": parsed.get("parsed_data"),
            # Store base64 for PDF preview in canvas
            "base64_data": base64.standard_b64encode(file_bytes).decode("utf-8") if file_type == "application/pdf" else None
        }
        session["files"].append(file_record)
        
        return JSONResponse(content={
            "success": True,
            "session_id": session_id,
            "file_id": file_id,
            "filename": filename,
            "file_type": file_type,
            "preview_available": file_type == "application/pdf",
            "extracted_text": parsed.get("extracted_text", "")[:2000] + "..." if len(parsed.get("extracted_text", "")) > 2000 else parsed.get("extracted_text", "")
        })
        
    except HTTPException:
        raise
    except Exception as e:
        log.exception(f"[Claude Chat] Upload error: {e}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "detail": str(e)}
        )


@router.get("/file/{session_id}/{file_id}")
async def get_file(session_id: str, file_id: str):
    """Get file data for preview in canvas."""
    session = get_session(session_id)
    
    for f in session["files"]:
        if f["file_id"] == file_id:
            return JSONResponse(content={
                "success": True,
                "file_id": file_id,
                "filename": f["filename"],
                "file_type": f["file_type"],
                "base64_data": f.get("base64_data"),
                "extracted_text": f.get("extracted_text")
            })
    
    raise HTTPException(status_code=404, detail="File not found")


@router.get("/session/{session_id}/files")
async def list_session_files(session_id: str):
    """List all files uploaded to a session."""
    session = get_session(session_id)
    
    files = [{
        "file_id": f["file_id"],
        "filename": f["filename"],
        "file_type": f["file_type"],
        "size": f["size"],
        "uploaded_at": f["uploaded_at"],
        "preview_available": f.get("base64_data") is not None
    } for f in session["files"]]
    
    return JSONResponse(content={"success": True, "files": files})


@router.post("/chat/stream")
async def chat_stream(request: Request):
    """
    Stream chat response from Claude.
    Uses Server-Sent Events (SSE) for real-time streaming.
    """
    try:
        data = await request.json()
        session_id = data.get("session_id")
        message = data.get("message", "")
        conversation_history = data.get("conversation_history", [])
        
        if not session_id:
            raise HTTPException(status_code=400, detail="session_id required")
        if not message:
            raise HTTPException(status_code=400, detail="message required")
        
        session = get_session(session_id)
        
        # Build messages for Claude
        messages = []
        for msg in conversation_history[-20:]:  # Last 20 messages for context
            messages.append({
                "role": msg["role"],
                "content": msg["content"]
            })
        
        # Add current message
        messages.append({
            "role": "user",
            "content": message
        })
        
        # Get file context from session
        files_context = session.get("files", [])
        
        # Build system prompt — inject deal context if present
        system = SYSTEM_PROMPT
        deal_context = session.get("deal_context", "")
        if deal_context:
            system += f"\n\n{'='*60}\nACTIVE DEAL CONTEXT (loaded from pipeline):\n{'='*60}\n{deal_context}"
        
        log.info(f"[Claude Chat] Streaming response for session {session_id}, {len(files_context)} files in context")
        
        return StreamingResponse(
            stream_claude_response(messages, system, files_context),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        log.exception(f"[Claude Chat] Chat error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/chat")
async def chat_non_streaming(request: Request):
    """
    Non-streaming chat endpoint (fallback for environments that don't support SSE).
    """
    try:
        data = await request.json()
        session_id = data.get("session_id")
        message = data.get("message", "")
        conversation_history = data.get("conversation_history", [])
        
        if not session_id:
            raise HTTPException(status_code=400, detail="session_id required")
        if not message:
            raise HTTPException(status_code=400, detail="message required")
        
        session = get_session(session_id)
        client = get_anthropic_client()
        
        # Build messages for Claude
        messages = []
        for msg in conversation_history[-20:]:
            messages.append({
                "role": msg["role"],
                "content": msg["content"]
            })
        messages.append({
            "role": "user",
            "content": message
        })
        
        # Build system with file context
        system = SYSTEM_PROMPT
        files_context = session.get("files", [])
        if files_context:
            file_summary = "\n\n---\n**UPLOADED DOCUMENTS:**\n"
            for f in files_context:
                file_summary += f"\n**{f['filename']}** ({f['file_type']})\n"
                if f.get('extracted_text'):
                    text = f['extracted_text'][:50000] if len(f.get('extracted_text', '')) > 50000 else f.get('extracted_text', '')
                    file_summary += f"```\n{text}\n```\n"
            file_summary += "\n---\n"
            system = system + file_summary
        
        response = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=8192,
            system=system,
            messages=messages
        )
        
        assistant_message = response.content[0].text
        
        return JSONResponse(content={
            "success": True,
            "response": assistant_message
        })
        
    except HTTPException:
        raise
    except Exception as e:
        log.exception(f"[Claude Chat] Non-streaming chat error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


class DraftRequest(BaseModel):
    doc_type: str  # "email" or "loi"
    topic: Optional[str] = None
    deal_name: Optional[str] = ""
    address: Optional[str] = ""
    units: Optional[int] = None
    findings: Optional[List[Dict[str, Any]]] = None
    extra_context: Optional[str] = ""


@router.post("/draft")
async def draft_document(payload: DraftRequest):
    """
    One-shot Claude draft for broker emails / LOI cover notes.
    Stateless — no chat session required.
    """
    client = get_anthropic_client()

    findings_text = ""
    if payload.findings:
        findings_text = "\n".join(
            f"- [{(f.get('severity') or 'note').upper()}] {f.get('label', '')}: {f.get('detail', '')}"
            for f in payload.findings
        )

    if payload.doc_type == "loi":
        system = (
            "You are a commercial real estate acquisitions professional drafting a "
            "concise, professional Letter of Intent (LOI) cover email to a listing broker. "
            "Be direct and businesslike. Do not invent numbers that were not provided — "
            "reference them generically (e.g. 'per our proposed terms') if missing."
        )
        user_msg = (
            f"Draft a short LOI submission cover email for this deal.\n\n"
            f"Property: {payload.deal_name or 'the property'}\n"
            f"Address: {payload.address or 'N/A'}\n"
            f"Units: {payload.units if payload.units is not None else 'N/A'}\n"
            f"{payload.extra_context or ''}\n\n"
            "Keep it under 200 words. Sign off as 'the Buyer'."
        )
    else:
        topic_line = f"Focus specifically on: {payload.topic}." if payload.topic else "Cover all findings below."
        system = (
            "You are a commercial real estate analyst drafting a concise, professional email "
            "to a listing broker summarizing due-diligence findings from a rent roll review. "
            "Be factual and specific, avoid alarmist language, and do not invent findings "
            "that were not provided."
        )
        user_msg = (
            f"Draft a brief email to the listing broker about the rent roll for "
            f"{payload.deal_name or 'this property'} ({payload.address or 'N/A'}).\n"
            f"{topic_line}\n\n"
            f"Findings:\n{findings_text or 'No specific findings provided — ask general clarifying questions about the rent roll.'}\n\n"
            "Keep it under 180 words, professional tone, end with a request for clarification/documentation."
        )

    try:
        response = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=1024,
            system=system,
            messages=[{"role": "user", "content": user_msg}],
        )
        draft_text = response.content[0].text
        return JSONResponse({"success": True, "draft": draft_text})
    except Exception as e:
        log.exception(f"[Claude Draft] error: {e}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=200)


@router.delete("/session/{session_id}/file/{file_id}")
async def delete_file(session_id: str, file_id: str):
    """Remove a file from a session."""
    session = get_session(session_id)
    
    session["files"] = [f for f in session["files"] if f["file_id"] != file_id]
    
    return JSONResponse(content={"success": True})


@router.delete("/session/{session_id}")
async def delete_session(session_id: str):
    """Delete a session and all its data."""
    if session_id in _sessions:
        del _sessions[session_id]
    
    return JSONResponse(content={"success": True})


# ============================================================================
# Artifact Execution - Generate Downloadable Files
# ============================================================================

def generate_spreadsheet_from_artifact(artifact_data: Dict[str, Any], title: str) -> bytes:
    """
    Generate an Excel file from artifact JSON data.
    Claude provides the structure, we just convert to xlsx.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    sheets = artifact_data.get("sheets", [])
    if not sheets:
        raise HTTPException(status_code=400, detail="No sheets in artifact data")
    
    for sheet_data in sheets:
        sheet_name = sheet_data.get("name", "Sheet")[:31]  # Excel limit
        ws = wb.create_sheet(title=sheet_name)
        
        # Write data
        data = sheet_data.get("data", [])
        for row_idx, row in enumerate(data, start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Handle formulas
                if isinstance(cell_value, str) and cell_value.startswith("="):
                    cell.value = cell_value
                else:
                    cell.value = cell_value
        
        # Apply column widths
        col_widths = sheet_data.get("columnWidths", {})
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Apply formatting
        formatting = sheet_data.get("formatting", {})
        for cell_range, fmt in formatting.items():
            try:
                for row in ws[cell_range]:
                    for cell in (row if hasattr(row, '__iter__') else [row]):
                        if fmt.get("bold"):
                            cell.font = Font(bold=True)
                        if fmt.get("background"):
                            cell.fill = PatternFill(start_color=fmt["background"].replace("#", ""), 
                                                   end_color=fmt["background"].replace("#", ""), 
                                                   fill_type="solid")
                        if fmt.get("numberFormat"):
                            cell.number_format = fmt["numberFormat"]
            except Exception as e:
                log.warning(f"[Artifact] Failed to apply formatting {cell_range}: {e}")
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_document_from_artifact(markdown_content: str, title: str) -> bytes:
    """
    Generate a PDF from markdown content.
    Uses markdown2 + weasyprint if available, otherwise returns HTML.
    """
    try:
        import markdown2
        html_content = markdown2.markdown(
            markdown_content, 
            extras=["tables", "fenced-code-blocks", "header-ids"]
        )
    except ImportError:
        # Fallback: basic markdown conversion
        html_content = f"<pre>{markdown_content}</pre>"
    
    # Wrap in styled HTML — professional CRE business plan format matching Appleby style
    full_html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800;900&display=swap');

        * {{ box-sizing: border-box; margin: 0; padding: 0; }}

        body {{
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #ffffff;
            color: #1a1a2e;
            font-size: 11pt;
            line-height: 1.6;
        }}

        /* Cover / first section */
        .cover {{
            background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 60%, #0d4429 100%);
            color: white;
            padding: 60px 50px 50px;
            min-height: 220px;
        }}
        .cover h1 {{
            font-size: 32pt;
            font-weight: 900;
            letter-spacing: -1px;
            margin-bottom: 8px;
            color: #ffffff;
        }}
        .cover .subtitle {{
            font-size: 12pt;
            color: #94a3b8;
            margin-bottom: 6px;
        }}
        .cover .tagline {{
            font-size: 10pt;
            color: #64748b;
            font-style: italic;
        }}

        .content {{
            padding: 40px 50px;
            max-width: 900px;
            margin: 0 auto;
        }}

        /* Section headers */
        h1 {{
            font-size: 20pt;
            font-weight: 800;
            color: #0f172a;
            border-bottom: 3px solid #10b981;
            padding-bottom: 10px;
            margin: 40px 0 20px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        h2 {{
            font-size: 14pt;
            font-weight: 700;
            color: #1e3a5f;
            margin: 28px 0 12px;
            padding-left: 12px;
            border-left: 4px solid #10b981;
        }}

        h3 {{
            font-size: 12pt;
            font-weight: 600;
            color: #374151;
            margin: 20px 0 10px;
            text-transform: uppercase;
            letter-spacing: 0.3px;
            font-size: 10pt;
        }}

        p {{
            margin: 0 0 12px;
            color: #374151;
        }}

        /* Tables — main styling */
        table {{
            border-collapse: collapse;
            width: 100%;
            margin: 16px 0 24px;
            font-size: 10pt;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 1px 4px rgba(0,0,0,0.08);
        }}

        thead tr {{
            background: #0f172a;
            color: #ffffff;
        }}

        thead th {{
            padding: 10px 14px;
            text-align: left;
            font-weight: 700;
            font-size: 9pt;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            border: none;
        }}

        tbody tr:nth-child(even) {{
            background: #f8fafc;
        }}

        tbody tr:nth-child(odd) {{
            background: #ffffff;
        }}

        tbody tr:last-child {{
            background: #e8f5e9;
            font-weight: 700;
            border-top: 2px solid #10b981;
        }}

        td {{
            padding: 9px 14px;
            border-bottom: 1px solid #e5e7eb;
            color: #1f2937;
        }}

        th {{
            padding: 10px 14px;
            text-align: left;
            font-weight: 700;
        }}

        /* Highlight boxes */
        .highlight-box {{
            background: linear-gradient(135deg, #f0fdf4, #ecfdf5);
            border: 1px solid #86efac;
            border-left: 4px solid #10b981;
            border-radius: 8px;
            padding: 16px 20px;
            margin: 16px 0;
        }}

        .highlight-box strong {{
            color: #065f46;
        }}

        /* Metric cards inline */
        .metrics-row {{
            display: flex;
            gap: 16px;
            margin: 20px 0;
            flex-wrap: wrap;
        }}

        .metric-card {{
            background: #f8fafc;
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            padding: 14px 18px;
            flex: 1;
            min-width: 140px;
        }}

        .metric-card .label {{
            font-size: 9pt;
            color: #6b7280;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.4px;
        }}

        .metric-card .value {{
            font-size: 16pt;
            font-weight: 800;
            color: #0f172a;
            margin-top: 4px;
        }}

        /* Scenario sections */
        .scenario-header {{
            background: #1e3a5f;
            color: white;
            padding: 12px 18px;
            border-radius: 8px 8px 0 0;
            font-weight: 700;
            font-size: 11pt;
            margin-top: 28px;
        }}

        .scenario-body {{
            border: 1px solid #e5e7eb;
            border-top: none;
            border-radius: 0 0 8px 8px;
            padding: 0;
            overflow: hidden;
        }}

        /* NOI waterfall */
        .waterfall-row {{
            display: flex;
            align-items: center;
            padding: 10px 16px;
            border-bottom: 1px solid #f3f4f6;
        }}

        .waterfall-row.total {{
            background: #e8f5e9;
            font-weight: 700;
            font-size: 11pt;
        }}

        .waterfall-row .arrow {{
            color: #10b981;
            font-weight: 700;
            margin: 0 8px;
        }}

        /* Lists */
        ul, ol {{
            padding-left: 20px;
            margin: 8px 0 12px;
        }}

        li {{
            margin-bottom: 6px;
            color: #374151;
        }}

        /* Blockquote / callout */
        blockquote {{
            background: #eff6ff;
            border-left: 4px solid #2563eb;
            padding: 12px 16px;
            margin: 16px 0;
            border-radius: 0 6px 6px 0;
            color: #1e40af;
            font-style: italic;
        }}

        /* Recommendation box */
        .recommendation {{
            background: linear-gradient(135deg, #0f172a, #1e3a5f);
            color: white;
            padding: 24px 28px;
            border-radius: 12px;
            margin: 24px 0;
        }}

        .recommendation h3 {{
            color: #10b981;
            font-size: 12pt;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-bottom: 12px;
        }}

        .recommendation p {{
            color: #cbd5e1;
            font-size: 11pt;
        }}

        /* Confidentiality footer */
        .confidentiality {{
            margin-top: 48px;
            padding: 16px;
            border-top: 1px solid #e5e7eb;
            font-size: 9pt;
            color: #9ca3af;
            text-align: center;
            font-style: italic;
        }}

        /* Page break hints */
        .page-break {{
            page-break-before: always;
        }}

        /* Positive/negative value colors */
        .positive {{ color: #065f46; font-weight: 600; }}
        .negative {{ color: #b91c1c; font-weight: 600; }}

        code {{
            background: #f4f4f4;
            padding: 2px 6px;
            border-radius: 3px;
            font-family: 'Monaco', 'Consolas', monospace;
            font-size: 9pt;
        }}

        strong {{ font-weight: 700; }}

        hr {{
            border: none;
            border-top: 1px solid #e5e7eb;
            margin: 24px 0;
        }}
    </style>
</head>
<body>
<div class="content">
{html_content}
</div>
<div class="confidentiality">
    This document is confidential and prepared for private investor review only. 
    All projections are estimates based on available data and subject to due diligence verification. 
    Past performance does not guarantee future results. Generated by DealSniper AI Underwriter.
</div>
</body>
</html>"""
    
    # Try to generate PDF with weasyprint
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=full_html).write_pdf()
        return pdf_bytes
    except ImportError:
        log.warning("[Artifact] weasyprint not available, returning HTML instead")
        return full_html.encode('utf-8')
    except Exception as e:
        log.warning(f"[Artifact] PDF generation failed: {e}, returning HTML")
        return full_html.encode('utf-8')


# ============================================================================
# Deal Context Injection - Load parsed deal data into the chat session
# ============================================================================

def _strip_wrapping_code_fence(text: str) -> str:
    """
    Claude is instructed to wrap the whole business plan in a single
    ```artifact:document:<title>``` fence, but it doesn't always follow that
    exactly — sometimes it wraps the entire response in a plain/generic
    ``` fence instead (with no 'artifact:document:' tag), and/or appends a
    trailing sentence after the closing fence (e.g. "Let me know if you'd
    like any changes!"). That trailing text breaks a naive "does the string
    end with ```" check, so the fence never gets stripped and the whole
    document renders client-side as ONE giant literal code block (raw '#',
    '**', '|' characters visible instead of real headers/bold/tables)
    instead of being parsed as markdown. This scans for the LAST line that
    is just a bare ``` fence (ignoring anything after it) instead of
    requiring the string to end exactly on the fence, so it's resilient to
    trailing commentary. Safe/idempotent to call even when there's no fence
    to strip.
    """
    t = (text or "").strip()
    for _ in range(2):  # handle at most one extra level of accidental double-wrapping
        if not t.startswith("```"):
            break
        lines = t.split("\n")
        close_idx = -1
        for i in range(len(lines) - 1, 0, -1):
            if re.match(r"^```\s*$", lines[i]):
                close_idx = i
                break
        if close_idx == -1:
            break  # no closing fence line found — leave as-is
        t = "\n".join(lines[1:close_idx]).strip()
    return t


BUSINESS_PLAN_PROMPT = """You are generating a professional Investment Underwriting & Business Plan document for a real estate deal. 

Based on the deal data provided in this session, generate a complete business plan document in the following structure. 
This must be formatted as a document artifact that can be downloaded as a PDF.

MANDATORY OUTPUT FORMAT — DO NOT SKIP THIS: your ENTIRE response must be wrapped in a single
```artifact:document:<Property Name> — Business Plan``` code fence (exactly like the DOCUMENT
ARTIFACT FORMAT described earlier in your instructions), with nothing before the opening fence and
nothing after the closing fence. This is a long, multi-section document — do not let that cause you
to drop the artifact wrapper; the whole document (every section below, start to finish) goes inside
that one fence. If you forget this, the plan will NOT be downloadable and will just dump as raw text
in the chat, which defeats the entire point of generating it.

CRITICAL — USE THE USER'S ACTUAL UNDERWRITING STRATEGY, DO NOT INVENT ONE:
The "SCENARIO / ASSUMPTIONS" data below is the EXACT financing and hold/exit strategy the user
already configured on the Results page of the underwriting model (interest rate/rate override,
LTV, amortization, IO period, refinance year/LTV/rate if refi is enabled, exit cap rate, hold
period, cost of sale %, rent growth assumption, and — if present — JV/waterfall structure such as
preferred return rate, GP promote %, and equity split). This plan must be built AROUND that real,
user-configured strategy — not a set of generic invented alternatives. If refinance is enabled in
the scenario data, the plan's exit/hold strategy is the refi described there (at that specific
year/LTV/rate), not a sale, and vice versa. If a JV/waterfall structure is present, the investor
returns section must reflect that exact preferred return and promote structure — do not substitute
a different equity structure. Only if the scenario data is genuinely sparse/missing for a given
assumption should you note an explicit assumption and flag it as such — never silently replace a
real user-configured number with a fabricated one.

If the user has any documents attached in this session (OMs, T12s, rent rolls, etc. — provided in
the UPLOADED DOCUMENTS context), pull real figures from them where the parsed data is incomplete or
ambiguous, and cite which document/section a figure came from when it materially differs from the
headline extracted numbers (e.g. "per the T-12, actual property taxes were $X, vs. the pro forma
figure of $Y").

The document must include ALL of the following sections — do not skip any:

1. **OFFERING HIGHLIGHTS** — property name/address, year built/renovated, total SF, asking price, T12 actual NOI, going-in cap rate, day-1 equity, occupancy, exit strategy

2. **SECTION 1: PROPERTY OVERVIEW** — property details table (address, type, unit mix, buildings, acreage, parking, occupancy, management, utilities), unit mix & rent roll table (unit type, count, avg rent, monthly, annual totals)

3. **SECTION 2: MARKET ANALYSIS** — why this market, market fundamentals table (metro area, population, median income, income growth, rent growth, vacancy, anchor institutions, major employers), rent comparable summary table (competing properties with units, SF, market rent, occupancy), cap rate context table (major metro vs suburban vs tertiary vs subject market)

4. **SECTION 3: BUSINESS PLAN** — overview of strategy, utility audit table (who pays what, annual cost, action), Phase 1 (RUBS or first value-add initiative with calculations and NOI impact), Phase 2 (rent increases or second initiative), Phase 3 (market rent upside), NOI Growth Waterfall table (baseline → each phase → exit value)

5. **SECTION 4: UNDERWRITING — T12 INCOME & EXPENSES** — full income statement (GPR, vacancy, other income, EGI) and expense breakdown (taxes, insurance, utilities, repairs, management, all line items) with NOI

6. **SECTION 5: THE USER'S DEAL STRUCTURE & STRATEGY** — present the ACTUAL financing/exit strategy from the scenario data described above as a single, clearly-labeled deal structure (not multiple invented alternatives): deal structure table (purchase price, down payment, closing costs, total equity, loan amount, rate, amortization, IO period, debt service), cash flow analysis (day-1/year1/stabilized NOI, debt service, cash flow, DSCR), exit strategy exactly as configured (refi at the user's refi year/LTV/rate, OR sale at the user's exit cap rate/hold period — whichever the scenario data specifies), investor returns using the user's actual preferred return/promote/equity split if a JV or waterfall structure is present, key metrics (going-in cap rate, market cap rate, day-1 equity, DSCR, total return, IRR/equity multiple if computable). If (and only if) the user explicitly asked for alternate scenarios to be compared, you may add a second labeled scenario — otherwise present ONE plan matching what they actually built.

7. **SECTION 6: EXECUTION & STABILIZATION TIMELINE** — a phased timeline table (Phase | Timeline | Key Actions | Target/KPI) covering: Month 1 (closing, transition, initial due diligence follow-ups), Month 2-4 (lease-up / initial value-add rollout), Month 3-6 (utility billback/RUBS implementation if applicable), Month 6-12 (stabilization), Month 12-18 (optimization / rent pushes), Month 18 through the user's hold period (steady-state hold), and the Exit/Refi window matching the user's actual configured exit year. Follow this with a First-90-Days action checklist broken into Month 1 / Month 2 / Month 3 concrete action items.

8. **SECTION 7: DEAL STRUCTURE SUMMARY** — a clean one-page recap table of the strategy from Section 5 (not a multi-scenario comparison, since this plan reflects the one real strategy the user configured) plus a short recommendation/conclusion paragraph.

Use the actual numbers from the deal data. Where data is missing, make reasonable assumptions and note them explicitly as assumptions.
Format all currency with $ and commas. Format all percentages with %.
Be thorough — this is a professional investor presentation document.

INTERNAL CONSISTENCY — DO NOT CONTRADICT YOURSELF: the "Day-1 Cash Flow" figure shown in
OFFERING HIGHLIGHTS at the top of the document MUST be the exact same number as the "Year 1 Cash
Flow" you compute in SECTION 5's cash flow analysis (including any adjustments/assumptions you
apply there, e.g. a property tax reassessment). Compute Section 5 first if needed, then use that
same final number at the top — never show two different "day-1" cash flow figures in the same
document.

REMINDER: wrap the entire document above in a single ```artifact:document:<title>``` fence as your
whole response — open the fence first, write all 8 sections inside it, then close the fence last."""


@router.post("/inject-deal-context")
async def inject_deal_context(request: Request):
    """
    Inject parsed deal data from the pipeline into a chat session.
    This gives Claude full context so the user can just say 'make me a business plan'.
    
    Body:
    {
        "session_id": "abc123",
        "deal_data": {
            "address": "...",
            "units": 17,
            "purchase_price": 1250000,
            "parsed_data": { ... },
            "scenario_data": { ... },
            "notes": "..."
        }
    }
    """
    try:
        body = await request.json()
        session_id = body.get("session_id")
        deal_data = body.get("deal_data", {})
        
        if not session_id:
            raise HTTPException(status_code=400, detail="session_id required")
        
        session = get_session(session_id)
        
        # Build a comprehensive deal context string
        parsed = deal_data.get("parsed_data") or {}
        scenario = deal_data.get("scenario_data") or {}
        
        deal_summary = f"""
DEAL CONTEXT — LOADED FROM PIPELINE
=====================================
Address: {deal_data.get('address', parsed.get('property', {}).get('address', 'N/A'))}
Units: {deal_data.get('units', parsed.get('property', {}).get('units', 'N/A'))}
Purchase Price: ${deal_data.get('purchase_price', parsed.get('pricing_financing', {}).get('price', 0)):,.0f}
Deal Structure: {deal_data.get('deal_structure', 'N/A')}
Notes: {deal_data.get('notes', 'None')}

FULL PARSED DATA:
{json.dumps(parsed, indent=2)[:30000]}

SCENARIO / ASSUMPTIONS (the user's actual configured underwriting strategy — financing terms, exit/refi assumptions, JV/waterfall structure if any):
{json.dumps(scenario, indent=2)[:8000]}
"""
        
        # Pull this deal's already-uploaded documents (OM/T12/rent roll/etc.
        # saved to the Deal Room's Document Vault) into the session so they
        # can actually be referenced/cited later (business plan generation,
        # follow-up chat questions) without asking the user to re-upload.
        deal_id = deal_data.get("deal_id")
        attached_count = _attach_deal_vault_documents(session, deal_id) if deal_id else 0
        if attached_count:
            deal_summary += f"\n{attached_count} saved document(s) from this deal's Document Vault (OM/T12/rent roll/etc.) are attached below under UPLOADED DOCUMENTS — reference them directly when relevant.\n"

        # Store as deal context in session
        session["deal_context"] = deal_summary
        session["deal_data"] = deal_data
        session["deal_address"] = deal_data.get("address", "Deal")
        
        log.info(f"[Claude Chat] Injected deal context into session {session_id}: {deal_data.get('address', 'unknown')} ({attached_count} vault documents attached)")
        
        return JSONResponse(content={
            "success": True,
            "session_id": session_id,
            "deal_address": session["deal_address"],
            "context_length": len(deal_summary),
            "documents_attached": attached_count,
        })
        
    except HTTPException:
        raise
    except Exception as e:
        log.exception(f"[Claude Chat] Deal context injection error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/business-plan/generate")
async def generate_business_plan(request: Request):
    """
    Trigger a business plan generation for the current session's deal.
    Streams the response exactly like /chat/stream.
    """
    try:
        data = await request.json()
        session_id = data.get("session_id")
        user_instructions = data.get("instructions", "")
        
        if not session_id:
            raise HTTPException(status_code=400, detail="session_id required")
        
        session = get_session(session_id)
        deal_context = session.get("deal_context", "")
        
        if not deal_context and not session.get("files"):
            raise HTTPException(
                status_code=400,
                detail="No deal data in session. Load a deal from the pipeline or upload documents first."
            )
        
        # Build the business plan trigger message
        user_message = BUSINESS_PLAN_PROMPT
        if user_instructions:
            user_message += f"\n\nADDITIONAL INSTRUCTIONS FROM USER:\n{user_instructions}"
        
        # Include conversation history
        conversation_history = data.get("conversation_history", [])
        messages = []
        for msg in conversation_history[-10:]:
            messages.append({"role": msg["role"], "content": msg["content"]})
        messages.append({"role": "user", "content": user_message})
        
        # Build system with deal context
        system = SYSTEM_PROMPT
        if deal_context:
            system += f"\n\n{'='*60}\nACTIVE DEAL CONTEXT:\n{'='*60}\n{deal_context}"
        
        files_context = session.get("files", [])
        
        log.info(f"[Claude Chat] Generating business plan for session {session_id}")
        
        return StreamingResponse(
            stream_claude_response(messages, system, files_context),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        log.exception(f"[Claude Chat] Business plan generation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/business-plan/generate-for-deal")
async def generate_business_plan_for_deal(request: Request):
    """
    One-shot (non-streaming) business plan generation tied directly to a
    deal_id — used by the Deal Room's "Generate Business Plan" button so it
    doesn't need to create/manage a chat session at all. Loads the deal
    straight from Supabase, attaches the deal's saved Document Vault files
    (OM/T12/rent roll/etc.), asks Claude for the full plan in one blocking
    call, and returns the finished markdown (already unwrapped from the
    ```artifact:document fence if present) so the frontend can render it to
    PDF and save it back onto the deal's documents.

    Body: { "deal_id": "..." }
    """
    try:
        body = await request.json()
        deal_id = body.get("deal_id")
        if not deal_id:
            raise HTTPException(status_code=400, detail="deal_id required")

        from token_manager import get_supabase
        sb = get_supabase()
        res = sb.table("deals").select("*").eq("deal_id", deal_id).limit(1).execute()
        rows = res.data or []
        if not rows:
            raise HTTPException(status_code=404, detail="Deal not found")
        deal = rows[0]

        parsed = deal.get("parsed_data") or {}
        scenario = deal.get("scenario_data") or {}
        address = deal.get("address") or (parsed.get("property") or {}).get("address") or "Deal"

        # Reuse the same session["files"] shape / helper used by the chat-
        # session flow, just not persisted anywhere — this call is stateless.
        session = {"files": []}
        attached_count = _attach_deal_vault_documents(session, deal_id)

        deal_summary = f"""
DEAL CONTEXT
=====================================
Address: {address}
Units: {deal.get('units', 'N/A')}
Purchase Price: ${(deal.get('purchase_price') or 0):,.0f}

FULL PARSED DATA:
{json.dumps(parsed, indent=2)[:30000]}

SCENARIO / ASSUMPTIONS (the user's actual configured underwriting strategy — financing terms, exit/refi assumptions, JV/waterfall structure if any):
{json.dumps(scenario, indent=2)[:8000]}
"""

        system = SYSTEM_PROMPT + f"\n\n{'='*60}\nACTIVE DEAL CONTEXT:\n{'='*60}\n{deal_summary}"

        files_context = session.get("files", [])
        if files_context:
            file_summary = "\n\n---\n**UPLOADED DOCUMENTS:**\n"
            for f in files_context:
                file_summary += f"\n**{f['filename']}** ({f['file_type']})\n"
                if f.get("extracted_text"):
                    text = f["extracted_text"][:50000]
                    file_summary += f"```\n{text}\n```\n"
            file_summary += "\n---\n"
            system += file_summary

        client = get_anthropic_client()
        log.info(f"[Claude Chat] Generating business plan document for deal {deal_id} ({attached_count} vault documents attached)")
        response = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=8192,
            system=system,
            messages=[{"role": "user", "content": BUSINESS_PLAN_PROMPT}],
        )
        full_text = response.content[0].text

        m = re.search(r"```artifact:document:([^\n]+)\n([\s\S]*?)```", full_text)
        if m:
            title = m.group(1).strip()
            markdown_content = m.group(2).strip()
        else:
            # Didn't get wrapped in the expected fence — fall back to using
            # the raw text rather than failing the whole request.
            title = f"{address} — Business Plan"
            markdown_content = full_text.strip()

        # Defensive cleanup: Claude sometimes wraps the whole response in a
        # plain/generic ``` fence instead of (or in addition to) the expected
        # ```artifact:document: fence. If that leftover fence isn't stripped,
        # the entire document renders client-side as one literal code block
        # instead of parsed markdown/tables. Safe to call even if there's
        # nothing to strip.
        markdown_content = _strip_wrapping_code_fence(markdown_content)

        return JSONResponse(content={
            "success": True,
            "title": title,
            "markdown": markdown_content,
            "documents_attached": attached_count,
        })

    except HTTPException:
        raise
    except Exception as e:
        log.exception(f"[Claude Chat] Business plan (deal) generation error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/artifact/execute")
async def execute_artifact(request: Request):
    """
    Execute an artifact to generate a downloadable file.
    
    Body:
    {
        "type": "spreadsheet" | "document",
        "title": "File Title",
        "data": { ... } | "markdown content"
    }
    """
    try:
        body = await request.json()
        artifact_type = body.get("type")
        title = body.get("title", "Untitled")
        data = body.get("data")
        
        if not artifact_type or not data:
            raise HTTPException(status_code=400, detail="type and data required")
        
        if artifact_type == "spreadsheet":
            # Parse JSON if string
            if isinstance(data, str):
                try:
                    data = json.loads(data)
                except json.JSONDecodeError as e:
                    raise HTTPException(status_code=400, detail=f"Invalid JSON: {e}")
            
            file_bytes = generate_spreadsheet_from_artifact(data, title)
            filename = f"{title.replace(' ', '_')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            
        elif artifact_type == "document":
            # data is markdown string
            if not isinstance(data, str):
                raise HTTPException(status_code=400, detail="Document data must be markdown string")
            
            file_bytes = generate_document_from_artifact(data, title)
            
            # Check if PDF or HTML was generated
            if file_bytes[:4] == b'%PDF':
                filename = f"{title.replace(' ', '_')}.pdf"
                media_type = "application/pdf"
            else:
                filename = f"{title.replace(' ', '_')}.html"
                media_type = "text/html"
        else:
            raise HTTPException(status_code=400, detail=f"Unknown artifact type: {artifact_type}")
        
        # Return as base64 for frontend to handle download
        file_b64 = base64.standard_b64encode(file_bytes).decode('utf-8')
        
        return JSONResponse(content={
            "success": True,
            "filename": filename,
            "media_type": media_type,
            "data": file_b64
        })
        
    except HTTPException:
        raise
    except Exception as e:
        log.exception(f"[Artifact] Execution error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
