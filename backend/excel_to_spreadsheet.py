"""
Simple Excel to x-data-spreadsheet converter
Reads Multifamily underwrite .xlsx and converts to spreadsheet format
"""
import openpyxl
import os

def load_excel_template():
    """
    Load the Excel file and convert to x-data-spreadsheet format
    """
    # Prefer public templates path; fallback to build path
    base_dir = os.path.dirname(__file__)
    public_path = os.path.join(base_dir, "..", "client", "public", "templates", "multifamily.xlsx")
    build_path = os.path.join(base_dir, "..", "client", "build", "multifamilytemplate.xlsx")
    excel_path = public_path if os.path.exists(public_path) else build_path
    
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")
    
    print(f"[EXCEL PARSER] Loading: {excel_path}")
    
    # Load workbook; use data_only=True to read computed values when present
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    rows = {}
    merges = []
    cols = {}
    
    # Use worksheet dimension to avoid EmptyCell issues and get bounds
    dim = ws.calculate_dimension()  # e.g., "A1:K200" or "A1"
    from openpyxl.utils.cell import range_boundaries
    try:
        min_col, min_row, max_col, max_row = range_boundaries(dim)
    except Exception:
        # Fallback: if dimension is invalid, scan a reasonable area
        min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1
    
    def _to_hex(rgb):
        try:
            if not rgb:
                return None
            s = str(rgb)
            # openpyxl uses ARGB like 'FF1F4E79'
            if len(s) == 8:
                return f"#{s[2:]}"  # drop alpha
            if len(s) == 6:
                return f"#{s}"
            return None
        except Exception:
            return None

    # Iterate with Cell objects to capture styles
    for r_idx, row_cells in enumerate(
        ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=False),
        start=min_row - 1  # 0-based index start
    ):
        cells = {}
        has_any = False
        for c_offset, cell in enumerate(row_cells, start=min_col - 1):
            val = cell.value
            if val is None and not cell.has_style:
                continue
            # number format handling
            numfmt = getattr(cell, "number_format", None)
            raw_val = None
            disp = None
            if isinstance(val, (int, float)):
                raw_val = float(val)
                fmt = (numfmt or "General").lower()
                try:
                    if "%" in fmt:
                        # Excel percent shows val*100 with % suffix according to format string decimals
                        decimals = 0
                        try:
                            if "." in fmt:
                                decimals = len(fmt.split(".")[1].split("%")[0])
                        except Exception:
                            pass
                        disp = f"{raw_val*100:.{decimals}f}%"
                    elif "$" in fmt or "[$" in fmt:
                        # Currency: infer decimals from format
                        decimals = 2
                        try:
                            if "." in fmt:
                                decimals = len(fmt.split(".")[1].split(";")[0].replace("#", "").replace("0", ""))
                        except Exception:
                            pass
                        disp = "$" + format(raw_val, f",.{decimals}f")
                    elif "#,##" in fmt or "0.00" in fmt or ".00" in fmt:
                        # Thousands with decimals
                        decimals = 2 if "." in fmt else 0
                        disp = format(raw_val, f",.{decimals}f")
                    else:
                        disp = str(val)
                except Exception:
                    disp = str(val)
            else:
                disp = None
            text = disp if disp is not None else ("" if val is None else str(val))
            # Extract basic styles
            bg = None
            try:
                if getattr(cell.fill, "fill_type", None) == "solid":
                    bg = _to_hex(getattr(getattr(cell.fill, "fgColor", None), "rgb", None))
            except Exception:
                pass
            fc = None
            try:
                fc = _to_hex(getattr(getattr(cell.font, "color", None), "rgb", None))
            except Exception:
                pass
            bold = bool(getattr(cell.font, "bold", False)) if getattr(cell, "font", None) else False
            italic = bool(getattr(cell.font, "italic", False)) if getattr(cell, "font", None) else False
            fs = int(getattr(cell.font, "size", 0)) if getattr(cell, "font", None) else None
            ht = getattr(getattr(cell, "alignment", None), "horizontal", None)
            vt = getattr(getattr(cell, "alignment", None), "vertical", None)

            cell_out = {"text": text}
            if raw_val is not None:
                cell_out["val"] = raw_val
                cell_out["disp"] = text
                cell_out["numfmt"] = numfmt or "General"
            if bg: cell_out["bg"] = bg
            if fc: cell_out["fc"] = fc
            if bold: cell_out["bold"] = True
            if italic: cell_out["italic"] = True
            if fs: cell_out["fs"] = fs
            if ht: cell_out["ht"] = ht
            if vt: cell_out["vt"] = vt

            cells[c_offset] = cell_out
            has_any = True

        if has_any:
            rows[r_idx] = rows.get(r_idx, {})
            rows[r_idx]["cells"] = cells
        
        # Apply row height if defined in Excel
        try:
            rd = ws.row_dimensions.get(r_idx + 1)
            if rd and rd.height:
                rows[r_idx] = rows.get(r_idx, {})
                rows[r_idx]["height"] = int(round(rd.height))
        except Exception:
            pass
    
    # Extract merges from Excel
    try:
        for rng in getattr(ws.merged_cells, "ranges", []):
            try:
                m_min_col, m_min_row, m_max_col, m_max_row = rng.bounds
                # Convert to 0-based and x-data-spreadsheet order col:row:col:row
                merges.append(f"{m_min_col - 1}:{m_min_row - 1}:{m_max_col - 1}:{m_max_row - 1}")
            except Exception:
                continue
    except Exception:
        pass

    # Extract column widths
    try:
        from openpyxl.utils import column_index_from_string
        for letter, dim in ws.column_dimensions.items():
            try:
                idx0 = column_index_from_string(letter) - 1
                if dim and dim.width:
                    # Rough conversion from Excel width (chars) to pixels
                    width_px = int(round(dim.width * 7))
                    cols[idx0] = {"width": max(40, width_px)}
            except Exception:
                continue
    except Exception:
        pass

    wb.close()
    
    print(f"[EXCEL PARSER] Parsed {len(rows)} non-empty rows from Excel")
    
    wb.close()
    
    print(f"[EXCEL PARSER] Loaded {len(rows)} rows from Excel")
    
    # Include column len if known
    cols_out = {"len": max_col - (min_col - 1)} if max_col and min_col else {}
    for c_idx, c_meta in cols.items():
        cols_out[c_idx] = c_meta

    return {
        "name": "Property",
        "rows": rows,
        "cols": cols_out,
        "merges": merges,
        "styles": []
    }
