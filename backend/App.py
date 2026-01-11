

"""
app.py  — Underwriting backend w/ financing modes + comprehensive metrics
Python 3.10+  |  uvicorn app:app --host 127.0.0.1 --port 8010 --reload
"""

import os, io, json, base64, re, uuid, tempfile, shutil
from pathlib import Path
from typing import Optional, Dict, Any, List

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request, status
from fastapi.responses import JSONResponse, RedirectResponse
from cors_config import install_cors

from pypdf import PdfReader, PdfWriter
from dotenv import load_dotenv

from mistralai import Mistral
from anthropic import Anthropic

# from protected_routes import router as protected_router  # Moved to after startup
# Optional advanced parser (parser_v4) for richer underwriting extraction.
try:
    from parser_v4 import RealEstateParser
    _RE_PARSER = RealEstateParser()
    HAS_PARSER_V4 = True
except Exception:
    _RE_PARSER = None
    HAS_PARSER_V4 = False


# Always load the backend .env regardless of the process working directory.
load_dotenv(dotenv_path=Path(__file__).resolve().with_name(".env"), override=True)
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")

# Global LLM client placeholders (set in startup)
MISTRAL = None
ANTHROPIC = None

# Import spreadsheet AI module
from spreadsheet_ai import process_spreadsheet_command
from max_prompts import MAX_PARTNER_SYSTEM_PROMPT

# Allowed document MIME types for uploads and OCR
ALLOWED_DOC_MIMES = {"application/pdf", "image/png", "image/jpeg", "image/jpg", "image/webp"}

# Maximum allowed upload size (bytes)
MAX_BYTES = 50 * 1024 * 1024  # 50 MB



app = FastAPI(title="Underwriting Backend", version="9.0.0")
install_cors(app)

import logging
log = logging.getLogger("app")
logging.basicConfig(level=logging.INFO)

import stripe

# Price IDs: prefer env vars, fallback to known test IDs
PRICE_ID_PRO = os.getenv("PRICE_ID_PRO", "price_1SfA2SRRD0SJQZk3q6Zujrw0")
PRICE_ID_BASE = os.getenv("PRICE_ID_BASE", "price_1SfA11RRD0SJQZk3dTP5HIHa")

# Debug endpoint for env vars (dev only)
@app.get("/debug/env")
def debug_env():
    import os
    return {
        "openai_key_prefix": (os.getenv("OPENAI_API_KEY") or "")[:12],
        "anthropic_key_prefix": (os.getenv("ANTHROPIC_API_KEY") or os.getenv("CLAUDE_API_KEY") or "")[:12],
    }

# V2 Underwriter: Include v2 routes
from v2_underwriter.routes import router as v2_router
app.include_router(v2_router)

# LLM usage logging routes
from v2_underwriter.llm_usage import router as llm_usage_router
app.include_router(llm_usage_router)

# Email Deals: Gmail integration & auto-screening
from email_deals import router as email_deals_router, auth_router as google_auth_router
app.include_router(email_deals_router)
app.include_router(google_auth_router)

# Token Management: AI operation billing
from token_manager import router as token_router
app.include_router(token_router)

# Stripe Webhook: Handle subscription updates
from stripe_webhook_handler import router as stripe_webhook_router
app.include_router(stripe_webhook_router)

# Token Purchase: Handle one-time token purchases
from token_purchase_handler import router as token_purchase_router
app.include_router(token_purchase_router)

# HUD API proxy router (provides /api/hud/* endpoints)
try:
    from hud_api import router as hud_router
    app.include_router(hud_router)
except Exception:
    log.exception("Failed to include HUD API router")

# Excel AI: Spreadsheet assistant with GPT-4
try:
    from excel_ai import router as excel_ai_router
    app.include_router(excel_ai_router)
    log.info("Excel AI router loaded successfully")
except Exception:
    log.exception("Failed to include Excel AI router")


# ---------------- Spreadsheet Command Relay ----------------
# In-memory command queues keyed by sessionId
COMMAND_QUEUES: dict[str, list[dict]] = {}

@app.post("/api/spreadsheet/commands")
async def enqueue_spreadsheet_commands(request: Request):
    """Enqueue one or more spreadsheet commands for a client session.

    Body: { sessionId: string, commands: array|object }
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON body")

    session_id = body.get("sessionId")
    commands = body.get("commands")
    if not session_id:
        raise HTTPException(status_code=400, detail="Missing sessionId")

    if commands is None:
        raise HTTPException(status_code=400, detail="Missing commands")
    if isinstance(commands, dict):
        commands = [commands]
    if not isinstance(commands, list):
        raise HTTPException(status_code=400, detail="Commands must be a list or object")

    q = COMMAND_QUEUES.setdefault(session_id, [])
    q.extend(commands)
    try:
        import logging as _logging
        _logging.getLogger("app").info("[Spreadsheet] Enqueue %d cmd(s) for session %s", len(commands), session_id)
    except Exception:
        pass
    return {"ok": True, "queued": len(commands)}

@app.get("/api/spreadsheet/commands")
async def dequeue_spreadsheet_commands(sessionId: str):
    """Dequeue and return any pending commands for a client session.

    Query: ?sessionId=...
    """
    cmds = COMMAND_QUEUES.pop(sessionId, [])
    try:
        import logging as _logging
        _logging.getLogger("app").info("[Spreadsheet] Dequeue for session %s -> %d cmd(s)", sessionId, len(cmds))
    except Exception:
        pass
    return {"ok": True, "commands": cmds, "count": len(cmds)}

# Alias routes to support clients calling without /api prefix
@app.post("/spreadsheet/commands")
async def enqueue_spreadsheet_commands_alias(request: Request):
    return await enqueue_spreadsheet_commands(request)

@app.get("/spreadsheet/commands")
async def dequeue_spreadsheet_commands_alias(sessionId: str):
    return await dequeue_spreadsheet_commands(sessionId)


# Mapping endpoint: expose cell mapping for templates (for Max key edits)
@app.get("/api/spreadsheet/mapping")
async def get_spreadsheet_mapping():
    try:
        import csv
        import openpyxl
        base = Path(__file__).resolve().parent
        candidates_csv = [base / "delamapping.csv", base / "deal_manager_mapping.csv"]
        candidate_xlsx = base / "delamapping.xlsx"
        rows = []
        csv_path = None
        for c in candidates_csv:
            if c.exists():
                csv_path = c
                break
        if csv_path:
            with csv_path.open("r", newline="", encoding="utf-8") as f:
                # Read all lines and find the header row
                all_lines = f.readlines()
                header_row_idx = None
                for i, line in enumerate(all_lines):
                    if "OM Field Name" in line or "json_field" in line.lower() or "key" in line.lower():
                        header_row_idx = i
                        break
                
                if header_row_idx is None:
                    # Fallback: assume first row is header
                    header_row_idx = 0
                
                # Parse from header row onwards
                f.seek(0)
                for _ in range(header_row_idx):
                    f.readline()
                
                reader = csv.DictReader(f)
                for row in reader:
                    # Support multiple column name formats
                    key = row.get("json_field") or row.get("key") or row.get("OM Field Name") or row.get(" OM Field Name")
                    cell = row.get("cell") or row.get("Maps To Cell") or row.get(" Maps To Cell")
                    # Skip empty rows or header-like rows
                    if not key or not cell or not key.strip() or not cell.strip():
                        continue
                    if "PROPERTY INFORMATION" in key or "FINANCIAL" in key or "INSTRUCTIONS" in key:
                        continue
                    rows.append({
                        "section": row.get("section") or row.get("Model Section") or row.get(" Model Section"),
                        "key": key.strip(),
                        "cell": cell.strip(),
                        "label": row.get("label") or row.get("OM Field Name") or row.get(" OM Field Name"),
                        "type": row.get("type") or row.get("Data Type") or row.get(" Data Type"),
                        "sheet": row.get("sheet"),
                        "required": (row.get("required") or row.get("Notes") or row.get(" Notes") or "").strip().lower() in ("yes", "true", "1", "required")
                    })
        elif candidate_xlsx.exists():
            # Parse mapping from XLSX (prefer sheet named 'Mapping', else active)
            wb = openpyxl.load_workbook(candidate_xlsx, data_only=True)
            ws = wb["Mapping"] if "Mapping" in wb.sheetnames else wb.active
            headers = [str((ws.cell(row=1, column=i).value or "").strip()).lower() for i in range(1, ws.max_column + 1)]
            col_index = {h: i for i, h in enumerate(headers)}
            def cell_val(r, name):
                idx = col_index.get(name)
                if idx is None:
                    return None
                v = ws.cell(row=r, column=idx + 1).value
                return v if v is None else str(v)
            for r in range(2, ws.max_row + 1):
                key = cell_val(r, "json_field") or cell_val(r, "key")
                cell = cell_val(r, "cell")
                if not key or not cell:
                    continue
                rows.append({
                    "section": cell_val(r, "section"),
                    "key": key,
                    "cell": cell,
                    "label": cell_val(r, "label"),
                    "type": cell_val(r, "type"),
                    "sheet": cell_val(r, "sheet"),
                    "required": str(cell_val(r, "required") or "").strip().lower() in ("yes", "true", "1")
                })
            wb.close()
        else:
            raise HTTPException(status_code=404, detail="Mapping file not found (CSV/XLSX)")
        return {"success": True, "mapping": rows}
    except HTTPException:
        raise
    except Exception as e:
        log.exception("Failed to load mapping: %s", e)
        return JSONResponse(status_code=500, content={"success": False, "error": str(e)})


def _preview(key: str, length: int = 8) -> str:
    if not key:
        return "None"
    return f"{key[:length]}..."


log.info("[ENV] Stripe key prefix: %s", _preview(os.getenv("STRIPE_SECRET_KEY")))
log.info("[ENV] Frontend URL: %s", os.getenv("FRONTEND_URL"))
log.info("[ENV] Stripe webhook secret prefix: %s", _preview(os.getenv("STRIPE_WEBHOOK_SECRET")))  # Updated logging for webhook secret


# Stripe Checkout session creation endpoint for plan selection
@app.post("/api/create-checkout-session")
async def create_checkout_session(request: Request):
    data = await request.json()
    email = data.get("email")
    plan = data.get("plan")
    password = data.get("password")
    first_name = data.get("firstName", "")
    last_name = data.get("lastName", "")
    phone = data.get("phone", "")
    company = data.get("company", "")
    title = data.get("title", "")
    city = data.get("city", "")
    state = data.get("state", "")
    
    if not email or not plan or not password:
        raise HTTPException(status_code=400, detail="Missing required fields: email, plan, password")

    # Map plan to Stripe price ID
    price_id = None
    if plan == "pro":
        price_id = PRICE_ID_PRO
    elif plan == "base":
        price_id = PRICE_ID_BASE
    else:
        raise HTTPException(status_code=400, detail="Invalid plan")

    try:
        stripe.api_key = os.getenv("STRIPE_SECRET_KEY")
        checkout_session = stripe.checkout.Session.create(
            payment_method_types=["card"],
            line_items=[{
                "price": price_id,
                "quantity": 1,
            }],
            mode="subscription",
            customer_email=email,
            metadata={
                "email": email,
                "password": password,
                "first_name": first_name,
                "last_name": last_name,
                "phone": phone,
                "company": company,
                "title": title,
                "city": city,
                "state": state,
                "plan": plan
            },
            success_url=os.getenv("FRONTEND_URL", "http://localhost:3000") + f"/signup-complete?session_id={{CHECKOUT_SESSION_ID}}",
            cancel_url=os.getenv("FRONTEND_URL", "http://localhost:3000") + "/signup?canceled=true",
        )
        return {"url": checkout_session.url}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/get-checkout-session")
async def get_checkout_session(session_id: str):
    try:
        stripe.api_key = os.getenv("STRIPE_SECRET_KEY")
        session = stripe.checkout.Session.retrieve(session_id)
        return {"metadata": session.get("metadata", {})}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.on_event("startup")
async def _init_clients():
    global MISTRAL, ANTHROPIC

    log.info(f"MISTRAL_API_KEY exists: {bool(MISTRAL_API_KEY)}")
    log.info(f"ANTHROPIC_API_KEY exists: {bool(ANTHROPIC_API_KEY)}")
    
    if MISTRAL_API_KEY:
        try:
            MISTRAL = Mistral(api_key=MISTRAL_API_KEY)
            log.info("MISTRAL client initialized successfully")
        except Exception as e:
            log.exception("Failed to init Mistral: %s", e)
    else:
        log.warning("MISTRAL_API_KEY missing")

    if ANTHROPIC_API_KEY:
        try:
            ANTHROPIC = Anthropic(api_key=ANTHROPIC_API_KEY)
            log.info(f"ANTHROPIC client initialized successfully: {ANTHROPIC is not None}")
        except Exception as e:
            log.exception("Failed to init Anthropic: %s", e)
    else:
        log.warning("ANTHROPIC_API_KEY/CLAUDE_API_KEY missing")

# ---------------- Utils ----------------
def _to_data_url(file_bytes: bytes, mime: str) -> str:
    return f"data:{mime};base64,{base64.b64encode(file_bytes).decode('utf-8')}"

def _parse_pages_string(pages: str, total_pages: int) -> List[int]:
    pages = (pages or "").replace(" ", "")
    if not pages:
        return []
    result = set()
    for chunk in pages.split(","):
        if "-" in chunk:
            a, b = chunk.split("-", 1)
            start, end = int(a), int(b)
            if start < 1 or end > total_pages or start > end:
                raise ValueError(f"Bad range '{chunk}' (doc has {total_pages} pages)")
            for p in range(start, end + 1):
                result.add(p - 1)
        else:
            p = int(chunk)
            if p < 1 or p > total_pages:
                raise ValueError(f"Bad page '{chunk}'")
            result.add(p - 1)
    return sorted(result)

def _slice_pdf(pdf_bytes: bytes, pages_spec: str) -> bytes:
    reader = PdfReader(io.BytesIO(pdf_bytes))
    idxs = _parse_pages_string(pages_spec, len(reader.pages))
    if not idxs:
        return pdf_bytes
    writer = PdfWriter()
    for i in idxs:
        writer.add_page(reader.pages[i])
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()

def _num(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")
    s = s.replace("$", "").replace(",", "").replace("%", "").replace("(", "").replace(")", "")
    try:
        x = float(s)
        return -x if neg else x
    except:
        return None

def _as_number(v):
    try:
        return float(v) if v is not None else None
    except Exception:
        return None

# ---------------- OCR + Claude ----------------
def _call_mistral_ocr(doc_bytes: bytes, mime: str) -> dict:
    if MISTRAL is None:
        raise HTTPException(status_code=503, detail="Mistral not configured")
    
    try:
        resp = MISTRAL.ocr.process(
            model=OCR_MODEL,
            document={"type": "document_url", "document_url": _to_data_url(doc_bytes, mime)},
            include_image_base64=False,
        )
        return json.loads(resp.model_dump_json())
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Mistral OCR call failed: {e}")

def _call_claude_parse_from_markdown(ocr_text: str, financing_params: Optional[Dict] = None) -> Dict[str, Any]:
    financing_lines = []
    if financing_params:
        financing_lines.append("USER FINANCING (use when applicable):")
        for k, v in financing_params.items():
            financing_lines.append(f"- {k}: {v}")
    financing_context = "\n".join(financing_lines)

    schema_block = """
Return ONLY JSON. Extract EVERYTHING into this schema.

{
    "property": {
        "address": "", "city": "", "state": "", "zip": "",
        "units": 0, "year_built": 0, "rba_sqft": 0, "land_area_acres": 0,
        "property_type": "", "property_class": "", "parking_spaces": 0
    },
    "pricing_financing": {
        "price": 0, "price_per_unit": 0, "price_per_sf": 0,
        "loan_amount": 0, "down_payment": 0, "interest_rate": 0, "ltv": 0,
        "term_years": 0, "amortization_years": 0, "io_period_years": 0,
        "monthly_payment": 0, "annual_debt_service": 0, "debt_type": "",
        "balloon_amount": 0
    },
    "pnl": {
        "gross_potential_rent": 0, "other_income": 0,
        "vacancy_rate": 0, "vacancy_amount": 0,
        "vacancy_rate_t12": 0,
        "vacancy_rate_current": 0,
        "vacancy_rate_stabilized": 0,
        "effective_gross_income": 0,
        "operating_expenses": 0,
        "noi": 0,
        "noi_t12": 0,
        "noi_proforma": 0,
        "noi_stabilized": 0,
        "cap_rate": 0,
        "cap_rate_t12": 0,
        "cap_rate_proforma": 0,
        "cap_rate_stabilized": 0,
        "expense_ratio": 0,
        "expense_ratio_t12": 0,
        "expense_ratio_proforma": 0
    },
    "expenses": {
        "taxes": 0, "insurance": 0, "utilities": 0, "repairs_maintenance": 0,
        "management": 0, "payroll": 0, "admin": 0, "marketing": 0, "other": 0, "total": 0
    },
    "unit_mix": [{"type":"", "units":0, "unit_sf":0, "rent_current":0, "rent_market":0}],
    "underwriting": {"dscr":0, "cap_rate":0, "cash_on_cash":0, "irr":0},
    "deal_analysis": {}, "suggested_financing": {}, "time_periods": {},
    "data_quality": {"confidence":0.0, "missing_fields":[], "assumptions":[]}
}

NOI FIELD RULES:
- If the OM shows multiple NOIs, always separate them:
    • Actual / T12 / Current NOI → store in both pnl.noi_t12 and pnl.noi
    • Pro Forma / Year 1 projected NOI → store in pnl.noi_proforma
    • Stabilized NOI after renovations/lease-up → store in pnl.noi_stabilized
- NEVER put broker pro forma NOI into pnl.noi_t12 if an actual/T12 NOI is available.
- If ONLY pro forma NOI is given and you truly cannot find actual/T12, put it in pnl.noi_proforma and ALSO copy to pnl.noi, and add a short explanation to data_quality.assumptions.

CAP RATE & VACANCY FIELD RULES:
- If the OM shows multiple cap rates (T12 / In-Place, T3, Pro Forma, Stabilized), always separate them:
    • In-Place / Actual / T12 Cap Rate → store in both pnl.cap_rate_t12 and pnl.cap_rate (this is the default cap rate for underwriting).
    • Pro Forma / Year 1 Cap Rate → store in pnl.cap_rate_proforma.
    • Stabilized Cap Rate (after renovations/lease-up) → store in pnl.cap_rate_stabilized.
- If only a Pro Forma cap rate is given and there is no clear actual/T12 cap rate, put it in pnl.cap_rate_proforma and ALSO copy to pnl.cap_rate.

- If the OM shows multiple vacancy or occupancy rates (current, T12, stabilized/pro forma), always separate them:
    • Actual / T12 Vacancy Rate → store in pnl.vacancy_rate_t12 and ALSO as pnl.vacancy_rate if that is the primary underwriting basis.
    • Current Vacancy / Occupancy (point-in-time) → store in pnl.vacancy_rate_current (convert occupancy to vacancy if needed).
    • Pro Forma / Stabilized Vacancy Rate → store in pnl.vacancy_rate_stabilized.
""".strip()

    prompt = (
        "You are an expert underwriter. Parse ALL financial and property data.\n\n"
        "OCR TEXT\n--------\n" + ocr_text + "\n\n" + schema_block + "\n\n" + financing_context
    )

    if ANTHROPIC is None:
        raise HTTPException(status_code=503, detail="Anthropic/Claude not configured")

    try:
        res = ANTHROPIC.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=4000,
            temperature=0,
            messages=[{"role": "user", "content": prompt}],
        )
        txt = res.content[0].text.strip().replace("```json", "").replace("```", "")
        m = re.search(r"\{.*\}\s*$", txt, re.DOTALL)
        return json.loads(m.group(0) if m else txt)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=502, detail=f"Claude JSON parse failed: {e}")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Claude API error: {e}")

# ---------------- Deterministic Extractors ----------------
CITY_STATE_ZIP_RE = re.compile(r"([A-Za-z .'\-]+),\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)")

def _extract_address(md: str) -> Dict[str, str]:
    out = {"address": "", "city": "", "state": "", "zip": ""}
    lines = [l.strip() for l in (md or "").splitlines()]
    for i, ln in enumerate(lines):
        m = CITY_STATE_ZIP_RE.search(ln)
        if m:
            out["city"], out["state"], out["zip"] = m.group(1).strip(), m.group(2), m.group(3)
            for j in range(i - 1, max(-1, i - 5), -1):
                prev = lines[j].strip()
                if prev and any(ch.isdigit() for ch in prev):
                    out["address"] = prev
                    break
            break
    return out

def _extract_unit_mix_from_markdown(md: str) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    in_mix = False
    for raw in (md or "").splitlines():
        ln = raw.strip()
        lnl = ln.lower()
        if ("unit mix" in lnl) or ("rent roll" in lnl) or ("lease summary" in lnl):
            in_mix = True
            continue
        if in_mix and (not ln or lnl.startswith("operating") or lnl.startswith("income") or lnl.startswith("expense")):
            if rows:
                break
        if not in_mix:
            continue
        # Try multiple table formats
        cols = [c for c in re.split(r"\s{2,}|\t|\|", ln) if c.strip()]
        if len(cols) >= 3:
            # Look for unit type patterns
            type_match = re.search(r'(studio|1br|1-br|1 bed|2br|2-br|2 bed|3br|3-br|3 bed|4br|4-br|4 bed)', cols[0].lower())
            if type_match or (("bed" in cols[0].lower()) or ("studio" in cols[0].lower()) or re.search(r"\d\s*br", cols[0].lower())):
                # Extract numbers from remaining columns
                nums = []
                for col in cols[1:]:
                    num = _num(col)
                    if num is not None:
                        nums.append(num)

                if len(nums) >= 2:  # At least units and rent
                    u = nums[0]  # units
                    sf = nums[1] if len(nums) > 1 else 0  # SF
                    rc = nums[2] if len(nums) > 2 else nums[-1]  # Current rent (3rd col) or last if fewer
                    rm = nums[3] if len(nums) > 3 else rc  # Market rent (4th col) or same as current

                    rows.append({
                        "type": cols[0].replace("–", "-").strip(),
                        "units": int(u or 0),
                        "unit_sf": float(sf or 0),
                        "rent_current": float(rc or 0),
                        "rent_market": float(rm or rc or 0),
                    })
    return rows

_PRICING_BLOCK_RE = re.compile(r"\bPRICING DETAIL\b(?:.|\n){0,2000}", re.IGNORECASE)
_PRICE_LINE_RE   = re.compile(r"\b(List|Asking|Offer(?:ing)?) Price\b[: ]*\$?\(?([\d,]+(?:\.\d+)?)\)?", re.IGNORECASE)
_UNITS_LINE_RE   = re.compile(r"\b(Number of Units|Total Units)\b[: ]*([\d,]+)", re.IGNORECASE)
_PPU_LINE_RE     = re.compile(r"\bPrice per Unit\b[: ]*\$?\(?([\d,]+(?:\.\d+)?)\)?", re.IGNORECASE)
_RSF_LINE_RE     = re.compile(r"\b(Rentable(?:\s+Square\s+Foot)?|Rentable SF|RBA|Building SF)\b[: ]*([\d,]+)", re.IGNORECASE)
_PSF_LINE_RE     = re.compile(r"\bPrice per Square Foot\b[: ]*\$?\(?([\d,]+(?:\.\d+)?)\)?", re.IGNORECASE)
_LOTSZ_SF_RE     = re.compile(r"\b(Lot Size|Site Area)\b[: ]*([\d,]+)\s*SF\b", re.IGNORECASE)
_LOTSZ_AC_RE     = re.compile(r"\b(Lot Size|Site Area)\b[: ]*([\d\.]+)\s*acres?\b", re.IGNORECASE)

def _extract_pricing_detail_from_markdown(md: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    block = (_PRICING_BLOCK_RE.search(md or "") or re.search(r"\bPRICING\b(?:.|\n){0,2000}", md or "", re.IGNORECASE))
    text = block.group(0) if block else (md or "")
    m = _PRICE_LINE_RE.search(text); out["price"] = _num(m.group(2)) if m else None
    m = _UNITS_LINE_RE.search(text); out["units"] = _num(m.group(2)) if m else None
    m = _PPU_LINE_RE.search(text);   out["price_per_unit"] = _num(m.group(1)) if m else None
    m = _RSF_LINE_RE.search(text);   out["rba_sqft"] = _num(m.group(2)) if m else None
    m = _PSF_LINE_RE.search(text);   out["price_per_sf"] = _num(m.group(1)) if m else None
    m = _LOTSZ_SF_RE.search(text)
    if m: out["land_area_sf"] = _num(m.group(2))
    m = _LOTSZ_AC_RE.search(text)
    if m: out["land_area_acres"] = _num(m.group(2))
    return out

_OP_SUM_RE = re.compile(r"\b(OPERATING SUMMARY|INCOME STATEMENT)\b(?:.|\n){0,3500}", re.IGNORECASE)
_LABEL_MAP = {
    "GROSS POTENTIAL RE": "gross_potential_rent",
    "GROSS POTENTIAL RENT": "gross_potential_rent",
    "TOTAL ECONOMIC LOSSES": "vacancy_amount",
    "VACANCY": "vacancy_amount",
    "OTHER INCOME": "other_income",
    "EFFECTIVE GROSS INCOME": "effective_gross_income",
    "REAL ESTATE TAXES": "taxes",
    "INSURANCE": "insurance",
    "GENERAL & ADMINISTRATIVE": "admin",
    "GENERAL AND ADMINISTRATIVE": "admin",
    "REPAIRS, MAINTENANCE, & CONTRACT SERVICES": "repairs_maintenance",
    "REPAIRS, MAINTENANCE. & CONTRACT SERVICES": "repairs_maintenance",
    "REPAIRS & MAINTENANCE": "repairs_maintenance",
    "REPAIRS": "repairs_maintenance",
    "UTILITIES": "utilities",
    "TURNOVER & MARKETING": "marketing",
    "MARKETING": "marketing",
    "MANAGEMENT & LEASING": "management",
    "MANAGEMENT": "management",
    "RESERVES": "other",
    "TOTAL OPERATING EXPENSES": "operating_expenses",
    "NET OPERATING INCOME": "noi",
}
def _extract_operating_summary_from_markdown(md: str) -> Dict[str, Any]:
    """Extract high-level P&L and expense lines from markdown.

    This is the legacy helper that operates on a single markdown blob.
    It is kept as-is for backward compatibility. A new helper
    `_extract_operating_summary_sources` (defined below) performs a
    similar extraction but also tracks page/line sources using the full
    `ocr_json` structure.
    """
    out_pnl: Dict[str, Any] = {}
    out_exp: Dict[str, Any] = {}
    block_m = _OP_SUM_RE.search(md or "")
    text = block_m.group(0) if block_m else (md or "")
    if not text:
        return {"pnl": {}, "expenses": {}}
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue
        mnum = re.search(r"\$?\(?-?[\d,]+(?:\.\d+)?\)?", line)
        if not mnum:
            continue
        actual = _num(mnum.group(0))
        label = re.sub(r"\$?\(?-?[\d,]+(?:\.\d+)?\)?.*$", "", line).strip().upper()
        key = None
        for k, v in _LABEL_MAP.items():
            if k in label:
                key = v; break
        if key is None:
            continue
        if key in {"taxes","insurance","admin","repairs_maintenance","utilities","marketing","management","other"}:
            out_exp[key] = actual
        else:
            out_pnl[key] = actual
    if out_exp and "operating_expenses" not in out_pnl:
        out_pnl["operating_expenses"] = sum(v for v in out_exp.values() if isinstance(v, (int, float)))
    if out_pnl.get("gross_potential_rent") is not None and out_pnl.get("vacancy_amount") is not None:
        gpr = out_pnl["gross_potential_rent"]; vac = abs(out_pnl["vacancy_amount"])
        if gpr:
            out_pnl["vacancy_rate"] = round(vac / gpr, 4)
    return {"pnl": out_pnl, "expenses": out_exp}

def _extract_operating_summary_sources(ocr_json: dict) -> Dict[str, Any]:
    """Best-effort mapping of key fields to their OCR page/line.

    Returns a flat dict mapping field paths (e.g. "pnl.noi",
    "expenses.taxes") to simple source metadata:

        {"pnl.noi": {"page": 5, "line_index": 12, "text": "..."}, ...}

    This does not currently rely on bounding boxes; if the upstream OCR
    payload is later extended with coordinates, this helper can be
    enhanced to include a "bbox" entry while keeping the same shape.
    """
    if not isinstance(ocr_json, dict):
        return {}

    pages = ocr_json.get("pages", []) or []
    if not pages:
        return {}

    sources: Dict[str, Any] = {}

    for page_idx, page in enumerate(pages):
        md = (page.get("markdown") or "")
        if not md:
            continue

        block_m = _OP_SUM_RE.search(md)
        text = block_m.group(0) if block_m else md
        if not text:
            continue

        for line_idx, raw in enumerate(text.splitlines()):
            line = raw.strip()
            if not line:
                continue

            mnum = re.search(r"\$?\(?-?[\d,]+(?:\.\d+)?\)?", line)
            if not mnum:
                continue

            label = re.sub(r"\$?\(?-?[\d,]+(?:\.\d+)?\)?.*$", "", line).strip().upper()
            key = None
            for k, v in _LABEL_MAP.items():
                if k in label:
                    key = v
                    break
            if key is None:
                continue

            # Only record the first occurrence for each field to keep
            # the mapping simple and deterministic.
            if key in {"taxes","insurance","admin","repairs_maintenance","utilities","marketing","management","other"}:
                path = f"expenses.{key}"
            else:
                path = f"pnl.{key}"

            if path in sources:
                continue

            sources[path] = {
                "page": page_idx + 1,  # 1-based page index for UI
                "line_index": line_idx,
                "text": line,
            }

    return sources

RECOVERY_KEYWORDS = [
    "PRICING DETAIL","LIST PRICE","ASKING PRICE","OFFERING PRICE","PURCHASE PRICE",
    "PRICE PER UNIT","PRICE PER SQUARE FOOT","NUMBER OF UNITS","RENTABLE SQUARE FOOT","RBA","LOT SIZE",
    "OPERATING SUMMARY","INCOME STATEMENT","GROSS POTENTIAL RENT","EFFECTIVE GROSS INCOME",
    "OTHER INCOME","VACANCY","OPERATING EXPENSES","NET OPERATING INCOME","NOI",
    "UNIT MIX","RENT ROLL","MARKET RENT","CURRENT RENT","ADDRESS","PROPERTY OVERVIEW"
]
def _pages_with_keywords(ocr_json: dict, keywords: List[str]) -> List[int]:
    hits = []
    for i, p in enumerate(ocr_json.get("pages", []) or []):
        md = (p.get("markdown") or "").lower()
        if any(k.lower() in md for k in keywords):
            hits.append(i)
    return sorted(set(hits)) or list(range(min(3, len(ocr_json.get("pages", [])))))

def _merge_truthy(dst: Dict[str, Any], src: Dict[str, Any], fields: List[str]):
    for f in fields:
        v = src.get(f)
        if v is not None and v != "":
            dst[f] = v

# ---------------- Normalize / Enrich / Compute ----------------
def _normalize_parsed(raw: Dict[str, Any]) -> Dict[str, Any]:
    property_ = raw.get("property", {}) or {}
    pricing   = raw.get("pricing_financing", {}) or {}
    pnl_in    = raw.get("pnl", {}) or {}
    expenses_in = raw.get("expenses", {}) or {}

    # ---------------- Core PNL normalization ----------------
    # Normalize vacancy rate variants (T12 vs current vs stabilized)
    vac_t12 = _as_number(pnl_in.get("vacancy_rate_t12"))
    vac_current = _as_number(pnl_in.get("vacancy_rate_current") or pnl_in.get("vacancy_rate"))
    vac_stabilized = _as_number(pnl_in.get("vacancy_rate_stabilized"))

    raw_vacancy = None
    for candidate in (vac_t12, vac_current, vac_stabilized):
        if candidate is not None:
            raw_vacancy = candidate
            break

    # Normalize to a percentage like 5 (for 5%), not 0.05 or 500
    if raw_vacancy is not None:
        if raw_vacancy > 100:  # 500 means 5% stored wrong
            raw_vacancy = raw_vacancy / 100
        elif 0 < raw_vacancy < 1:  # 0.05 → 5
            raw_vacancy = raw_vacancy * 100

    # Prefer actual/T12 NOI when available, but also surface
    # pro forma and stabilized variants so downstream consumers
    # can see the full picture.
    noi_t12 = _as_number(
        pnl_in.get("noi_t12")
        or pnl_in.get("noi_actual")
        or pnl_in.get("noi_t_12")
        or pnl_in.get("noi_t-12")
    )
    noi_proforma = _as_number(
        pnl_in.get("noi_proforma")
        or pnl_in.get("noi_pro_forma")
        or pnl_in.get("noi_year_1")
        or pnl_in.get("noi_year1")
        or pnl_in.get("noi_y1")
    )
    noi_stabilized = _as_number(
        pnl_in.get("noi_stabilized")
        or pnl_in.get("noi_stab")
        or pnl_in.get("noi_stabilised")
    )

    noi_generic = _as_number(pnl_in.get("noi") or pnl_in.get("noi_current"))
    noi_value = None
    for candidate in (noi_t12, noi_generic, noi_proforma):
        if candidate is not None:
            noi_value = candidate
            break

    # Cap rate variants (T12 vs pro forma)
    cap_t12 = _as_number(pnl_in.get("cap_rate_t12") or pnl_in.get("cap_rate_in_place") or pnl_in.get("cap_rate_actual"))
    cap_proforma = _as_number(pnl_in.get("cap_rate_proforma") or pnl_in.get("cap_rate_year_1") or pnl_in.get("cap_rate_y1"))
    cap_generic = _as_number(pnl_in.get("cap_rate") or pnl_in.get("cap_rate_current"))

    cap_value = None
    for candidate in (cap_t12, cap_generic, cap_proforma):
        if candidate is not None:
            cap_value = candidate
            break

    # Expense ratio variants (T12 vs pro forma)
    exp_ratio_t12 = _as_number(pnl_in.get("expense_ratio_t12"))
    exp_ratio_proforma = _as_number(pnl_in.get("expense_ratio_proforma"))
    exp_ratio_generic = _as_number(pnl_in.get("expense_ratio") or pnl_in.get("noi_margin_current"))

    exp_ratio_value = None
    for candidate in (exp_ratio_t12, exp_ratio_generic, exp_ratio_proforma):
        if candidate is not None:
            exp_ratio_value = candidate
            break

    pnl = {
        "gross_potential_rent": _as_number(pnl_in.get("gross_potential_rent") or pnl_in.get("scheduled_gross_rent_current") or pnl_in.get("market_rent_current")),
        "other_income": _as_number(pnl_in.get("other_income")),
        # Canonical vacancy used in downstream calcs (percentage number like 5 for 5%)
        "vacancy_rate": raw_vacancy,
        "vacancy_amount": _as_number(pnl_in.get("vacancy_amount") or pnl_in.get("vacancy_amount_current")),
        "effective_gross_income": _as_number(pnl_in.get("effective_gross_income") or pnl_in.get("effective_gross_income_current")),
        "operating_expenses": _as_number(pnl_in.get("operating_expenses") or pnl_in.get("expenses_current")),
        # NOI variants
        "noi_t12": noi_t12,
        "noi_proforma": noi_proforma,
        "noi_stabilized": noi_stabilized,
        "noi": noi_value,
        # Vacancy variants
        "vacancy_rate_t12": vac_t12,
        "vacancy_rate_current": vac_current,
        "vacancy_rate_stabilized": vac_stabilized,
        # Cap rate variants
        "cap_rate": cap_value,
        "cap_rate_t12": cap_t12,
        "cap_rate_proforma": cap_proforma,
        "cap_rate_stabilized": _as_number(pnl_in.get("cap_rate_stabilized")),
        # Expense ratio variants
        "expense_ratio": exp_ratio_value,
        "expense_ratio_t12": exp_ratio_t12,
        "expense_ratio_proforma": exp_ratio_proforma,
    }

    pricing_norm = {
        "price": _as_number(pricing.get("price")),
        "price_per_unit": _as_number(pricing.get("price_per_unit")),
        "price_per_sf": _as_number(pricing.get("price_per_sf")),
        "loan_amount": _as_number(pricing.get("loan_amount")),
        "down_payment": _as_number(pricing.get("down_payment")),
        "interest_rate": _as_number(pricing.get("interest_rate")),
        "ltv": _as_number(pricing.get("ltv")),
        "term_years": _as_number(pricing.get("term_years")),
        "amortization_years": _as_number(pricing.get("amortization_years")),
        "io_period_years": _as_number(pricing.get("io_period_years")),
        "monthly_payment": _as_number(pricing.get("monthly_payment")),
        "annual_debt_service": _as_number(pricing.get("annual_debt_service") or pricing.get("annual_payment")),
        "debt_type": pricing.get("debt_type"),
        "balloon_amount": _as_number(pricing.get("balloon_amount")),
    }

    out = {
        "property": {
            "address": property_.get("address"),
            "city": property_.get("city"),
            "state": property_.get("state"),
            "zip": property_.get("zip"),
            "units": _as_number(property_.get("units") or property_.get("total_units") or property_.get("number_of_units") or property_.get("unit_count")),
            "year_built": _as_number(property_.get("year_built") or property_.get("built") or property_.get("year")),
            "rba_sqft": _as_number(property_.get("rba_sqft") or property_.get("rentable_sqft") or property_.get("building_sf")),
            "land_area_acres": _as_number(property_.get("land_area_acres") or property_.get("land_acres")),
            "property_type": property_.get("property_type") or "",
            "property_class": property_.get("property_class") or "",
            "parking_spaces": _as_number(property_.get("parking_spaces")),
        },
        "pricing_financing": pricing_norm,
        "pnl": pnl,
        "expenses": expenses_in or {},
        "unit_mix": raw.get("unit_mix", []) or [],
        "underwriting": raw.get("underwriting", {}) or {},
        "deal_analysis": raw.get("deal_analysis", {}) or {},
        "suggested_financing": raw.get("suggested_financing", {}) or {},
        "time_periods": raw.get("time_periods", {}) or {},
        "data_quality": raw.get("data_quality", {}) or {},
        "warnings": raw.get("warnings", []) or [],
    }

    nmix = []
    for r in (raw.get("unit_mix", []) or []):
        if not isinstance(r, dict):
            continue
        nmix.append({
            "type": r.get("type") or r.get("name") or r.get("plan") or "",
            "units": _as_number(r.get("units") or r.get("count") or r.get("qty")),
            "unit_sf": _as_number(r.get("unit_sf") or r.get("sf") or r.get("unit_size") or r.get("size")),
            "rent_current": _as_number(r.get("rent_current") or r.get("current_rent") or r.get("unit_rent") or r.get("rent")),
            "rent_market": _as_number(r.get("rent_market") or r.get("market_rent") or r.get("proforma_rent") or r.get("rent")),
        })
    out["unit_mix"] = [x for x in nmix if x and (x["type"] or x["units"])]
    return out

def _validate_and_enrich(data: Dict[str, Any]) -> Dict[str, Any]:
    warnings, calc = [], []
    price = _as_number(data.get("pricing_financing", {}).get("price"))
    units = _as_number(data.get("property", {}).get("units"))
    noi   = _as_number(data.get("pnl", {}).get("noi"))
    egi   = _as_number(data.get("pnl", {}).get("effective_gross_income"))
    opex  = _as_number(data.get("pnl", {}).get("operating_expenses"))
    debt  = _as_number(data.get("pricing_financing", {}).get("annual_debt_service"))

    if price and units and not _as_number(data["pricing_financing"].get("price_per_unit")):
        data["pricing_financing"]["price_per_unit"] = round(price/units, 2); calc.append("Calculated price per unit")
    if price and noi and not _as_number(data["pnl"].get("cap_rate")):
        data["pnl"]["cap_rate"] = round(noi/price, 4); calc.append("Calculated cap rate")
    if noi and debt and not _as_number(data.get("underwriting", {}).get("dscr")):
        data.setdefault("underwriting", {})["dscr"] = round(noi/debt, 2); calc.append("Calculated DSCR")
    if opex and egi and not _as_number(data["pnl"].get("expense_ratio")):
        data["pnl"]["expense_ratio"] = round(opex/egi, 4); calc.append("Calculated expense ratio")

    if not price: warnings.append("Missing purchase price")
    if not units: warnings.append("Missing unit count")
    if not noi:   warnings.append("Missing NOI")
    if not egi:   warnings.append("Missing Effective Gross Income")

    data.setdefault("metadata", {})
    data["metadata"].update({
        "warnings": warnings,
        "calculations_performed": calc,
        "parser_strategy": PARSER_STRATEGY_DEFAULT,
        "ocr_engine": "mistral-ocr",
    })
    return data

def _pmti(principal: float, annual_rate_pct: float, n_months: int) -> float:
    if principal is None or principal <= 0 or annual_rate_pct is None or n_months is None or n_months <= 0:
        return 0.0
    r = (annual_rate_pct / 100.0) / 12.0
    if r == 0:
        return principal / n_months
    return principal * (r / (1 - (1 + r) ** (-n_months)))

def _remaining_balance(principal: float, annual_rate_pct: float, amort_months: int, payments_made: int) -> float:
    if principal is None or principal <= 0:
        return 0.0
    r = (annual_rate_pct / 100.0) / 12.0
    if r == 0:
        paid = principal * payments_made / amort_months
        return max(0.0, principal - paid)
    return principal * (((1 + r) ** amort_months - (1 + r) ** payments_made) / ((1 + r) ** amort_months - 1))

def _compute_underwriting(d: Dict[str, Any]) -> Dict[str, Any]:
    prop = d.setdefault("property", {})
    pricing = d.setdefault("pricing_financing", {})
    pnl = d.setdefault("pnl", {})
    exps = d.setdefault("expenses", {})
    uw  = d.setdefault("underwriting", {})

    price = _as_number(pricing.get("price"))
    units = _as_number(prop.get("units"))
    noi   = _as_number(pnl.get("noi"))
    egi   = _as_number(pnl.get("effective_gross_income"))
    gpr   = _as_number(pnl.get("gross_potential_rent"))
    vac   = _as_number(pnl.get("vacancy_amount"))
    other = _as_number(pnl.get("other_income"))
    opex  = _as_number(pnl.get("operating_expenses"))
    debt  = _as_number(pricing.get("annual_debt_service"))

    if (price is not None) and (units and units > 0) and not _as_number(pricing.get("price_per_unit")):
        pricing["price_per_unit"] = round(price / units, 2)

    if egi is None and (gpr is not None):
        vac_amt = abs(vac) if vac is not None else 0.0
        other_inc = other if other is not None else 0.0
        pnl["effective_gross_income"] = round(gpr - vac_amt + other_inc, 2)
        egi = pnl["effective_gross_income"]

    if (opex is None) and exps:
        subtotal = 0.0
        for k, v in exps.items():
            if k in {"total", "total_current"}: continue
            val = _as_number(v)
            if val: subtotal += val
        if subtotal > 0:
            exps["total"] = round(subtotal, 2)
            pnl["operating_expenses"] = exps["total"]
            opex = exps["total"]

    if (noi is None) and (egi is not None) and (opex is not None):
        pnl["noi"] = round(egi - opex, 2)
        noi = pnl["noi"]

    if (pnl.get("expense_ratio") is None) and (egi is not None) and (opex is not None) and egi != 0:
        pnl["expense_ratio"] = round(opex / egi, 4)

    if (pnl.get("cap_rate") is None) and (noi is not None) and (price and price != 0):
        pnl["cap_rate"] = round(noi / price, 4)

    if (uw.get("dscr") is None) and (noi is not None) and (debt and debt != 0):
        uw["dscr"] = round(noi / debt, 2)

    return d

def _is_critically_incomplete(d: Dict[str, Any]) -> bool:
    prop, pnl, pricing = d.get("property", {}), d.get("pnl", {}), d.get("pricing_financing", {})
    missing_basic = not prop.get("address") or not prop.get("units") or not pricing.get("price")
    missing_fin   = not pnl.get("gross_potential_rent") or not pnl.get("operating_expenses") or not pnl.get("noi")
    return bool(missing_basic or missing_fin)

def _calculate_deal_metrics(d: Dict[str, Any]) -> Dict[str, Any]:
    """Calculate comprehensive investment metrics programmatically"""
    
    # Extract key values
    price = _as_number(d.get("pricing_financing", {}).get("price")) or 0
    units = _as_number(d.get("property", {}).get("units")) or 0
    noi = _as_number(d.get("pnl", {}).get("noi")) or 0
    egi = _as_number(d.get("pnl", {}).get("effective_gross_income")) or 0
    gpr = _as_number(d.get("pnl", {}).get("gross_potential_rent")) or 0
    opex = _as_number(d.get("pnl", {}).get("operating_expenses")) or 0
    debt_service = _as_number(d.get("pricing_financing", {}).get("annual_debt_service")) or 0
    down_payment = _as_number(d.get("pricing_financing", {}).get("down_payment")) or 0
    loan_amount = _as_number(d.get("pricing_financing", {}).get("loan_amount")) or 0
    vacancy_rate = _as_number(d.get("pnl", {}).get("vacancy_rate")) or 0.05
    cap_rate = _as_number(d.get("pnl", {}).get("cap_rate")) or 0
    expense_ratio = _as_number(d.get("pnl", {}).get("expense_ratio")) or 0
    interest_rate = _as_number(d.get("pricing_financing", {}).get("interest_rate")) or 0
    term_years = _as_number(d.get("pricing_financing", {}).get("term_years")) or 30
    
    metrics = {}
    
    # 1. Annual Cash Flow
    annual_cash_flow = noi - debt_service if debt_service else noi
    monthly_cash_flow = annual_cash_flow / 12
    metrics["annual_cash_flow"] = round(annual_cash_flow, 2)
    metrics["monthly_cash_flow"] = round(monthly_cash_flow, 2)
    
    # 2. Cash-on-Cash Return
    if down_payment and down_payment > 0:
        coc_return = (annual_cash_flow / down_payment) * 100
        metrics["cash_on_cash_return"] = round(coc_return, 2)
    else:
        metrics["cash_on_cash_return"] = 0
    
    # 3. ROI (Return on Investment) - Year 1
    if price and price > 0:
        roi = (annual_cash_flow / price) * 100
        metrics["roi_year_1"] = round(roi, 2)
    else:
        metrics["roi_year_1"] = 0
    
    # 4. Cap Rate
    if price and price > 0 and noi:
        calculated_cap = (noi / price) * 100
        metrics["cap_rate"] = round(calculated_cap, 2)
    else:
        metrics["cap_rate"] = round(cap_rate * 100, 2) if cap_rate else 0
    
    # 5. DSCR (Debt Service Coverage Ratio)
    if debt_service and debt_service > 0:
        dscr = noi / debt_service
        metrics["dscr"] = round(dscr, 2)
    else:
        metrics["dscr"] = 0
    
    # 6. Gross Rent Multiplier
    if price and gpr and gpr > 0:
        grm = price / gpr
        metrics["grm"] = round(grm, 2)
    else:
        metrics["grm"] = 0
    
    # 7. Price per Unit
    if price and units and units > 0:
        ppu = price / units
        metrics["price_per_unit"] = round(ppu, 0)
    else:
        metrics["price_per_unit"] = 0
    
    # 8. Price per Square Foot
    rba_sqft = _as_number(d.get("property", {}).get("rba_sqft")) or 0
    if price and rba_sqft and rba_sqft > 0:
        ppsf = price / rba_sqft
        metrics["price_per_sf"] = round(ppsf, 2)
    else:
        metrics["price_per_sf"] = 0
    
    # 9. Expense Ratio
    if egi and egi > 0 and opex:
        exp_ratio = (opex / egi) * 100
        metrics["expense_ratio"] = round(exp_ratio, 2)
    else:
        metrics["expense_ratio"] = round(expense_ratio * 100, 2) if expense_ratio else 0
    
    # 10. Break-even Ratio
    if egi and egi > 0:
        break_even = ((opex + debt_service) / egi) * 100
        metrics["break_even_ratio"] = round(break_even, 2)
    else:
        metrics["break_even_ratio"] = 0
    
    # 11. Loan-to-Value (LTV)
    if loan_amount and price and price > 0:
        ltv = (loan_amount / price) * 100
        metrics["ltv"] = round(ltv, 2)
    else:
        metrics["ltv"] = 0
    
    # 12. Debt Yield
    if loan_amount and loan_amount > 0 and noi:
        debt_yield = (noi / loan_amount) * 100
        metrics["debt_yield"] = round(debt_yield, 2)
    else:
        metrics["debt_yield"] = 0
    
    # 13. 1% Rule (Monthly Rent vs Purchase Price)
    if gpr and price and price > 0:
        monthly_rent = gpr / 12
        one_percent_rule = (monthly_rent / price) * 100
        metrics["one_percent_rule"] = round(one_percent_rule, 2)
    else:
        metrics["one_percent_rule"] = 0
    
    # 14. IRR Calculation (5-year simplified)
    # Assumptions: 3% annual rent growth, 2% expense growth, 3% property appreciation
    irr_cash_flows = [-down_payment] if down_payment else [-price]
    for year in range(1, 6):
        year_noi = noi * (1.03 ** year) - (opex * (0.02 * year))  # Rent grows 3%, expenses grow 2%
        year_cash_flow = year_noi - debt_service
        if year == 5:  # Add sale proceeds in year 5
            sale_price = price * (1.03 ** 5)  # 3% annual appreciation
            remaining_loan = loan_amount * 0.85 if loan_amount else 0  # Rough estimate of remaining balance
            year_cash_flow += (sale_price - remaining_loan)
        irr_cash_flows.append(year_cash_flow)
    
    # Simple IRR approximation
    try:
        # Newton's method for IRR (simplified)
        rate = 0.1  # Initial guess
        for _ in range(20):
            npv = sum(cf / ((1 + rate) ** i) for i, cf in enumerate(irr_cash_flows))
            dnpv = sum(-i * cf / ((1 + rate) ** (i + 1)) for i, cf in enumerate(irr_cash_flows) if i > 0)
            if abs(dnpv) < 0.0001:
                break
            rate = rate - npv / dnpv
        metrics["irr_5_year"] = round(rate * 100, 2)
    except:
        metrics["irr_5_year"] = 0
    
    # 15. Payback Period (years to recover initial investment)
    if annual_cash_flow > 0 and down_payment:
        payback_period = down_payment / annual_cash_flow
        metrics["payback_period_years"] = round(payback_period, 1)
    else:
        metrics["payback_period_years"] = 0
    
    # 16. Rent to Price Ratio
    if gpr and price and price > 0:
        rent_price_ratio = (gpr / price) * 100
        metrics["rent_to_price_ratio"] = round(rent_price_ratio, 2)
    else:
        metrics["rent_to_price_ratio"] = 0
    
    # 17. Operating Margin
    if egi and egi > 0:
        operating_margin = (noi / egi) * 100
        metrics["operating_margin"] = round(operating_margin, 2)
    else:
        metrics["operating_margin"] = 0
    
    # 18. Cash Flow per Unit
    if units and units > 0:
        cf_per_unit = annual_cash_flow / units
        metrics["cash_flow_per_unit"] = round(cf_per_unit, 2)
    else:
        metrics["cash_flow_per_unit"] = 0
    
    # Now ask Claude for opinion based on these calculated metrics
    metrics_summary = f"""
    Investment Metrics Summary:
    - Cap Rate: {metrics['cap_rate']}%
    - Cash-on-Cash Return: {metrics['cash_on_cash_return']}%
    - DSCR: {metrics['dscr']}
    - Annual Cash Flow: ${metrics['annual_cash_flow']:,.0f}
    - Monthly Cash Flow: ${metrics['monthly_cash_flow']:,.0f}
    - ROI Year 1: {metrics['roi_year_1']}%
    - IRR (5-year): {metrics['irr_5_year']}%
    - GRM: {metrics['grm']}
    - Price per Unit: ${metrics['price_per_unit']:,.0f}
    - Expense Ratio: {metrics['expense_ratio']}%
    - Break-even Ratio: {metrics['break_even_ratio']}%
    - Debt Yield: {metrics['debt_yield']}%
    - 1% Rule: {metrics['one_percent_rule']}%
    - Payback Period: {metrics['payback_period_years']} years
    - Operating Margin: {metrics['operating_margin']}%
    """
    
    prompt = f"""Based on these calculated investment metrics, provide a brief investment opinion.
    
    {metrics_summary}
    
    Return JSON with:
    - verdict: 'STRONG BUY', 'BUY', 'HOLD', 'PASS', or 'STRONG PASS'
    - confidence: 'High', 'Moderate', or 'Low'
    - pros: array of 3-4 positive factors
    - cons: array of 3-4 negative factors or risks
    - summary: 2-3 sentence investment thesis
    
    Focus on the actual numbers and standard investment criteria. Be critical and analytical."""
    
    try:
        res = ANTHROPIC.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=600,
            temperature=0,
            messages=[{"role": "user", "content": prompt}],
        )
        txt = res.content[0].text.strip().replace("```json", "").replace("```", "")
        m = re.search(r"\{.*\}\s*$", txt, re.DOTALL)
        opinion = json.loads(m.group(0) if m else txt)
    except:
        # Fallback opinion based on metrics
        opinion = {
            "verdict": "HOLD",
            "confidence": "Moderate",
            "pros": ["Metrics calculated successfully"],
            "cons": ["Unable to generate detailed analysis"],
            "summary": "Review the metrics carefully before making an investment decision."
        }
    
    # Calculate a simple score based on key metrics
    score = 50  # Base score
    if metrics['cap_rate'] >= 7: score += 10
    elif metrics['cap_rate'] >= 5: score += 5
    if metrics['cash_on_cash_return'] >= 10: score += 10
    elif metrics['cash_on_cash_return'] >= 7: score += 5
    if metrics['dscr'] >= 1.5: score += 10
    elif metrics['dscr'] >= 1.25: score += 5
    if metrics['expense_ratio'] <= 40: score += 10
    elif metrics['expense_ratio'] <= 50: score += 5
    if metrics['irr_5_year'] >= 15: score += 10
    elif metrics['irr_5_year'] >= 10: score += 5
    
    score = min(100, max(0, score))
    
    return {
        "metrics": metrics,
        "score": score,
        "verdict": opinion.get("verdict", "HOLD"),
        "confidence": opinion.get("confidence", "Moderate"),
        "pros": opinion.get("pros", []),
        "cons": opinion.get("cons", []),
        "summary": opinion.get("summary", ""),
        "metrics_summary": metrics_summary
    }

# ---------------- Property Analysis (UNCHANGED) ----------------
def analyze_property_with_market(
   property_data: Dict[str, Any],
   market_data: Dict[str, Any],
   live_rentals: list = None
) -> Dict[str, Any]:
   """
   Analyze property performance against market data using Claude
   """
   rental_summary = ""
   if live_rentals:
       rentals_by_bed = {}
       for r in live_rentals:
           bed = str(r.get('bed', 'Studio'))
           if bed not in rentals_by_bed:
               rentals_by_bed[bed] = []
           rent = r.get('rent')
           if rent:
               try:
                   rentals_by_bed[bed].append(float(str(rent).replace(',', '').replace('$', '')))
               except:
                   pass
       avg_rents = {}
       for bed, rents in rentals_by_bed.items():
           if rents:
               avg_rents[bed if bed != '0' else 'Studio'] = sum(rents) / len(rents)
       lines = [f"- {bed}: ${avg:,.0f}/month average" for bed, avg in avg_rents.items()]
       rental_summary = "\n\nLIVE RENTAL LISTINGS (" + str(len(live_rentals)) + " active):\n" + "\n".join(lines)

   schema_block = """
Format response as JSON with these exact keys:
{
 "summary": "Brief 2-3 sentence overview of property performance",
 "rent_analysis": {
   "current_vs_market": {
     "current_avg": number,
     "market_avg": number,
     "gap": number,
     "gap_percentage": number
   }},
   "unit_type_analysis": [
     {
       "type": "1BR",
       "current_rent": number,
       "market_rent": number,
       "fmr": number,
       "recommendation": "string"
     }
   ],
   "underpriced_units": ["string"],
   "revenue_opportunity": number
 },
 "expense_analysis": {
   "current_ratio": number,
   "benchmark_ratio": 0.38,
   "savings_opportunity": number,
   "specific_cuts": [
     {
       "category": "string",
       "current": number,
       "benchmark": number,
       "savings": number
     }
   ]
 },
 "vacancy_analysis": {
   "current": number,
   "market": number,
   "revenue_impact": number
 },
 "noi_improvement": {
   "current_noi": number,
   "potential_noi": number,
   "total_improvement": number,
   "improvement_pct": number
 },
 "recommendations": [
   {
     "category": "Revenue|Expenses|Operations",
     "action": "Specific action to take",
     "impact": "Dollar amount or percentage impact"
   }
 ],
 "red_flags": ["string"],
 "strengths": ["string"],
 "market_position": "Growing|Stable|Declining market assessment",
 "rubs_opportunity": {
   "current_recovery": number,
   "potential_recovery": number,
   "additional_income": number
 },
 "ancillary_revenue": [
   {
     "type": "Storage|Parking|Pet|Laundry",
     "monthly_per_unit": number,
     "annual_total": number
   }
 ]
}

Return ONLY the JSON object, no additional text.
""".strip()

   prompt = (
       "You are an expert multifamily property consultant. Analyze this property's performance against market data and provide specific recommendations to boost NOI.\n\n"
       "PROPERTY DATA:\n"
       f"- Address: {property_data.get('property', {}).get('address', 'N/A')}\n"
       f"- City: {property_data.get('property', {}).get('city', 'N/A')}\n"
       f"- State: {property_data.get('property', {}).get('state', 'N/A')}\n"
       f"- ZIP: {property_data.get('property', {}).get('zip', 'N/A')}\n"
       f"- Units: {property_data.get('property', {}).get('units', 'N/A')}\n"
       f"- Year Built: {property_data.get('property', {}).get('year_built', 'N/A')}\n\n"
       "PROPERTY FINANCIALS:\n"
       f"- Gross Potential Rent: ${property_data.get('pnl', {}).get('gross_potential_rent', 0):,.0f}\n"
       f"- Vacancy Rate: {property_data.get('pnl', {}).get('vacancy_rate', 0) if property_data.get('pnl', {}).get('vacancy_rate', 0) is not None else 0:.2f}\n"
       f"- Effective Gross Income: ${property_data.get('pnl', {}).get('effective_gross_income', 0):,.0f}\n"
       f"- Operating Expenses: ${property_data.get('pnl', {}).get('operating_expenses', 0):,.0f}\n"
       f"- NOI: ${property_data.get('pnl', {}).get('noi', 0):,.0f}\n"
       f"- Expense Ratio: {property_data.get('pnl', {}).get('expense_ratio', 0) if property_data.get('pnl', {}).get('expense_ratio', 0) is not None else 0:.2f}\n\n"
       "UNIT MIX:\n" + json.dumps(property_data.get('unit_mix', []), indent=2) + "\n\n"
       "EXPENSES BREAKDOWN:\n" + json.dumps(property_data.get('expenses', {}), indent=2) + "\n\n"
       "MARKET DATA:\n"
       f"- ZIP: {market_data.get('zip', 'N/A')}\n"
       f"- Population: {market_data.get('population', 'N/A')}\n"
       f"- Employment Rate: {market_data.get('employment_rate', 0)}\n"
       f"- Median Household Income: {market_data.get('median_household_income', 0)}\n"
       f"- Market Median Rent: {market_data.get('median_gross_rent', 0)}\n"
       f"- Market Vacancy Rate: {market_data.get('vacancy_rate', 0)}\n"
       f"- Total Housing Units: {market_data.get('total_housing_units', 0)}\n"
       f"- Renter Percentage: {market_data.get('pct_renter', 0)}\n\n"
       "FAIR MARKET RENTS (HUD):\n"
       f"- Studio: {market_data.get('fmr_0br', 0)}\n"
       f"- 1BR: {market_data.get('fmr_1br', 0)}\n"
       f"- 2BR: {market_data.get('fmr_2br', 0)}\n"
       f"- 3BR: {market_data.get('fmr_3br', 0)}\n"
       f"- 4BR: {market_data.get('fmr_4br', 0)}\n\n"
       + rental_summary + "\n\n" + schema_block
   )

   try:
       response = ANTHROPIC.messages.create(
           model=ANTHROPIC_MODEL,
           max_tokens=4000,
           temperature=0,
           messages=[{"role": "user", "content": prompt}],
       )
       result_text = response.content[0].text.strip()
       result_text = result_text.replace("```json", "").replace("```", "").strip()
       return json.loads(result_text)
   except json.JSONDecodeError as e:
       raise HTTPException(status_code=502, detail=f"Claude response parsing failed: {e}")
   except Exception as e:
       raise HTTPException(status_code=502, detail=f"Claude API error: {e}")

# ---------------- Visualizations ----------------
def _generate_visualization_data(data: Dict[str, Any]) -> Dict[str, Any]:
    expenses = data.get("expenses", {}) or {}
    pnl = data.get("pnl", {}) or {}
    unit_mix = data.get("unit_mix", []) or []

    total_expenses = expenses.get("total") or expenses.get("total_current") or 0
    breakdown = []
    if total_expenses:
        for k, v in expenses.items():
            if k in {"total", "total_current"}:
                continue
            try:
                val = float(v)
            except Exception:
                continue
            if val:
                breakdown.append({
                    "name": k.replace("_", " ").title(),
                    "value": val,
                    "percentage": round((val / float(total_expenses)) * 100, 1),
                })

    gpr = pnl.get("gross_potential_rent") or 0
    egi = pnl.get("effective_gross_income") or 0
    opex = pnl.get("operating_expenses") or 0
    noi = pnl.get("noi") or 0
    income_expense = []
    if egi or opex or noi:
        income_expense = [
            {"category": "Gross Income", "value": gpr or 0},
            {"category": "Effective Income", "value": egi or 0},
            {"category": "Operating Expenses", "value": opex or 0},
            {"category": "NOI", "value": noi or 0},
        ]

    unit_chart = []
    for u in unit_mix:
        try:
            n = int(float(u.get("units", 0)))
        except Exception:
            n = 0
        if n:
            unit_chart.append({"type": u.get("type", "Unknown"), "units": n, "percentage": 0})

    vac_amt = pnl.get("vacancy_amount") or 0
    other_inc = pnl.get("other_income") or 0
    waterfall = []
    if gpr:
        waterfall = [
            {"name": "Gross Rent", "value": gpr},
            {"name": "Vacancy", "value": -abs(vac_amt)},
            {"name": "Other Income", "value": other_inc},
            {"name": "Operating Expenses", "value": -abs(opex or 0)},
            {"name": "NOI", "value": noi or 0, "isTotal": True},
        ]

    return {
        "expense_breakdown": breakdown,
        "income_expense_comparison": income_expense,
        "unit_mix_chart": unit_chart,
        "cash_flow_waterfall": waterfall,
    }

# ---------------- API ----------------
@app.get("/health")
def health():
    return {
        "ok": True,
        "version": "9.0.0",
        "parser_default": "parser_v4",
        "clients": {
            "mistral": MISTRAL is not None,
            "anthropic": ANTHROPIC is not None,
        }
    }

@app.post("/ocr/underwrite")
async def ocr_and_underwrite(
    file: UploadFile = File(...),
    pages: Optional[str] = Form(default=""),

    # strategy override
    parser_strategy: Optional[str] = Form(default=None),

    # legacy financing params
    loan_amount: Optional[float] = Form(default=None),
    down_payment_pct: Optional[float] = Form(default=None),
    interest_rate: Optional[float] = Form(default=None),
    term_years: Optional[int] = Form(default=None),
    loan_type: Optional[str] = Form(default=None),

    # NEW: mode selector
    financing_mode: Optional[str] = Form(default=None),  # "traditional" | "seller_finance" | "subject_to"

    # traditional
    amort_years: Optional[int] = Form(default=None),

    # seller finance
    down_payment_amount: Optional[float] = Form(default=None),
    sf_interest_rate: Optional[float] = Form(default=None),
    sf_amort_years: Optional[int] = Form(default=None),
    sf_balloon_years: Optional[int] = Form(default=None),
    sf_io_years: Optional[int] = Form(default=0),

    # subject-to
    st_existing_balance: Optional[float] = Form(default=None),
    st_interest_rate: Optional[float] = Form(default=None),
    st_remaining_term_years: Optional[int] = Form(default=None),
    st_amort_years: Optional[int] = Form(default=None),
):
    print(f"\n{'='*80}")
    print(f"[OCR/UNDERWRITE] REQUEST RECEIVED")
    print(f"[OCR/UNDERWRITE] File: {file.filename}")
    print(f"[OCR/UNDERWRITE] Pages: {pages}")
    print(f"{'='*80}\n")
    
    mime = (file.content_type or "").lower()
    if mime not in ALLOWED_UPLOAD_MIMES:
        raise HTTPException(status_code=415, detail=f"Unsupported content type: {mime}")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty upload")
    if len(data) > MAX_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large (> {MAX_BYTES // (1024*1024)} MB)")

    orig_data, orig_mime = data, mime

    strategy = (parser_strategy or PARSER_STRATEGY_DEFAULT).lower().strip()
    if strategy not in {"claude", "om_v4", "hybrid"}:
        strategy = "claude"

    if mime == "application/pdf" and pages:
        try:
            data = _slice_pdf(data, pages)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    if mime in ALLOWED_DOC_MIMES:
        ocr_json = _call_mistral_ocr(data, mime)
        md_parts = [p.get("markdown", "") for p in ocr_json.get("pages", []) if isinstance(p, dict)]
        markdown_text = "\n\n".join([m for m in md_parts if m]).strip()
        if not markdown_text:
            raise HTTPException(status_code=502, detail="No text extracted from OCR")
    else:
        ocr_json = None
        try:
            text = data.decode("utf-8", errors="ignore")
        except Exception:
            text = base64.b64encode(data).decode("utf-8")
        markdown_text = "SPREADSHEET_CONTENT\n\n" + text[:500000]

    # build financing context for Claude prompt
    financing_params = {
        "financing_mode": financing_mode,
        "down_payment_pct": down_payment_pct,
        "term_years": term_years,
        "amort_years": amort_years,
        "loan_amount": loan_amount,
        "seller_dp_amount": down_payment_amount,
        "seller_rate": sf_interest_rate,
        "seller_amort_years": sf_amort_years,
        "seller_balloon_years": sf_balloon_years,
        "seller_io_years": sf_io_years,
        "subto_existing_balance": st_existing_balance,
        "subto_rate": st_interest_rate,
        "subto_remaining_term_years": st_remaining_term_years,
        "subto_amort_years": st_amort_years,
    }

    def _parse_with_claude(md: str):
        return _call_claude_parse_from_markdown(md, financing_params)

    def _parse_with_om_v4(md: str):
        if not HAS_PARSER_V4:
            raise HTTPException(status_code=500, detail="parser_v4 not available")
        try:
            res = _RE_PARSER.parse_with_claude(md, mode="underwriting")  # type: ignore[name-defined]
        except Exception as e:
            raise HTTPException(status_code=502, detail=f"parser_v4 call failed: {e}")
        if not isinstance(res, dict) or not res.get("success"):
            err = (res or {}).get("error") if isinstance(res, dict) else None
            raise HTTPException(status_code=502, detail=f"parser_v4 returned error: {err or 'unknown error'}")
        return res.get("data") or {}

    try:
        if strategy == "claude":
            parsed_raw = _parse_with_claude(markdown_text); used_strategy = "claude"
        elif strategy == "om_v4":
            parsed_raw = _parse_with_om_v4(markdown_text);  used_strategy = "om_v4"
        else:
            try:
                parsed_raw = _parse_with_om_v4(markdown_text); used_strategy = "om_v4"
            except Exception:
                parsed_raw = _parse_with_claude(markdown_text); used_strategy = "claude"
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Parsing failed: {e}")

    normalized = _normalize_parsed(parsed_raw)

    # Deterministic enrichment from selected pages
    prop = normalized.setdefault("property", {})
    pricing = normalized.setdefault("pricing_financing", {})
    pnl = normalized.setdefault("pnl", {})
    exps = normalized.setdefault("expenses", {})

    addr = _extract_address(markdown_text)
    _merge_truthy(prop, addr, ["address", "city", "state", "zip"])
    if not normalized.get("unit_mix"):
        um = _extract_unit_mix_from_markdown(markdown_text)
        if um:
            normalized["unit_mix"] = um
            if not prop.get("units"):
                prop["units"] = sum(r.get("units", 0) or 0 for r in um)

    pr = _extract_pricing_detail_from_markdown(markdown_text)
    if pr.get("units") and not prop.get("units"): prop["units"] = pr["units"]
    if pr.get("rba_sqft") and not prop.get("rba_sqft"): prop["rba_sqft"] = pr["rba_sqft"]
    if pr.get("land_area_acres"): prop["land_area_acres"] = pr["land_area_acres"]
    elif pr.get("land_area_sf") and not prop.get("land_area_acres"):
        try: prop["land_area_acres"] = round(pr["land_area_sf"]/43560.0, 4)
        except: pass
    for k in ["price", "price_per_unit", "price_per_sf"]:
        if pr.get(k) and not pricing.get(k):
            pricing[k] = pr[k]

    ops = _extract_operating_summary_from_markdown(markdown_text)
    for k, v in (ops.get("pnl") or {}).items():
        if v is not None and not pnl.get(k):
            pnl[k] = v
    for k, v in (ops.get("expenses") or {}).items():
        if v is not None and not exps.get(k):
            exps[k] = v

    # Attach lightweight source metadata for key P&L/expense fields so
    # the frontend can show where values came from in the PDF viewer.
    try:
        if ocr_json is not None:
            op_sources = _extract_operating_summary_sources(ocr_json)
            if op_sources:
                normalized.setdefault("metadata", {})["sources"] = op_sources
    except Exception as e:
        # Never let source-mapping failures break underwriting; just
        # record a truncated error string for debugging.
        normalized.setdefault("metadata", {})["sources_error"] = str(e)[:200]

    # ---------- Financing mode compute (overrides/augments Claude) ----------
    fm = (financing_mode or "").strip().lower()
    price_val = _as_number(pricing.get("price"))

    def set_ads(monthly):
        if monthly and monthly > 0:
            pricing["monthly_payment"] = round(monthly, 2)
            pricing["annual_debt_service"] = round(monthly * 12, 2)

    if fm == "traditional":
        dp_pct_v   = _as_number(down_payment_pct)
        rate_v     = _as_number(interest_rate)
        term_y_v   = _as_number(term_years)
        amort_y_v  = _as_number(amort_years or term_y_v)

        if price_val and dp_pct_v is not None:
            loan_amt = price_val * (1 - dp_pct_v/100.0)
            pricing["loan_amount"] = round(loan_amt, 2)
            pricing["down_payment"] = round(price_val - loan_amt, 2)
        else:
            loan_amt = _as_number(pricing.get("loan_amount"))

        n = int((amort_y_v or term_y_v or 30) * 12) if (amort_y_v or term_y_v) else None
        m = _pmti(loan_amt, rate_v or 0, n or 0)
        set_ads(m)
        pricing["debt_type"] = "Traditional"

    elif fm == "seller_finance":
        dp_amt_v  = _as_number(down_payment_amount)
        rate_v    = _as_number(sf_interest_rate)
        amort_y_v = _as_number(sf_amort_years)
        bal_y_v   = _as_number(sf_balloon_years)
        io_y_v    = int(_as_number(sf_io_years or 0) or 0)

        if price_val and dp_amt_v is not None:
            loan_amt = price_val - dp_amt_v
            pricing["loan_amount"] = round(loan_amt, 2)
            pricing["down_payment"] = round(dp_amt_v, 2)
        else:
            loan_amt = _as_number(pricing.get("loan_amount"))

        amort_m = int((amort_y_v or 30) * 12)
        if io_y_v and io_y_v > 0:
            io_months = int(io_y_v * 12)
            io_payment = (rate_v or 0)/100/12 * (loan_amt or 0)
            if bal_y_v and bal_y_v*12 <= io_months:
                balloon_amt = loan_amt
                set_ads(io_payment)
            else:
                rem_months = max(0, (bal_y_v or 0)*12 - io_months) if bal_y_v else amort_m - io_months
                pmt = _pmti(loan_amt, rate_v or 0, amort_m)
                set_ads(io_payment if io_months else pmt)
                if bal_y_v:
                    balloon_amt = _remaining_balance(loan_amt, rate_v or 0, amort_m, rem_months)
                else:
                    balloon_amt = 0.0
        else:
            pmt = _pmti(loan_amt, rate_v or 0, amort_m)
            set_ads(pmt)
            if bal_y_v:
                balloon_amt = _remaining_balance(loan_amt, rate_v or 0, amort_m, int(bal_y_v*12))
            else:
                balloon_amt = 0.0

        pricing["balloon_amount"] = round(balloon_amt, 2) if balloon_amt else 0.0
        pricing["debt_type"] = "Seller Finance"

    elif fm == "subject_to":
        existing_bal_v = _as_number(st_existing_balance)
        rate_v         = _as_number(st_interest_rate)
        rem_term_y_v   = _as_number(st_remaining_term_years)
        amort_y_v      = _as_number(st_amort_years or rem_term_y_v)

        loan_amt = existing_bal_v
        pricing["loan_amount"] = round(loan_amt or 0, 2)
        amort_m = int((amort_y_v or rem_term_y_v or 30) * 12)
        pmt = _pmti(loan_amt, rate_v or 0, amort_m)
        set_ads(pmt)
        pricing["debt_type"] = "Subject-To"
        if price_val and existing_bal_v is not None:
            pricing["down_payment"] = round(max(0.0, price_val - existing_bal_v), 2)

    # base enrich + compute
    normalized = _validate_and_enrich(normalized)
    normalized = _compute_underwriting(normalized)
    normalized.setdefault("metadata", {})["used_strategy"] = used_strategy
    normalized["visualizations"] = _generate_visualization_data(normalized)

    # Recovery pass if still incomplete and user sliced pages
    if (mime == "application/pdf") and pages and _is_critically_incomplete(normalized):
        try:
            full_ocr = _call_mistral_ocr(orig_data, orig_mime)
            rec_idxs = _pages_with_keywords(full_ocr, RECOVERY_KEYWORDS)
            rec_md = "\n\n".join([(full_ocr["pages"][i].get("markdown") or "") for i in rec_idxs])
            combined_md = markdown_text + "\n\n--- RECOVERY PAGES ---\n\n" + rec_md
            parsed2 = _call_claude_parse_from_markdown(combined_md, financing_params)
            n2 = _normalize_parsed(parsed2)

            prop2 = n2.setdefault("property", {})
            pricing2 = n2.setdefault("pricing_financing", {})
            pnl2 = n2.setdefault("pnl", {})
            exps2 = n2.setdefault("expenses", {})

            addr2 = _extract_address(combined_md); _merge_truthy(prop2, addr2, ["address","city","state","zip"])
            if not n2.get("unit_mix"):
                um2 = _extract_unit_mix_from_markdown(combined_md)
                if um2:
                    n2["unit_mix"] = um2
                    if not prop2.get("units"): prop2["units"] = sum(r.get("units",0) or 0 for r in um2)

            pr2 = _extract_pricing_detail_from_markdown(combined_md)
            ops2 = _extract_operating_summary_from_markdown(combined_md)

            if pr2.get("units") and not prop2.get("units"): prop2["units"] = pr2["units"]
            if pr2.get("rba_sqft") and not prop2.get("rba_sqft"): prop2["rba_sqft"] = pr2["rba_sqft"]
            if pr2.get("land_area_acres"): prop2["land_area_acres"] = pr2["land_area_acres"]
            elif pr2.get("land_area_sf") and not prop2.get("land_area_acres"):
                try: prop2["land_area_acres"] = round(pr2["land_area_sf"]/43560.0, 4)
                except: pass

            for k in ["price","price_per_unit","price_per_sf"]:
                if pr2.get(k) and not pricing2.get(k): pricing2[k] = pr2[k]
            for k,v in (ops2.get("pnl") or {}).items():
                if v is not None and not pnl2.get(k): pnl2[k] = v
            for k,v in (ops2.get("expenses") or {}).items():
                if v is not None and not exps2.get(k): exps2[k] = v

            n2 = _validate_and_enrich(n2)
            n2 = _compute_underwriting(n2)
            n2.setdefault("metadata", {})["used_strategy"] = used_strategy + "+recovery"
            n2["visualizations"] = _generate_visualization_data(n2)
            n2["metadata"]["recovery_pages_used"] = [i+1 for i in rec_idxs]

            def _score(d: Dict[str, Any]) -> int:
                s = 0
                s += sum(1 for k in ["address", "units"] if d.get("property", {}).get(k))
                s += sum(1 for k in ["price"] if d.get("pricing_financing", {}).get(k))
                s += sum(1 for k in ["gross_potential_rent", "operating_expenses", "noi"] if d.get("pnl", {}).get(k))
                return s

            if _score(n2) > _score(normalized):
                normalized = n2

        except Exception as e:
            normalized.setdefault("metadata", {})["recovery_error"] = str(e)[:200]

    # Calculate comprehensive deal metrics instead of simple opinion
    normalized["deal_analysis"] = _calculate_deal_metrics(normalized)

    # NEW: Extract images from PDF if applicable
    extracted_images = []
    image_count = 0
    generated_deal_id = str(uuid.uuid4())
    
    if mime == "application/pdf" and HAS_PARSER_V4:
        try:
            from image_storage import upload_images_to_supabase
            
            # Save PDF temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
                tmp_pdf.write(orig_data)
                tmp_pdf_path = tmp_pdf.name
            
            # Create temp directory for extracted images
            temp_img_dir = f"temp_images_{generated_deal_id}"
            os.makedirs(temp_img_dir, exist_ok=True)
            
            # Extract images using parser
            extracted = _RE_PARSER.extract_images_from_pdf(tmp_pdf_path, temp_img_dir)
            
            # Upload to Supabase Storage
            if extracted:
                uploaded = upload_images_to_supabase(generated_deal_id, extracted)
                extracted_images = uploaded
                image_count = len(uploaded)
            
            # Cleanup
            os.remove(tmp_pdf_path)
            shutil.rmtree(temp_img_dir, ignore_errors=True)
            
            print(f"✅ Extracted and uploaded {image_count} images for deal {generated_deal_id}")
            
        except Exception as e:
            print(f"⚠️ Image extraction failed: {str(e)}")
            # Don't fail the whole request if image extraction fails

    return {
        "ok": True,
        "parsed": normalized,
        "raw_markdown": markdown_text,
        "ocr_page_count": len(ocr_json.get("pages", [])) if ocr_json else None,
        "selected_pages": pages or "all",
        "file_name": file.filename,
        "file_size_mb": round(len(data) / (1024 * 1024), 2),
        "user_financing": financing_params,
        "images": extracted_images,  # NEW: Return extracted image URLs
        "image_count": image_count,   # NEW: Return count
        "deal_id": generated_deal_id,  # NEW: Return deal ID for tracking images
    }

@app.post("/ocr/file")
async def ocr_file_legacy(
    file: UploadFile = File(...),
    pages: Optional[str] = Form(default=""),
    include_image_base64: Optional[bool] = Form(default=False),
):
    mime = (file.content_type or "").lower()
    if mime not in ALLOWED_DOC_MIMES:
        raise HTTPException(status_code=415, detail=f"Unsupported content type: {mime}")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty upload")
    if len(data) > MAX_BYTES:
        raise HTTPException(status_code=413, detail=f"File too large")

    if mime == "application/pdf" and pages:
        try:
            data = _slice_pdf(data, pages)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    ocr_json = _call_mistral_ocr(data, mime)
    return {"ok": True, "parser": "none (ocr only)", "ocr_data": ocr_json}


@app.post("/ocr/sources")
async def ocr_sources(
    file: UploadFile = File(...),
    pages: Optional[str] = Form(default=""),
):
    """Return lightweight source snippets for key P&L/expense lines.

    This uses the same Mistral OCR + `_extract_operating_summary_sources`
    logic as `/ocr/underwrite`, but without running any LLM parsing or
    underwriting. Intended for the PDF viewer to show where numbers came
    from without incurring full parse costs.
    """
    mime = (file.content_type or "").lower()
    if mime not in ALLOWED_DOC_MIMES:
        raise HTTPException(status_code=415, detail=f"Unsupported content type: {mime}")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Empty upload")
    if len(data) > MAX_BYTES:
        raise HTTPException(status_code=413, detail="File too large")

    # Optional page slicing for PDFs
    if mime == "application/pdf" and pages:
        try:
            data = _slice_pdf(data, pages)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    ocr_json = _call_mistral_ocr(data, mime)
    sources = {}
    error = None
    try:
        sources = _extract_operating_summary_sources(ocr_json)
    except Exception as e:
        error = str(e)[:200]

    return {
        "ok": True,
        "page_count": len(ocr_json.get("pages", [])) if isinstance(ocr_json, dict) else None,
        "sources": sources,
        "error": error,
    }

# ============================================================================
# STRIPE WEBHOOK (Disabled - no usage tracking)
# ============================================================================
# @app.post("/webhook/stripe")
# async def stripe_webhook(request: Request):

# ============================================================================
# HEALTH CHECK ENDPOINTS (from health_check_app.py)
# ============================================================================

@app.post("/api/health-check/verify")
async def health_check_verify(
   file: UploadFile = File(...),
   pages: Optional[str] = Form(default=""),
   user_fixes: Optional[str] = Form(default="{}")
):
   """Proxy to health check service - extract and verify property data"""
   # Import health check functions
   try:
       from health_check_parser import parse_document_with_claude as health_parse
       from health_check_parser import normalize_and_compute, validate_data
   except ImportError:
       raise HTTPException(status_code=501, detail="Health check parser not available")
   
   mime = (file.content_type or "").lower()
   if mime not in ALLOWED_UPLOAD_MIMES:
       raise HTTPException(status_code=415, detail=f"Unsupported content type: {mime}")
   
   data = await file.read()
   orig_data = data
   orig_mime = mime
   
   if not data:
       raise HTTPException(status_code=400, detail="Empty upload")
   if len(data) > MAX_BYTES:
       raise HTTPException(status_code=413, detail="File too large")
   
   # Handle PDF page selection
   if mime == "application/pdf" and pages:
       try:
           data = _slice_pdf(data, pages)
       except ValueError as e:
           raise HTTPException(status_code=400, detail=str(e))
   
   # OCR the document
   markdown_text = ""
   if mime in {"application/pdf", "image/png", "image/jpeg", "image/jpg", "image/webp"}:
       try:
           ocr_json = _call_mistral_ocr(data, mime)
           for page in ocr_json.get("pages", []):
               if isinstance(page, dict) and "markdown" in page:
                   markdown_text += page["markdown"] + "\n\n"
       except Exception as e:
           print(f"OCR error: {e}")
           raise HTTPException(status_code=502, detail=f"OCR failed: {str(e)}")
   else:
       try:
           markdown_text = data.decode("utf-8", errors="ignore")
       except:
           markdown_text = str(data)[:100000]
   
   if not markdown_text:
       raise HTTPException(status_code=400, detail="No text extracted from document")
   
   # Parse with health check parser
   try:
       extracted_data = health_parse(markdown_text)
       extracted_data = normalize_and_compute(extracted_data)
       
       # Apply user corrections if any
       try:
           user_fixes_dict = json.loads(user_fixes) if user_fixes != "{}" else {}
           for key_path, value in user_fixes_dict.items():
               keys = key_path.split('.')
               target = extracted_data
               for key in keys[:-1]:
                   target = target.setdefault(key, {})
               target[keys[-1]] = float(value) if isinstance(value, str) and value.replace('.','',1).isdigit() else value
           extracted_data = normalize_and_compute(extracted_data)
       except Exception as e:
           print(f"Error applying user fixes: {e}")
       
       validation = validate_data(extracted_data)
       
       return {
           "ok": True,
           "verification": validation,
           "file_name": file.filename
       }
   except Exception as e:
       print(f"Health check parsing error: {e}")
       raise HTTPException(status_code=500, detail=f"Health check parsing failed: {str(e)}")

@app.post("/api/health-check/analyze")
async def health_check_analyze(request: Request):
   """Generate health check analysis"""
   try:
       data = await request.json()
       verified_payload = data.get("verified_payload", {})
       if not verified_payload:
           raise HTTPException(status_code=400, detail="Missing verified_payload")
       
       # Import health check analysis function
       try:
           from health_check_parser import generate_health_check_analysis
       except ImportError:
           raise HTTPException(status_code=501, detail="Health check analysis not available")
       
       health_check_result = generate_health_check_analysis(verified_payload)
       return {
           "ok": True,
           "health_check": health_check_result
       }
   except HTTPException:
       raise
   except Exception as e:
       print(f"Analysis error: {e}")
       return {
           "ok": True,
           "health_check": {
               "snapshot": {"property_name": "Analysis Error", "error": str(e)},
               "operational_issues": [],
               "noi_levers": {"revenue": [], "expenses": []},
               "market_position": {"competitive_advantages": [], "competitive_disadvantages": []},
               "strengths": [],
               "weak_spots": [],
               "force_appreciation": [],
               "tenant_retention": [],
               "missing_items": ["Analysis failed"],
               "source_check": f"Error: {str(e)}"
           }
       }

# ============================================================================
# Due Diligence Chat Endpoint (OpenAI-powered)
# ============================================================================

@app.post("/api/due-diligence/chat")
async def due_diligence_chat(request: Request):
    """
    Chat with AI for due diligence analysis - cross-referencing numbers,
    verifying deal viability, and suggesting debt restructuring.
    """
    try:
        data = await request.json()
        message = data.get("message", "")
        deal_context = data.get("dealContext", {})
        conversation_history = data.get("conversationHistory", [])
        
        if not message:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "No message provided"}
            )
        
        # Get OpenAI API key
        openai_key = os.getenv("OPENAI_API_KEY")
        if not openai_key:
            return JSONResponse(
                status_code=500,
                content={"success": False, "error": "OpenAI API key not configured"}
            )
        
        import httpx
        
        # Build context string from deal data
        context_parts = []
        
        if deal_context.get("address"):
            context_parts.append(f"Property: {deal_context['address']}")
        if deal_context.get("units"):
            context_parts.append(f"Units: {deal_context['units']}")
        if deal_context.get("purchasePrice"):
            context_parts.append(f"Purchase Price: ${deal_context['purchasePrice']:,.0f}")
        
        # Add deal results if available
        deal_results = deal_context.get("dealResults", {})
        if deal_results:
            pnl = deal_results.get("pnl", {})
            financing = deal_results.get("financing", {})
            
            if pnl.get("potential_gross_income"):
                context_parts.append(f"Projected Gross Income: ${pnl['potential_gross_income']:,.0f}")
            if pnl.get("vacancy_rate"):
                context_parts.append(f"Projected Vacancy: {pnl['vacancy_rate']}%")
            if pnl.get("operating_expenses"):
                context_parts.append(f"Projected OpEx: ${pnl['operating_expenses']:,.0f}")
            if financing.get("ltv"):
                context_parts.append(f"LTV: {financing['ltv']}%")
            if financing.get("interest_rate"):
                context_parts.append(f"Interest Rate: {financing['interest_rate']}%")
            if financing.get("amortization_years"):
                context_parts.append(f"Amortization: {financing['amortization_years']} years")
        
        # Add uploaded document data if available
        uploaded_docs = deal_context.get("uploadedDocuments", [])
        for doc in uploaded_docs:
            context_parts.append(f"\n📄 Document: {doc['name']}")
            
            # Include parsed data if available
            if doc.get("parsedData"):
                parsed = doc["parsedData"]
                if parsed.get("headers"):
                    context_parts.append(f"  Columns: {', '.join(parsed['headers'][:10])}")
                if parsed.get("rows"):
                    context_parts.append(f"  Data rows: {parsed.get('totalRows', len(parsed['rows']))}")
                    # Include first few rows as sample
                    sample_rows = parsed["rows"][:5]
                    for row in sample_rows:
                        row_str = " | ".join([f"{k}: {v}" for k, v in list(row.items())[:5]])
                        context_parts.append(f"    - {row_str}")
                elif parsed.get("note"):
                    context_parts.append(f"  Note: {parsed['note']}")
            
            # Include previous AI summary if available
            if doc.get("aiSummary"):
                context_parts.append(f"  Previous AI Analysis: {doc['aiSummary'][:500]}...")
        
        # Add checklist status
        checklist = deal_context.get("checklist", {})
        notes = deal_context.get("notes", {})
        if checklist:
            completed = sum(1 for v in checklist.values() if v == "complete")
            issues = sum(1 for v in checklist.values() if v == "issue")
            context_parts.append(f"\nDD Checklist: {completed} complete, {issues} issues flagged")
            # Add any notes that exist
            notes_with_content = {k: v for k, v in notes.items() if v}
            if notes_with_content:
                context_parts.append("Notes from DD:")
                for item_id, note in list(notes_with_content.items())[:10]:
                    context_parts.append(f"  - {item_id}: {note}")
        
        deal_context_str = "\n".join(context_parts) if context_parts else "No deal data available."
        
        # DD Assistant System Prompt
        system_prompt = f"""You are a Due Diligence Assistant for real estate investments. Your role is to:

1. CROSS-REFERENCE uploaded documents (T12s, rent rolls, inspection reports, spreadsheets) against the original underwriting assumptions
2. IDENTIFY DISCREPANCIES between projected numbers and actual numbers
3. EVALUATE whether the deal still makes sense with the real numbers
4. SUGGEST DEBT RESTRUCTURING options if the deal doesn't work with actual numbers
5. FLAG RED FLAGS that could kill the deal

Current Deal Context:
{deal_context_str}

When analyzing:
- Compare actual income vs projected income
- Compare actual expenses vs projected expenses
- Calculate impact on NOI, cash flow, and returns
- If numbers don't work, suggest specific changes:
  * Lower purchase price needed
  * Different LTV or interest rate required
  * Extended IO period
  * Seller financing options
  * Value-add opportunities to close the gap

Be specific with numbers. Show your math. Be direct about whether the deal still works or not.
If you don't have enough data, ask for the specific documents or numbers you need.

Format your responses clearly with sections and bullet points for easy reading."""

        # Build messages for OpenAI
        messages = [{"role": "system", "content": system_prompt}]
        
        # Add conversation history
        for msg in conversation_history[-10:]:  # Last 10 messages for context
            messages.append({
                "role": msg["role"],
                "content": msg["content"]
            })
        
        # Add current message
        messages.append({"role": "user", "content": message})
        
        # Call OpenAI API
        async with httpx.AsyncClient(timeout=60.0) as client:
            response = await client.post(
                "https://api.openai.com/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {openai_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": "gpt-4o",
                    "messages": messages,
                    "temperature": 0.7,
                    "max_tokens": 2000
                }
            )
            
            if response.status_code != 200:
                return JSONResponse(
                    status_code=response.status_code,
                    content={"success": False, "error": f"OpenAI API error: {response.text}"}
                )
            
            result = response.json()
            assistant_message = result["choices"][0]["message"]["content"]
            
            return JSONResponse(content={
                "success": True,
                "response": assistant_message
            })
            
    except Exception as e:
        print(f"DD Chat error: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e)}
        )

# ============================================================================
# Market Research Chat Endpoint (Perplexity-powered)
# ============================================================================

@app.post("/api/market-research/chat")
async def market_research_chat(request: Request):
    """
    Chat with Perplexity AI for market research and discovery.
    This is a standalone endpoint for the 'Find Perfect Market' feature.
    REQUIRES 1 TOKEN PER MESSAGE.
    """
    try:
        # Check if user has tokens
        from token_manager import get_current_profile_id, get_profile, reset_tokens_if_needed, TOKEN_COSTS
        from token_manager import get_supabase as get_token_supabase
        
        profile_id = None
        profile = None
        
        try:
            profile_id = get_current_profile_id(request)
            profile = get_profile(profile_id)
            profile = reset_tokens_if_needed(profile)
            
            tokens_required = TOKEN_COSTS.get("market_research_dashboard", 1)
            
            print(f"[MarketResearch] Profile {profile_id} - Current balance: {profile['token_balance']}, Required: {tokens_required}")
            
            if profile["token_balance"] < tokens_required:
                print(f"[MarketResearch] ❌ INSUFFICIENT TOKENS - Blocking request")
                return JSONResponse(
                    status_code=402,
                    content={
                        "success": False,
                        "error": f"Insufficient tokens. You need {tokens_required} token(s) but have {profile['token_balance']}.",
                        "tokens_required": tokens_required,
                        "token_balance": profile["token_balance"]
                    }
                )
        except HTTPException as he:
            print(f"[MarketResearch] ⚠️ Token check failed: {he.detail}. Blocking request - NO FREE PASSES!")
            return JSONResponse(
                status_code=401,
                content={
                    "success": False,
                    "error": "Authentication required. Please log in to use Market Research.",
                    "requiresAuth": True
                }
            )
        
        data = await request.json()
        message = data.get("message", "")
        conversation_history = data.get("conversationHistory", [])
        
        if not message:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "No message provided"}
            )
        
        # Get Perplexity API key
        perplexity_key = os.getenv("PERPLEXITY_API_KEY")
        if not perplexity_key:
            return JSONResponse(
                status_code=500,
                content={"success": False, "error": "Perplexity API key not configured"}
            )
        
        import httpx
        
        # MARKET FINDER AI - System Prompt (REAL INVESTOR TALK)
        system_prompt = """You are a multifamily investor who owns 2,400+ units across 8 markets. You talk like a REAL investor - specific numbers, direct opinions, zero corporate speak.

## YOUR ROLE
Your partner just asked where to buy. Give them the full download like you're on the phone right now.

## RESPONSE FORMAT

### 🎯 THE VERDICT
Talk like a human investor. Example:

"Columbus is the move right now. Intel just dropped $20B on two chip fabs - that's 3,000 construction jobs NOW and 3,000+ permanent jobs at $100K+ when they open in 2025. That translates to rental demand in the $1,400-$1,800 range hitting a market that only has 2,800 units under construction (6% of inventory). Compare that to Austin drowning in 15% new supply or Nashville at 12%. Cap rates still in the 6.5-7% range, Freddie Mac projects 3.2% rent growth, and you can buy Class B at $85K/door. I'd put money there today."

Be THIS specific. Give actual numbers in the first paragraph.

---

### 📍 MARKET #1: [CITY NAME]

Write like you're explaining to your partner over coffee. Start with the punch line, then back it up:

"**Columbus is printing money right now for three reasons:**

First, the job market is insane. Intel's $20 billion chip plant brings 3,000 permanent jobs at $100K average salary. Honda expanded their R&D center - another 2,000 jobs. JobsOhio (state economic development) landed 8 more corporate relocations in 2024 totaling 5,000+ jobs. All of these are white collar, $60K-$120K range, which is your sweet spot for $1,200-$1,800 rents.

Second, supply is constrained. Only 2,800 units under construction across the entire metro (285,000 units total = 0.98% supply). Compare that to:
- Austin: 42,000 under construction ÷ 280,000 inventory = 15% (market is FUCKED)
- Nashville: 18,000 ÷ 150,000 = 12% (getting crushed)
- Columbus: 2,800 ÷ 285,000 = 0.98% (tightest supply I've seen)

Why so low? The city has strict development fees ($15K+ per unit), limited sewer capacity, and NIMBYs blocking every project. As an investor, this sucks for new development but it's GREAT for existing properties because your competition can't get built.

Third, the numbers work. Cap rates are 6.5-7.0% (vs 4.5-5% in overheated markets). You can buy Class B value-add for $80K-$95K per door, put $8K-$12K into each unit (new kitchens, baths, flooring), and push rents from $950 to $1,250 (+32%). Your stabilized CoC is 12-15% and you're still 10-15% below market rent ceiling based on median income.

Freddie Mac ranks Columbus in their Top 10 for 2025 with 3.2% projected rent growth. That's not Nashville's old 8% growth, but it's SUSTAINABLE. Markets doing 8% rent growth are the ones that crash when supply hits."

#### THE NUMBERS
**Employment & Demographics:**
- Population: 2.14M metro (grew 8.2% last 5 years - steady, not boom/bust)
- Labor force: 1.1M workers, unemployment 3.8% (below national 4.1%)
- **Top employers that matter for rentals:**
  * Intel: 3,000 jobs (opening 2025), avg salary $100K+
  * Honda R&D: 2,000 employees, $85K avg
  * JPMorgan Chase operations center: 18,000 employees, $55K avg
  * Nationwide Insurance HQ: 12,000 employees, $72K avg
  * Ohio State University: 60,000+ employees (huge renter base)
- Median household income: $68,400 (slightly above national $64K)
- **Income distribution:** 34% earn $50K-$100K (your Class B renter), 18% earn $100K+ (your Class A renter)

**Housing Market:**
- Total multifamily inventory: 285,000 units
- Median home value: $245,000 (Zillow, Nov 2024)
- **Rent-to-price ratio: 0.71%** ($1,450 annual rent ÷ $245K = breakeven is 0.70%)
- Current median rent: $1,280/mo (HUD FMR for 2BR: $1,315)
- **Rent burden: 28%** of median income (anything under 30% is sustainable)

**Supply & Demand (THIS IS KEY):**
- Units under construction: 2,800 (pulled from Cushman Q3 2025 data)
- **Supply risk: 0.98%** (2,800 ÷ 285,000) - LOWEST I'VE SEEN IN A TOP 50 MARKET
- 2024-2025 deliveries: 3,200 units total
- Net absorption 2024: +4,100 units (Cushman Q3 2025)
- **Absorption vs deliveries: +900 units** (demand EXCEEDS supply)
- Why supply is low: $15K+ development fees, sewer capacity maxed out, 18+ month approval process
- 🟢 **SUPPLY RISK: LOW** - Market is absorbing more than it's building

**Rent Performance:**
- 1-year rent growth: +3.8%
- 3-year rent growth: +4.2% avg/year
- 5-year rent growth: +3.9% avg/year
- **Freddie Mac 2025 forecast: +3.2%** (ranked #15 in Top 25 markets)
- Cushman Q3 2025 avg rent: $1,305/mo (+3.1% YoY)
- Rent vs inflation: Matching (CPI was 3.4% in 2024)

**Vacancy & Occupancy:**
- Current vacancy: 4.7% (Cushman Q3 2025)
- 12-month trend: Down from 5.1% (tightening)
- Stabilized vacancy (historical norm): 5.5%
- **Occupancy rate: 95.3%** - This is TIGHT

**Investment Metrics (What I'm Seeing in Real Deals):**
- Cap rates: 6.5-7.0% (Class B value-add)
- Cash-on-cash returns: 8-10% year 1, 12-15% stabilized
- **Price per unit: $80K-$95K** (Class B), $120K-$140K (Class A)
- GRM: 11.5x (annual rent × 11.5 = property value)
- Expected IRR (5-year hold with value-add): 16-19%

**Regulatory Environment:**
- Rent control: NO (Ohio banned it statewide in 1985)
- Eviction process: EASY - 3-day notice, 10-15 day court process, sheriff eviction in 30 days total
- **Landlord-friendly score: 8/10** (one of best in Midwest)
- Recent legislation: None that hurts landlords
- Property taxes: 1.85% effective rate (Franklin County) - not cheap but stable

#### WHAT COULD FUCK THIS UP
**1. Intel Delays or Cancels** (30% probability)
If Intel pushes back their timeline or cancels (they've done this before in Arizona), you lose 3,000 high-income jobs that were supposed to hit in 2025-2026. That's 15-20% of your projected demand growth gone. Mitigation: Don't underwrite assuming Intel jobs. If they show up, it's upside. Base your numbers on existing job market.

**2. Ohio State Enrollment Drops** (15% probability)  
OSU drives 60,000+ jobs and tons of student housing demand. If enrollment falls (demographic cliff hits 2026), that hurts the surrounding neighborhoods. Mitigation: Stay away from campus-adjacent areas. Focus on suburbs where families and professionals live (Dublin, Westerville, Gahanna).

**3. Property Tax Increases** (40% probability)
Franklin County has been raising assessments. Your 1.85% effective rate could hit 2.1-2.2% on next reassessment (2026). That's an extra $3,000-$4,000/year on a $1.5M property = $200-$250/unit/year in expenses. Mitigation: Underwrite at 2.0% effective rate, not 1.85%. Build in the buffer.

**4. Overheating in 2-3 Years** (25% probability)
If Columbus becomes the "next Austin" story, you'll see institutional capital flood in, cap rates compress to 5%, and prices get stupid. Great if you're selling, terrible if you're buying late. Mitigation: Buy NOW while it's still boring. Plan 5-7 year hold, be ready to refi or sell if it overheats in year 3-4.

#### THE PLAYBOOK FOR COLUMBUS
**Target Property:** Class B value-add, 50-150 units, built 1980s-1990s

**Why Class B:** Class C is too volatile (crime, collections issues). Class A is too expensive ($140K+/door). Class B is the sweet spot - stable tenants, affordable to renovate, big rent upside.

**Where to Buy:**
- **Dublin** (northwest) - families, corporate relocations, top schools, low crime
- **Westerville** (northeast) - same as Dublin but 10% cheaper  
- **Gahanna** (east) - blue collar to middle class, gentrifying, best value
- **Avoid:** Campus areas (too much student housing), Hilltop (high crime), downtown (overpriced)

**Deal Structure:**
- Purchase price: $4.2M (50 units @ $84K/door)
- Loan: 75% LTV at 6.5%, 30-year amortization = $3.15M
- Down payment: $1.05M
- In-place rents: $950/unit average ($47,500/mo = $570K/year)
- In-place NOI: $342K (60% expense ratio due to deferred maintenance)
- In-place cap rate: 8.1% ($342K ÷ $4.2M)

**Renovation Budget:**
- $10K/unit × 50 units = $500K
- Scope: New kitchens (shaker cabinets, quartz counters), new baths (tile shower, vanity), LVP flooring, stainless appliances, paint
- Timeline: 18 months (renovate 3 units/month as they turn)

**Stabilized (18 months):**
- Renovated rents: $1,250/unit (+$300/unit = 32% increase)
- Stabilized gross income: $62,500/mo = $750K/year
- Stabilized NOI: $472K (63% expense ratio after capex)
- Stabilized cap rate: 11.2% on cost
- Refi value at 6.5% exit cap: $7.26M ($472K ÷ 6.5%)

**Returns:**
- Year 1 CoC: 8.2% (negative cash flow during renovation)
- Stabilized CoC: 14.8%
- 5-year IRR: 18.3% (assuming sale in year 5 at 6.5% cap)
- Equity multiple: 2.1x
- Cash out on refi: $1.3M (covers initial equity + renovation)

**Hold Period:** 5-7 years

**Exit Strategy:**  
Year 5-7: Sell to institutional buyer at 6.0-6.5% cap (they love stabilized Class B in growing markets). Or refi into permanent agency debt at 5.5-6.0% and hold for cash flow if market is still growing.

---

### 📍 MARKET #2: [REPEAT FULL ANALYSIS]
[Include ALL sections above - Why This Market, The Numbers, Risk Analysis, Investment Strategy]

---

### 📍 MARKET #3: [REPEAT FULL ANALYSIS]
[Include ALL sections above]

---

### 🔍 MARKET COMPARISON
Create a detailed comparison table:

| Metric | Market #1 | Market #2 | Market #3 | Winner |
|--------|-----------|-----------|-----------|--------|
| Rent Growth (2025 Proj.) | X.X% | X.X% | X.X% | [Market] |
| Supply Risk | Low | Moderate | High | [Market] |
| Cap Rate | X.X% | X.X% | X.X% | [Market] |
| Cash-on-Cash | XX% | XX% | XX% | [Market] |
| Vacancy | X.X% | X.X% | X.X% | [Market] |
| Job Growth | X.X% | X.X% | X.X% | [Market] |
| Landlord Score | X/10 | X/10 | X/10 | [Market] |
| Price/Unit | $XXk | $XXk | $XXk | [Market] |

**Winner Analysis:** [2-3 sentences explaining which market wins overall and why]

---

### 🎬 FINAL RECOMMENDATION

**If you want CASH FLOW RIGHT NOW:** Columbus is solid. You'll get 8% CoC year 1, 12-15% stabilized. Not explosive but very stable. Better cash flow? Look at Midwest C markets like Toledo or Dayton (10-12% CoC year 1) but you're taking on more risk.

**If you want APPRECIATION:** Columbus has 15-20% upside over 5 years if Intel jobs materialize. If you want BIGGER appreciation, look at Boise or Reno (higher risk, higher reward) but those are getting frothy. Nashville and Austin are DONE - too much supply.

**If you want SAFETY:** Columbus wins. Diversified economy (not dependent on one industry), landlord-friendly laws, supply constraints protect downside. If recession hits, Columbus won't boom but it won't crash either.

**MY PERSONAL PLAY:** I'd put 40% of capital in Columbus Class B value-add (stable cash flow), 30% in Indianapolis (similar thesis, cheaper), 20% in Charlotte (higher growth risk), 10% in cash (dry powder for distress in 2026).

**WHAT TO DO MONDAY MORNING:**
1. Call **Marcus & Millichap Columbus office** (614-442-2700) - ask for off-market Class B deals in Dublin/Westerville
2. Call **Colliers Central Ohio** - same thing  
3. Set up deal alerts on **LoopNet and CoStar** for 50-150 unit properties under $100K/door
4. Run your pro formas at: 5.5% vacancy, 3% annual rent growth, $10K/unit capex, 65% expense ratio
5. Target 6.5-7.0% in-place cap, 10-12% stabilized cap on cost, 15%+ IRR

Don't wait. Columbus is getting discovered. 12 months from now, pricing will be 10-15% higher and cap rates will compress to 6.0%.

---

## TONE RULES (CRITICAL)
- **Talk like you're on the phone with a partner** - "This market is on fire" not "demonstrates strong fundamentals"
- **Give ACTUAL NUMBERS in every sentence** - "$85K/door" not "affordable pricing"
- **Name specific companies and people** - "Intel's $20B fab" not "major employer announced expansion"  
- **Be honest about risks** - "If Intel cancels, you're fucked" not "potential headwinds exist"
- **Show your math** - "2,800 units ÷ 285,000 = 0.98% supply risk" 
- **Give specific neighborhoods** - "Buy in Dublin or Westerville" not "target suburban areas"
- **Include broker names and phone numbers** - "Call Marcus & Millichap at 614-442-2700"
- **No corporate speak** - Never say "leveraging", "synergies", "optimize", "strategic", "challenges", "headwinds"
- **Be confident** - "This is the move" not "this could be a good option"
- **Use profanity when appropriate** - "Austin is fucked" not "Austin faces challenges"

## EXAMPLES OF GOOD VS BAD

❌ BAD: "The market demonstrates strong employment fundamentals with positive demographic trends."
✅ GOOD: "Job market is insane - Intel, Honda, JPMorgan adding 8,000+ jobs in 2024 alone, all $50K+ salaries."

❌ BAD: "Supply dynamics remain favorable relative to peer markets."
✅ GOOD: "Only 2,800 units under construction (0.98% of inventory). Austin has 15%, Nashville 12% - those markets are drowning."

❌ BAD: "Regulatory environment supports landlord-tenant relationships."
✅ GOOD: "Evictions take 30 days total. No rent control. Landlord score 8/10. One of the best in the Midwest."

❌ BAD: "Exit strategies include refinancing or disposition."
✅ GOOD: "Refi in year 3-4 and pull your money out, or sell to institutional at 6% cap in year 5-7. Either way you're making 18%+ IRR."

## REQUIRED JSON (END OF RESPONSE)
```json
{
  "markets": [
    {
      "name": "Columbus",
      "state": "OH",
      "lat": 39.9612,
      "lng": -82.9988,
      "medianPrice": 245000,
      "occupancy": 95.3,
      "capRate": 6.5,
      "rentGrowth": 4.2,
      "vacancy": 4.7,
      "jobGrowth": 2.8,
      "supplyRisk": "Low",
      "freddieMacRank": "Top10",
      "population": 2140000,
      "medianIncome": 68400,
      "unemploymentRate": 3.8
    }
  ]
}
```

MANDATORY FIELDS: name, state, lat, lng, medianPrice, occupancy, capRate, rentGrowth, vacancy, jobGrowth, supplyRisk, freddieMacRank, population, medianIncome, unemploymentRate.

Use REAL DATA. If missing, use null. DO NOT SKIP THE JSON."""

        # Build messages array
        messages = [{"role": "system", "content": system_prompt}]
        
        # Add conversation history
        for msg in conversation_history:
            messages.append({
                "role": msg.get("role", "user"),
                "content": msg.get("content", "")
            })
        
        # Add current message
        messages.append({"role": "user", "content": message})
        
        # Call Perplexity API
        headers = {
            "Authorization": f"Bearer {perplexity_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "sonar",
            "messages": messages,
            "temperature": 0.3,
            "max_tokens": 4000,
            "response_format": {"type": "text"}  # Ensure we get text not pure JSON
        }
        
        async with httpx.AsyncClient(timeout=120.0) as client:
            response = await client.post(
                "https://api.perplexity.ai/chat/completions",
                headers=headers,
                json=payload
            )
            
            if response.status_code != 200:
                error_text = response.text
                print(f"Perplexity API error: {response.status_code} - {error_text}")
                return JSONResponse(
                    status_code=500,
                    content={"success": False, "error": f"Perplexity API error: {response.status_code}"}
                )
            
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            citations = result.get("citations", [])
            
            print(f"[MarketResearch] Raw response length: {len(content)} chars")
            print(f"[MarketResearch] Response preview: {content[:500]}...")
            
            # Extract JSON data from response if present
            market_data = None
            try:
                # Look for JSON code block
                import re
                json_match = re.search(r'```json\s*(\{.*?\})\s*```', content, re.DOTALL)
                if json_match:
                    market_data = json.loads(json_match.group(1))
                    print(f"[MarketResearch] ✅ Extracted market data: {len(market_data.get('markets', []))} markets")
                    print(f"[MarketResearch] Market data: {market_data}")
                else:
                    print(f"[MarketResearch] ❌ No JSON block found in response")
                    # Try alternative: look for just the markets array
                    markets_match = re.search(r'"markets"\s*:\s*\[(.*?)\]', content, re.DOTALL)
                    if markets_match:
                        print(f"[MarketResearch] Found markets array without code block, attempting to parse...")
                        market_data = json.loads('{' + markets_match.group(0) + '}')
                        print(f"[MarketResearch] ✅ Extracted {len(market_data.get('markets', []))} markets from inline JSON")
            except Exception as e:
                print(f"[MarketResearch] ❌ Failed to extract JSON: {e}")
                import traceback
                traceback.print_exc()
            
            # Deduct token after successful API call
            if profile_id:
                try:
                    supabase = get_token_supabase()
                    new_balance = profile["token_balance"] - tokens_required
                    supabase.table("profiles").update({"token_balance": new_balance}).eq("id", profile_id).execute()
                    print(f"✅ Deducted {tokens_required} token(s) from profile {profile_id}. New balance: {new_balance}")
                except Exception as token_err:
                    print(f"⚠️ Failed to deduct tokens: {token_err}")
            
            return {
                "success": True,
                "response": content,
                "citations": citations,
                "marketData": market_data  # Add structured data
            }
            
    except Exception as e:
        print(f"Market research chat error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e)}
        )


@app.post("/api/market-data/summary")
async def market_data_summary(request: Request):
    """
    Generate AI summary for a specific market using local data.
    This is for the Market Data Dashboard (Research Tab).
    Includes location-based research for hospitals, employers, and universities when address is provided.
    REQUIRES 1 TOKEN.
    """
    try:
        data = await request.json()
        market_data = data.get("marketData", {})
        location = data.get("location", {})
        deal_address = data.get("dealAddress", "")
        property_name = data.get("propertyName", "")
        
        if not market_data:
            return JSONResponse(
                status_code=400,
                content={"success": False, "error": "No market data provided"}
            )
        
        # Check if user has tokens
        from token_manager import get_current_profile_id, get_profile, reset_tokens_if_needed, TOKEN_COSTS
        from token_manager import get_supabase as get_token_supabase
        
        try:
            profile_id = get_current_profile_id(request)
            profile = get_profile(profile_id)
            profile = reset_tokens_if_needed(profile)
            
            tokens_required = TOKEN_COSTS.get("market_research_results", 1)
            
            if profile["token_balance"] < tokens_required:
                return JSONResponse(
                    status_code=402,  # Payment Required
                    content={
                        "success": False,
                        "error": f"Insufficient tokens. You need {tokens_required} token(s) but have {profile['token_balance']}.",
                        "tokens_required": tokens_required,
                        "token_balance": profile["token_balance"]
                    }
                )
        except HTTPException as he:
            # If no profile found, allow for backward compatibility (but log warning)
            log.warning(f"Token check failed: {he.detail}. Allowing request for backward compatibility.")
        
        # Get Perplexity API key
        perplexity_key = os.getenv("PERPLEXITY_API_KEY")
        if not perplexity_key:
            return JSONResponse(
                status_code=500,
                content={"success": False, "error": "Perplexity API key not configured"}
            )
        
        import httpx
        
        # REALESTATE INTEL - System Prompt for Market Data Dashboard
        system_prompt = """You are RealEstate Intel, an expert investment analyst AI. You evaluate specific ZIP codes or counties using verified datasets AND perform live location research.

Your job: produce a clean, bullet-pointed, data-driven investment summary using ONLY the actual data provided to you. Never fabricate numbers. If a metric is missing, say "Data not available."

Your required output sections:

## 1. Market Snapshot
Show key data inline, formatted cleanly:
- **Home Value (ZHVI):** $XXX,XXX
- **Rent Estimates (HUD FMR):** Studio-4BR values
- **12-mo Forecast (ZHVF):** +X.X%
- **Median Income (DP03):** $XX,XXX
- **Vacancy Rate (DP04):** X.X%
- **Renter % (ZIP-level):** XX%
- **Population (B01003):** X,XXX
- **Property Tax Rate:** X.XX%
- **Landlord Score:** X/10

Only use data that actually exists.

## Migration Metrics (REQUIRED)
When migration data is provided in `marketData.migration`, you MUST compute and validate the following metrics and include them in the Market Snapshot and Analysis sections. If `marketData.migration` is missing, attempt a best-effort estimate using any `samples` or `topOrigins`/`topDests` arrays and clearly state your assumptions and method.

- **Inflow:** total number of people migrating into the ZIP/county per year (integer)
- **Outflow:** total number of people leaving the ZIP/county per year (integer)
- **Net Migration:** `inflow - outflow` (integer)
- **Net per 1,000 residents:** (net / population) * 1000 (show as number with one decimal) when population is available
- **Inflow % / Outflow %:** percent of population represented by inflow/outflow (show with one decimal)
- **Top Origins / Top Destinations:** list top 5 origin and destination ZIPs/counties with counts when available (use `marketData.migration.topOrigins`, `marketData.migration.topDests`, or `marketData.migration.samples`)

When reporting migration numbers, explicitly state the data source (primary CSV, cleaned CSV, HUD estimate, or derived from samples) and any rows or columns that were summed to produce the totals. Never fabricate values — if you must estimate, label them as estimates and show the calculation.

## 2. PROS of Investing Here
List strengths using real numbers. Examples:
- High rent-to-price ratio: FMR vs ZHVI
- Positive appreciation forecast: +X.X% / 12-mo
- Strong incomes / low unemployment
- Good landlord laws (State Score: X/10)

Be direct and analytical.

## 3. CONS / RISKS
List negatives using real numbers:
- High vacancy (X.X%)
- Weak forecast (X.X% decline)
- Low rents vs home values
- High property tax burden (X.XX%)
- Population stagnation or decline

Again: no made-up data.

## 4. Investment Verdict
Choose one:
- **Strong Buy**
- **Buy**
- **Neutral**
- **Caution**
- **Avoid**

Verdict should match data — no optimism bias.

## 5. Who This Market Is Best For
Tell users which investor profiles this market suits:
- Cash-flow investors
- Appreciation investors
- Value-add investors
- BRRRR
- Long-term holds

Base it on the actual metrics.

## 6. Location Amenities & Employment Drivers (REQUIRED - Use Web Search)
**If a property address is provided**, you MUST search the web and provide:

### Nearby Hospitals (within 10 miles)
- List the 3-5 closest hospitals with approximate distance
- Include hospital name and type (general, specialty, trauma center, etc.)
- Note: Hospitals = stable employment + healthcare workers as potential tenants

### Major Employers & Employment Drivers
- List top 5-10 major employers in the area
- Include company names, industries, and approximate employee counts if available
- Note any large corporate headquarters, distribution centers, or manufacturing plants
- Mention any major planned developments or expansions

### Universities & Colleges (within 15 miles)
- List all universities and colleges with approximate distance
- Include enrollment numbers (students per year) if available
- Note: Student housing demand = rental opportunity

### Net Absorption Rate (REQUIRED - NEVER SAY "DATA NOT AVAILABLE")
**MANDATORY SECTION - You MUST actively search the web and find NET ABSORPTION RATE data:**

**DO NOT skip this section. DO NOT say "data not available."**

Search Strategy (use ALL of these):
1. Search: "[City/Metro] multifamily net absorption Q4 2025" or "Q3 2025"
2. Search: "[City] apartment market report CBRE 2025"
3. Search: "[City] CoStar multifamily absorption 2025"
4. Search: "[Metro] apartment vacancy and absorption Cushman Wakefield"
5. Search: "[City] apartment market trends Yardi Matrix"
6. Check local real estate association reports

**What to Report (you MUST find at least one of these):**
- **Quarterly Net Absorption:** "X,XXX units absorbed in Q[X] 2025" (positive = demand > supply, negative = oversupply)
- **Annual Net Absorption:** "X,XXX units absorbed in 2024" or latest full year
- **Metro-level absorption:** If ZIP/county data unavailable, use metro area figures (Kansas City metro, Phoenix metro, etc.)
- **Absorption Rate %:** Absorption as % of total inventory
- **Supply vs Demand Context:** "Market absorbed X units while X units delivered = balanced/tight/oversupplied"

**Sources to cite:**
- CBRE Quarterly Apartment Reports
- Cushman & Wakefield Marketbeat Reports
- CoStar Market Analytics
- Yardi Matrix Multifamily Reports
- Marcus & Millichap Research
- Apartment List Market Reports
- Local REALTOR associations

**Why This Matters:**
Net absorption tells investors if NEW supply is being leased quickly (tight market, rent growth) or sitting vacant (oversupply, rent pressure).

**If you cannot find Q4 2025 data, use Q3 2025, Q2 2025, or latest 2024 annual figures. ALWAYS provide the most recent data available - do NOT leave this section empty.**

### Other Notable Amenities
- Major shopping centers, entertainment districts
- Public transit access (if applicable)
- Military bases (stable tenant pool)
- Tech parks or business districts

This section is CRITICAL for multifamily investment analysis - proximity to employment and education drives tenant demand.

Style Requirements:
- Clean headers
- Bullet points
- Bold all key numbers
- Never fabricate stats
- If a metric is missing: "Data not available"
- Tone: conversational, confident, analytical
- For location research: USE YOUR WEB SEARCH CAPABILITY to find real, current data"""

        # Build the user message with actual data
        user_message = f"""Analyze this market for real estate investment:

**Location:** {location.get('city', 'Unknown')}, {location.get('state', '')} - ZIP: {location.get('zip', 'N/A')}, County: {location.get('county', 'N/A')}

**Available Data:**

HOME VALUES (Zillow):
- Current ZHVI: ${market_data.get('homeValue', 'N/A'):,} if isinstance(market_data.get('homeValue'), (int, float)) else market_data.get('homeValue', 'N/A')
- 1-Month Forecast: {market_data.get('forecast1m', 'N/A')}%
- 3-Month Forecast: {market_data.get('forecast3m', 'N/A')}%
- 12-Month Forecast: {market_data.get('forecast12m', 'N/A')}%

RENTS (HUD Fair Market Rent):
- Studio: ${market_data.get('fmr0br', 'N/A')}
- 1BR: ${market_data.get('fmr1br', 'N/A')}
- 2BR: ${market_data.get('fmr2br', 'N/A')}
- 3BR: ${market_data.get('fmr3br', 'N/A')}
- 4BR: ${market_data.get('fmr4br', 'N/A')}

ECONOMIC (Census DP03):
- Median Household Income: ${market_data.get('medianIncome', 'N/A'):,} if isinstance(market_data.get('medianIncome'), (int, float)) else market_data.get('medianIncome', 'N/A')
- Unemployment Rate: {market_data.get('unemploymentRate', 'N/A')}%
- Poverty Rate: {market_data.get('povertyRate', 'N/A')}%

HOUSING (Census DP04):
- Total Housing Units: {market_data.get('totalUnits', 'N/A')}
- Median Home Value (Census): ${market_data.get('censusHomeValue', 'N/A')}
- Median Gross Rent (Census): ${market_data.get('censusRent', 'N/A')}
- Vacancy Rate: {market_data.get('vacancyRate', 'N/A')}%
- Owner Occupied: {market_data.get('ownerOccupied', 'N/A')}%
- Renter Occupied: {market_data.get('renterOccupied', 'N/A')}%

POPULATION (Census B01003):
- Total Population: {market_data.get('population', 'N/A')}

MIGRATION TRENDS (IRS/Census Data):
- Annual Inflow (people moving IN): {market_data.get('migrationInflow', 'N/A')}
- Annual Outflow (people moving OUT): {market_data.get('migrationOutflow', 'N/A')}
- Net Migration: {market_data.get('migrationNet', 'N/A')}
{f"- Net per 1,000 residents: {(market_data.get('migrationNet', 0) / market_data.get('population', 1) * 1000):.1f}" if market_data.get('population') and market_data.get('migrationNet') else ""}

LANDLORD SCORES (State: {location.get('state', 'N/A')}):
- Eviction Score: {market_data.get('evictionScore', 'N/A')}/10
- Deposit Score: {market_data.get('depositScore', 'N/A')}/10
- Rent Control Score: {market_data.get('rentControlScore', 'N/A')}/10
- Termination Score: {market_data.get('terminationScore', 'N/A')}/10

PROPERTY TAX:
- Effective Tax Rate: {market_data.get('propertyTaxRate', 'N/A')}%
- Median Taxes Paid: ${market_data.get('medianTaxesPaid', 'N/A')}

Generate your investment analysis based on this data."""

        # Add location research request if address is provided
        if deal_address:
            location_research = f"""

---

**IMPORTANT: LOCATION RESEARCH REQUIRED**

Property Address: {deal_address}
{f"Property Name: {property_name}" if property_name else ""}

Please use your web search capability to research and include in Section 6:

1. **Nearby Hospitals (within 10 miles)**: Search for hospitals near "{deal_address}" and list the closest 3-5 with approximate distances.

2. **Major Employers**: Search for major employers in {location.get('city', 'the area')}, {location.get('state', '')}. List top 5-10 companies, their industries, and any notable facts (HQ location, employee counts, recent expansions).

3. **Universities & Colleges (within 15 miles)**: Search for higher education institutions near "{deal_address}". Include enrollment numbers/student population for each.

4. **Employment Drivers**: What industries drive employment in this area? Any major business parks, industrial zones, or planned developments?

5. **Migration Patterns**: Research migration trends for {location.get('city', 'the area')}, {location.get('state', '')}. Are people moving to or from this area? What are the top origin cities/states? Why are people relocating here (jobs, cost of living, climate, etc.)?

This location research is CRITICAL for evaluating tenant demand - include specific names, distances, and numbers."""

            user_message += location_research

        # Build messages array
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_message}
        ]
        
        # Call Perplexity API
        headers = {
            "Authorization": f"Bearer {perplexity_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "sonar",
            "messages": messages,
            "temperature": 0.2,
            "max_tokens": 4000
        }
        
        async with httpx.AsyncClient(timeout=120.0) as client:
            response = await client.post(
                "https://api.perplexity.ai/chat/completions",
                headers=headers,
                json=payload
            )
            
            if response.status_code != 200:
                error_text = response.text
                print(f"Perplexity API error: {response.status_code} - {error_text}")
                return JSONResponse(
                    status_code=500,
                    content={"success": False, "error": f"Perplexity API error: {response.status_code}"}
                )
            
            result = response.json()
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "")
            
            # Deduct token after successful generation
            try:
                token_supabase = get_token_supabase()
                new_balance = profile["token_balance"] - tokens_required
                
                token_supabase.table("profiles").update({
                    "token_balance": new_balance
                }).eq("id", profile_id).execute()
                
                # Log usage
                token_supabase.table("token_usage").insert({
                    "profile_id": profile_id,
                    "operation_type": "market_research_results",
                    "tokens_used": tokens_required,
                    "deal_id": None,
                    "deal_name": property_name,
                    "location": f"{location.get('city', '')}, {location.get('state', '')} {location.get('zip', '')}"
                }).execute()
                
                log.info(f"Deducted {tokens_required} token(s) for market research. New balance: {new_balance}")
            except Exception as token_error:
                log.error(f"Failed to deduct token: {token_error}")
                # Don't fail the request if token deduction fails
            
            return {
                "success": True,
                "summary": content,
                "tokens_deducted": tokens_required,
                "new_balance": new_balance,
                "message": f"✓ {tokens_required} token deducted. Remaining balance: {new_balance}"
            }
            
    except Exception as e:
        print(f"Market data summary error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e)}
        )


# ============================================================================
# Deal Structure Recommendation Endpoint (AI-powered)
# ============================================================================

# Dedicated system prompt for Deal Structure Analysis - DO NOT USE ELSEWHERE
DEAL_STRUCTURE_SYSTEM_PROMPT = """You are Financing Architect AI, a specialist real estate investment analyst focused on designing optimal deal financing structures.

You are given deal-level numbers and terms (purchase price, NOI, current rents/expenses, cap rate, loan terms, etc.).

Your job is to brutally, quantitatively evaluate which one of the following 6 structures is best:

1. **Traditional** – Agency/bank loan (Freddie Mac, Fannie Mae, local bank)
2. **Seller Finance** – Seller holds a note (usually higher LTV, flexible terms)
3. **Subject To** – Taking over seller's existing mortgage
4. **Hybrid** – Subject To + Seller Carry combination
5. **Equity Partner** – JV partner funds down payment for equity split
6. **Seller Carry** – Bank in first position + seller second position note

---

## CORE METRICS & DEFINITIONS

Calculate and reference these explicitly for each structure:

**Annual Debt Service (ADS)**: Sum of all annual loan payments across all notes.

**Annual Cash Flow (CF)**: CF = NOI – Annual Debt Service. If CF < 0, clearly state it is NEGATIVE CASH FLOW.

**Debt Service Coverage Ratio (DSCR)**: DSCR = NOI ÷ Annual Debt Service. Banks want ≥ 1.25x. Below 1.0x = cash flow negative.

**Cash-on-Cash Return (CoC)**: CoC = Annual Cash Flow ÷ Total Cash Out of Pocket

**Value-Add & Forced Appreciation**: Value = NOI ÷ Cap Rate. Show equity created by improving NOI.

**Velocity of Money**: How fast investor can recycle capital via refinance or sale.

---

## DISQUALIFICATION RULES

Automatically reject a structure if:
- DSCR < 0.90x (severe negative cash flow, not viable)
- Subject To: No existing mortgage data provided
- Equity Partner: Profit split makes investor's effective CoC < 3%

Mark rejected structures as "NOT VIABLE: [reason]"

---

## TIE-BREAKER HIERARCHY

When structures are within 10% of each other:
1. **DSCR ≥ 1.20x** beats higher CoC with DSCR < 1.10x (safety first)
2. **Lower cash out of pocket** wins if DSCR is acceptable for both
3. **Simpler structure** wins if numbers are equal (fewer parties = fewer failure points)
4. **Better exit optionality** wins (easier refi path)

---

## INVESTOR PROFILES

Align recommendation with investor profile:
- **Cash Flow Focus** – Prioritizes maximum current cash flow and stable DSCR
- **Appreciation Play** – Accepts lower short-term cash flow for strong forced appreciation
- **Low Money Down** – Minimum capital out of pocket, high CoC, even if DSCR is tight
- **1031 Exchange** – Needs to place specific capital; may accept lower returns for safety

If profile not stated, assume balanced investor wanting positive cash flow, DSCR ≥ 1.20x, and reasonable equity growth.

---

## REQUIRED OUTPUT FORMAT

### 1. QUICK COMPARISON TABLE (Required First)

| Structure | Cash Required | Monthly Payment | Annual CF | DSCR | CoC | Verdict |
|-----------|---------------|-----------------|-----------|------|-----|---------|
| Traditional | $X | $X | $X | X.XXx | X.X% | ✅/⚠️/❌ |
| Seller Finance | ... | ... | ... | ... | ... | ... |
| Subject To | ... | ... | ... | ... | ... | ... |
| Hybrid | ... | ... | ... | ... | ... | ... |
| Equity Partner | ... | ... | ... | ... | ... | ... |
| Seller Carry | ... | ... | ... | ... | ... | ... |

Use: ✅ = Recommended, ⚠️ = Viable but risky, ❌ = Not viable

### 2. RECOMMENDED STRUCTURE

State clearly: "**RECOMMENDED: [Structure Name]**"

Explain WHY in 2-3 direct sentences with specific numbers.

### 3. STEP-BY-STEP CAPITAL BREAKDOWN

**At Closing:**
- Bank lends: $X
- Seller carries: $X (if applicable)
- Investor brings: $X
- Partner brings: $X (if applicable)

**Monthly/Annual:**
- Bank payment: $X/mo ($X/yr)
- Seller note payment: $X/mo (if applicable)
- Net cash flow to investor: $X/mo ($X/yr)

### 4. VALUE-ADD & EXIT STRATEGY

- Current NOI: $X → Proforma NOI: $X
- Current Value: $X → Stabilized Value: $X
- **Equity Created: $X**
- Recommended exit: Refi in X years / Sell at X cap / Long-term hold

### 5. RISKS & MITIGATIONS

**Key Risks:**
- Risk 1
- Risk 2

**Mitigations:**
- Mitigation 1
- Mitigation 2

### 6. ACTIONABLE NEXT STEPS

1. Step 1
2. Step 2
3. Step 3

---

## JSON OUTPUT (Required at END)

```json
{
  "recommendedStructure": "structure-key",
  "confidence": "high/medium/low",
  "primaryReason": "One sentence explanation",
  "keyMetrics": {
    "annualCashflow": 0,
    "dscr": 0.00,
    "cashOnCash": 0.00,
    "cashRequired": 0
  }
}
```

Structure keys: traditional, seller-finance, subject-to, hybrid, equity-partner, seller-carry

---

## STYLE REQUIREMENTS

- Professional investment memo tone
- Bold all key metrics and numbers
- Be direct and specific - no hand-waving
- If data is missing, state assumption clearly
- Do NOT sugar-coat bad numbers
- You are here to show hard numbers, not tell investor what they want to hear
"""

@app.post("/api/deal-structure/recommend")
async def deal_structure_recommend(request: Request):
    """
    Analyze all deal structures and recommend the optimal one based on investor goals.
    Uses Claude/Anthropic for intelligent analysis with dedicated Deal Structure prompt.
    REQUIRES 1 TOKEN.
    """
    try:
        # Check if user has tokens
        from token_manager import get_current_profile_id, get_profile, reset_tokens_if_needed, TOKEN_COSTS
        from token_manager import get_supabase as get_token_supabase
        
        try:
            profile_id = get_current_profile_id(request)
            profile = get_profile(profile_id)
            profile = reset_tokens_if_needed(profile)
            
            tokens_required = TOKEN_COSTS.get("deal_structure_analysis", 1)
            
            if profile["token_balance"] < tokens_required:
                return JSONResponse(
                    status_code=402,
                    content={
                        "success": False,
                        "error": f"Insufficient tokens. You need {tokens_required} token(s) but have {profile['token_balance']}.",
                        "tokens_required": tokens_required,
                        "token_balance": profile["token_balance"]
                    }
                )
        except HTTPException as he:
            log.warning(f"Token check failed: {he.detail}. Allowing request for backward compatibility.")
            profile_id = None
            profile = None
        
        data = await request.json()
        
        property_info = data.get("property", {})
        financials = data.get("financials", {})
        structures = data.get("structures", [])
        user_preferred = data.get("userPreferredStructure", "traditional")
        
        # Build the structure comparison table for the prompt
        structure_table = "\n".join([
            f"| {s['structure']} | ${s['cashOutOfPocket']:,.0f} | ${s['monthlyPayment']:,.0f} | ${s['annualCashflow']:,.0f} | {s['dscr']:.2f}x | {s['cashOnCash']:.1f}% | ? |"
            for s in structures
        ])
        
        user_prompt = f"""## DEAL ANALYSIS REQUEST

### Property Overview
- **Address**: {property_info.get('address', 'Unknown')}
- **Units**: {property_info.get('units', 0)}
- **Year Built**: {property_info.get('yearBuilt', 'Unknown')}
- **Type**: {property_info.get('type', 'Multifamily')}

### Financial Summary
- **Purchase Price**: ${financials.get('purchasePrice', 0):,.0f}
- **Current NOI**: ${financials.get('currentNOI', 0):,.0f}
- **Proforma NOI**: ${financials.get('proformaNOI', 0):,.0f}
- **Value-Add Potential**: ${financials.get('valueAddPotential', 0):,.0f}
- **As-Is Value**: ${financials.get('asIsValue', 0):,.0f}
- **Stabilized Value**: ${financials.get('stabilizedValue', 0):,.0f}
- **Cap Rate**: {financials.get('capRate', 5.5):.2f}%

### Pre-Calculated Structure Metrics

| Structure | Cash Required | Monthly Payment | Annual CF | DSCR | CoC | Verdict |
|-----------|---------------|-----------------|-----------|------|-----|---------|
{structure_table}

### User's Current Preferred Structure: {user_preferred}

---

**TASK**: Analyze all 6 structures above. Fill in the Verdict column. Recommend the OPTIMAL structure based on the metrics. Follow the exact output format specified in your instructions.

Provide analysis even if some structures show negative cash flow - explain why they're not viable and which one IS viable."""

        # Use Claude for analysis with dedicated Deal Structure prompt
        if not ANTHROPIC:
            # Fallback response if no API key
            return {
                "success": True,
                "recommendation": {
                    "recommendedStructure": "subject-to",
                    "summary": """## Quick Comparison

| Structure | Cash Required | Annual CF | DSCR | CoC | Verdict |
|-----------|---------------|-----------|------|-----|---------|
| Traditional | $756,000 | -$11,996 | 0.92x | -1.6% | ❌ Negative CF |
| Subject To | $270,000 | $33,543 | 1.31x | 12.4% | ✅ BEST |
| Seller Carry | $459,000 | $8,798 | 1.07x | 1.9% | ⚠️ Tight |

**RECOMMENDED: Subject To**

Subject To is the clear winner with **12.4% Cash-on-Cash** and **1.31x DSCR** - the only structure with both strong cash flow AND acceptable debt coverage. Traditional financing produces negative cash flow at current NOI.

**Next Steps:**
1. Verify seller's existing loan balance and terms
2. Confirm loan is assumable or has no due-on-sale clause
3. Structure earnest money to protect both parties
4. Execute value-add to increase NOI for future refi"""
                }
            }
        
        response = ANTHROPIC.messages.create(
            model="claude-3-haiku-20240307",
            max_tokens=4000,
            system=DEAL_STRUCTURE_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_prompt}]
        )
        
        ai_response = response.content[0].text
        
        # Try to parse JSON from the response
        recommended_structure = "traditional"  # default
        
        # First try to extract JSON block
        import re
        json_match = re.search(r'```json\s*(\{.*?\})\s*```', ai_response, re.DOTALL)
        if json_match:
            try:
                import json
                parsed = json.loads(json_match.group(1))
                if parsed.get("recommendedStructure"):
                    recommended_structure = parsed["recommendedStructure"]
            except:
                pass
        
        # Fallback: keyword matching
        if recommended_structure == "traditional":
            structure_keywords = {
                "traditional": ["traditional", "bank loan", "agency", "freddie", "fannie"],
                "seller-finance": ["seller finance", "seller financing", "owner finance"],
                "subject-to": ["subject to", "subject-to", "subto", "sub-to", "sub to"],
                "hybrid": ["hybrid", "combination"],
                "equity-partner": ["equity partner", "jv", "joint venture"],
                "seller-carry": ["seller carry", "seller second", "seller 2nd", "bank + seller"]
            }
            
            response_lower = ai_response.lower()
            for key, keywords in structure_keywords.items():
                for keyword in keywords:
                    if "recommend" in response_lower and keyword in response_lower:
                        recommended_structure = key
                        break
        
        # Deduct token after successful generation
        try:
            if profile_id and profile:
                token_supabase = get_token_supabase()
                new_balance = profile["token_balance"] - tokens_required
                
                token_supabase.table("profiles").update({
                    "token_balance": new_balance
                }).eq("id", profile_id).execute()
                
                # Log usage
                token_supabase.table("token_usage").insert({
                    "profile_id": profile_id,
                    "operation_type": "deal_structure_analysis",
                    "tokens_used": tokens_required,
                    "deal_id": None,
                    "deal_name": property_info.get('address', 'Unknown'),
                    "location": f"{property_info.get('address', '')}"
                }).execute()
                
                log.info(f"Deducted {tokens_required} token(s) for deal structure analysis. New balance: {new_balance}")
        except Exception as token_error:
            log.error(f"Failed to deduct token: {token_error}")
        
        return {
            "success": True,
            "recommendation": {
                "recommendedStructure": recommended_structure,
                "summary": ai_response
            }
        }
        
    except Exception as e:
        print(f"Deal structure recommendation error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e)}
        )


# ============================================================================
# SPREADSHEET AI ENDPOINT - Max can control the spreadsheet
# ============================================================================

@app.post("/api/spreadsheet/command")
async def spreadsheet_command(request: Request):
    """
    Process natural language commands and return spreadsheet operations
    Similar to Shortcut AI for Google Sheets
    """
    try:
        data = await request.json()
        user_message = data.get("message", "")
        property_data = data.get("propertyData", {})
        current_sheet_state = data.get("sheetState", None)
        
        if not user_message:
            raise HTTPException(status_code=400, detail="Message is required")
        
        # Process command through AI
        result = process_spreadsheet_command(
            user_message=user_message,
            property_data=property_data,
            current_sheet_state=current_sheet_state
        )
        
        return JSONResponse(content=result)
        
    except Exception as e:
        print(f"Spreadsheet command error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e), "operations": []}
        )


@app.get("/api/spreadsheet/get-template")
async def get_spreadsheet_template():
    """
    Return the multifamily underwriting model template from Excel file
    """
    try:
        from excel_to_spreadsheet import load_excel_template
        
        print("\n[GET TEMPLATE] Loading Excel template...")
        template_data = load_excel_template()
        print(f"[GET TEMPLATE] Loaded {len(template_data['rows'])} rows")
        
        return JSONResponse(content={
            "success": True,
            "data": template_data
        })
    except Exception as e:
        print(f"[GET TEMPLATE] Error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={
            "success": False,
            "error": str(e)
        })
        print(f"[GET TEMPLATE] Error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e)}
        )


# ============================================================================
# MAX PARTNER CHAT - Dedicated Max-only conversational endpoint
# ============================================================================

@app.post("/api/max/chat")
async def max_partner_chat(request: Request):
    """
    Conversational partner for the Property page.
    Uses Max-only partner system prompt. Not used elsewhere.
    """
    try:
        data = await request.json()
        messages = data.get("messages", [])
        # Optional page context
        sheet_state_json = data.get("sheet_state_json")
        sheet_calc_json = data.get("sheet_calc_json")
        sheet_structure = data.get("sheet_structure")
        active_tab = data.get("active_tab")
        scenario_data = data.get("scenario_data")
        full_calculations = data.get("full_calculations")
        property_data = data.get("property_data")

        if not messages or not isinstance(messages, list):
            raise HTTPException(status_code=400, detail="messages array is required")

        # Build compact context for the model
        parts = []
        if active_tab:
            parts.append(f"Active Tab: {active_tab}")
        if property_data:
            try:
                parts.append("Property: " + json.dumps(property_data)[:2000])
            except Exception:
                pass
        if scenario_data:
            try:
                parts.append("Scenario: " + json.dumps(scenario_data)[:2000])
            except Exception:
                pass
        if sheet_structure:
            try:
                parts.append("Sheet Structure: " + json.dumps(sheet_structure)[:1200])
            except Exception:
                pass
        if sheet_calc_json:
            try:
                parts.append("Calcs: " + json.dumps(sheet_calc_json)[:1200])
            except Exception:
                pass
        if sheet_state_json:
            try:
                parts.append("State: " + json.dumps(sheet_state_json)[:1200])
            except Exception:
                pass

        context_text = "\n".join(parts)

        if ANTHROPIC is None:
            # Graceful fallback to avoid 500s when Claude is not configured
            assistant_text = (
                "Max partner chat is offline right now. Your spreadsheet and mapping still work. "
                "Please configure ANTHROPIC_API_KEY (or CLAUDE_API_KEY) in backend/.env to enable chat, "
                "or continue with spreadsheet actions."
            )
            return JSONResponse(content={
                "success": True,
                "message": {"role": "assistant", "content": assistant_text}
            })

        res = ANTHROPIC.messages.create(
            model="claude-3-haiku-20240307",
            max_tokens=4000,
            system=MAX_PARTNER_SYSTEM_PROMPT,
            messages=[
                {"role": "user", "content": context_text},
                *messages
            ],
        )

        try:
            if isinstance(res.content, list) and len(res.content) and hasattr(res.content[0], "text"):
                assistant_text = res.content[0].text
            elif isinstance(res, dict) and isinstance(res.get("content"), list):
                assistant_text = str(res["content"][0])
            else:
                assistant_text = str(getattr(res, "content", ""))
            if not assistant_text:
                assistant_text = "(No response text)"
        except Exception:
            assistant_text = "(Failed to parse model response)"
        return JSONResponse(content={
            "success": True,
            "message": {"role": "assistant", "content": assistant_text}
        })
    except Exception as e:
        print(f"[MAX PARTNER CHAT] Error: {e}")
        import traceback
        traceback.print_exc()
        # Do not surface 500s to the client; return a friendly message
        return JSONResponse(content={
            "success": True,
            "message": {"role": "assistant", "content": (
                "Partner chat encountered an issue but your spreadsheet features are unaffected. "
                "Please configure ANTHROPIC_API_KEY or try again later."
            )}
        })


@app.post("/api/spreadsheet/build-model")
async def spreadsheet_build_model_direct(request: Request):
    """
    Build full underwriting model directly without Claude API
    """
    try:
        from spreadsheet_ai import build_full_underwriting_model
        
        data = await request.json()
        property_data = data.get("propertyData", {})
        
        print("\n" + "="*80)
        print("[BUILD MODEL DIRECT] ENDPOINT HIT!")
        print("="*80 + "\n")
        
        # Build model with default parameters
        model_spec = {
            "parameters": {
                "units": 102,
                "purchasePrice": 7400000,
                "propertyName": "New Town Apartments",
                "totalSf": 75904,
                "yearBuilt": 1985,
                "loanAmount": 5550000,
                "interestRate": 0.06,
                "loanTermMonths": 360,
                "closingCosts": 36000,
                "dueDiligence": 15000,
                "capexBudgetYr1": 150000,
                "financingCosts": 12000,
                "operatingReserves": 50000,
                "lpEquity": 1901700,
                "gpEquity": 211300,
                "rentGrowth": 0.03,
                "expenseGrowth": 0.025,
                "vacancyRate": 0.05,
                "exitCap5Yr": 0.06,
                "exitCap10Yr": 0.065,
                "prefReturn": 0.08
            }
        }
        
        print("[BUILD MODEL DIRECT] Calling build_full_underwriting_model...")
        operations = build_full_underwriting_model(model_spec, property_data)
        print(f"[BUILD MODEL DIRECT] Generated {len(operations)} operations")
        print(f"[BUILD MODEL DIRECT] First 3 operations: {operations[:3]}")
        
        return JSONResponse(content={
            "success": True,
            "operations": operations
        })
        
    except Exception as e:
        print(f"[BUILD MODEL DIRECT] ERROR: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={"success": False, "error": str(e), "operations": []}
        )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", "8010")))