"""
Template-Based Excel Generator for Deal Builder

Reads the underwriting_template.json and populates it with deal data,
producing a comprehensive institutional-grade underwriting model.
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Styling constants
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="2E5A88", end_color="2E5A88", fill_type="solid")
INPUT_FONT = Font(color="0066CC")  # Blue for changeable inputs
CALC_FONT = Font(color="000000")  # Black for calculated
GREEN_FONT = Font(color="006600")  # Green for references
YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

CURRENCY_FORMAT = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
PERCENT_FORMAT = '0.00%'
NUMBER_FORMAT = '#,##0'


def safe_num(val, default=0):
    """Safely convert to number."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def build_full_underwriting_model(deal_data: Dict[str, Any]) -> openpyxl.Workbook:
    """
    Build a complete institutional underwriting model from deal data.
    
    Matches the comprehensive template structure with:
    - Deal Snapshot + Financial Performance + Equity Returns (rows 4-12)
    - Property Information + Cashflow Summary (rows 18-38)
    - Purchase & Financing (rows 31-44)
    - Year 5 Buyout Analysis (rows 40-76)
    - Cashflow Scenarios with 3 scenarios (rows 60-98)
    - Sensitivity Analysis tables (rows 101-120)
    - Rent Roll (rows 122+)
    """
    log.info("[TemplateGen] Building full underwriting model...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Underwriting Model"
    
    # Extract deal data sections
    prop = deal_data.get("property", {})
    fin = deal_data.get("financials", {})
    exp = deal_data.get("expenses", {})
    inc = deal_data.get("income", {})
    occ = deal_data.get("occupancy", {})
    unit_mix = deal_data.get("unit_mix", [])
    financing = deal_data.get("financing", {})
    
    # Core values
    deal_name = prop.get("name") or prop.get("address", "Deal")
    address = prop.get("address", "")
    city = prop.get("city", "")
    state = prop.get("state", "")
    zip_code = prop.get("zip", "")
    units = safe_num(prop.get("units", 0))
    year_built = prop.get("year_built", "")
    building_sf = safe_num(prop.get("building_sf") or prop.get("gross_building_sf", 0))
    
    # Financial values
    purchase_price = safe_num(fin.get("asking_price") or fin.get("purchase_price", 0))
    noi = safe_num(fin.get("noi", 0))
    cap_rate = safe_num(fin.get("cap_rate", 0))
    if cap_rate > 1:
        cap_rate = cap_rate / 100
    
    # Financing assumptions (with defaults)
    down_payment_pct = safe_num(financing.get("down_payment_pct", 0.20))
    interest_rate = safe_num(financing.get("interest_rate", 0.07))
    amort_years = safe_num(financing.get("amortization_years", 30))
    closing_costs_pct = safe_num(financing.get("closing_costs_pct", 0.01))
    
    # Calculated financing
    down_payment = purchase_price * down_payment_pct
    loan_amount = purchase_price - down_payment
    closing_costs = purchase_price * closing_costs_pct
    total_cash_to_close = down_payment + closing_costs
    
    # Monthly payment calculation (P&I)
    monthly_rate = interest_rate / 12
    num_payments = amort_years * 12
    if monthly_rate > 0 and num_payments > 0:
        monthly_payment = loan_amount * (monthly_rate * (1 + monthly_rate)**num_payments) / ((1 + monthly_rate)**num_payments - 1)
    else:
        monthly_payment = 0
    annual_debt_service = monthly_payment * 12
    
    # Cashflow
    annual_cashflow = noi - annual_debt_service
    monthly_cashflow = annual_cashflow / 12
    
    # Returns
    dscr = noi / annual_debt_service if annual_debt_service > 0 else 0
    price_per_unit = purchase_price / units if units > 0 else 0
    
    # Occupancy
    current_occupancy = safe_num(occ.get("current_occupancy", 0.90))
    if current_occupancy > 1:
        current_occupancy = current_occupancy / 100
    
    # Expenses breakdown
    property_tax = safe_num(exp.get("property_tax") or exp.get("taxes", 0))
    insurance = safe_num(exp.get("insurance", 0))
    repairs = safe_num(exp.get("repairs_maintenance") or exp.get("repairs_and_maintenance", 0))
    utilities = safe_num(exp.get("utilities", 0))
    water_sewer = safe_num(exp.get("water_sewer", 0))
    trash = safe_num(exp.get("trash") or exp.get("trash_service", 0))
    landscaping = safe_num(exp.get("landscaping", 0))
    payroll = safe_num(exp.get("payroll", 0))
    admin = safe_num(exp.get("admin") or exp.get("general_admin", 0))
    marketing = safe_num(exp.get("marketing", 0))
    management = safe_num(exp.get("management_fee") or exp.get("management", 0))
    reserves = safe_num(exp.get("reserves", 0))
    total_expenses = safe_num(fin.get("total_expenses", 0))
    if total_expenses == 0:
        total_expenses = property_tax + insurance + repairs + utilities + water_sewer + trash + landscaping + payroll + admin + marketing + management + reserves
    
    # Income
    gross_income = safe_num(inc.get("scheduled_gross_income") or fin.get("gross_potential_rent", 0))
    other_income = safe_num(inc.get("other_income", 0))
    vacancy_loss = safe_num(inc.get("vacancy_loss") or fin.get("vacancy_loss", 0))
    if vacancy_loss == 0 and gross_income > 0:
        vacancy_loss = gross_income * 0.10  # Default 10% vacancy
    
    # ═══════════════════════════════════════════════════════════════════
    # ROW 1-2: TITLE
    # ═══════════════════════════════════════════════════════════════════
    ws["A1"] = f"{deal_name} — INVESTMENT UNDERWRITING MODEL"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:K1")
    
    full_address = f"{address} | {city}, {state} {zip_code}".strip(" |")
    ws["A2"] = f"{full_address}  |  {int(units)} Units  |  Built {year_built}  |  ${purchase_price:,.0f}  |  {down_payment_pct:.0%} Down  |  {interest_rate:.2%} / {int(amort_years)} yr"
    ws["A2"].font = Font(size=10, color="666666")
    ws.merge_cells("A2:K2")
    
    # ═══════════════════════════════════════════════════════════════════
    # ROW 4: SECTION HEADERS
    # ═══════════════════════════════════════════════════════════════════
    ws["A4"] = "DEAL SNAPSHOT"
    ws["A4"].font = HEADER_FONT
    ws["A4"].fill = HEADER_FILL
    ws.merge_cells("A4:B4")
    
    ws["G4"] = "FINANCIAL PERFORMANCE"
    ws["G4"].font = HEADER_FONT
    ws["G4"].fill = HEADER_FILL
    ws.merge_cells("G4:H4")
    
    ws["J4"] = "EQUITY RETURNS"
    ws["J4"].font = HEADER_FONT
    ws["J4"].fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    ws.merge_cells("J4:K4")
    
    # ═══════════════════════════════════════════════════════════════════
    # ROWS 5-12: DEAL SNAPSHOT + FINANCIAL PERFORMANCE + EQUITY RETURNS
    # ═══════════════════════════════════════════════════════════════════
    snapshot_data = [
        ("Purchase Price", purchase_price, "NOI — Current", noi, "Total Investment", total_cash_to_close),
        ("Price Per Unit", price_per_unit, "Cap Rate", cap_rate, "Annual Cashflow", annual_cashflow),
        ("Total Cash to Close", total_cash_to_close, "DSCR", dscr, "Monthly Cashflow", monthly_cashflow),
        ("Loan Amount", loan_amount, "Annual Debt Service", annual_debt_service, "Cash-on-Cash", annual_cashflow / total_cash_to_close if total_cash_to_close > 0 else 0),
        ("Monthly Debt Service", monthly_payment, "Interest Rate", interest_rate, "", ""),
        ("Occupancy", current_occupancy, "Amortization", f"{int(amort_years)} yrs", "", ""),
        ("Total Units", units, "", "", "", ""),
        ("Year Built", year_built, "", "", "", ""),
    ]
    
    for i, (label1, val1, label2, val2, label3, val3) in enumerate(snapshot_data):
        row = 5 + i
        
        # Column A-B: Deal Snapshot
        ws[f"A{row}"] = label1
        if isinstance(val1, (int, float)) and label1 in ["Purchase Price", "Price Per Unit", "Total Cash to Close", "Loan Amount", "Monthly Debt Service"]:
            ws[f"B{row}"] = val1
            ws[f"B{row}"].number_format = CURRENCY_FORMAT
        elif isinstance(val1, float) and label1 == "Occupancy":
            ws[f"B{row}"] = val1
            ws[f"B{row}"].number_format = PERCENT_FORMAT
        else:
            ws[f"B{row}"] = val1
        
        # Column G-H: Financial Performance
        if label2:
            ws[f"G{row}"] = label2
            if "Cap Rate" in label2 or "Interest" in label2:
                ws[f"H{row}"] = val2
                ws[f"H{row}"].number_format = PERCENT_FORMAT
            elif "NOI" in label2 or "Debt Service" in label2:
                ws[f"H{row}"] = val2
                ws[f"H{row}"].number_format = CURRENCY_FORMAT
            elif "DSCR" in label2:
                ws[f"H{row}"] = val2
                ws[f"H{row}"].number_format = "0.00x"
            else:
                ws[f"H{row}"] = val2
        
        # Column J-K: Equity Returns
        if label3:
            ws[f"J{row}"] = label3
            if "Cashflow" in label3 or "Investment" in label3:
                ws[f"K{row}"] = val3
                ws[f"K{row}"].number_format = CURRENCY_FORMAT
            elif "Cash-on-Cash" in label3:
                ws[f"K{row}"] = val3
                ws[f"K{row}"].number_format = PERCENT_FORMAT
    
    # ═══════════════════════════════════════════════════════════════════
    # ROW 14: LEGEND
    # ═══════════════════════════════════════════════════════════════════
    ws["A14"] = "Blue = changeable inputs  |  Black = calculated formulas  |  All scenarios update automatically when inputs change"
    ws["A14"].font = Font(size=9, color="666666")
    ws.merge_cells("A14:K14")
    
    # ═══════════════════════════════════════════════════════════════════
    # ROW 16: SECTION DIVIDERS
    # ═══════════════════════════════════════════════════════════════════
    ws["A16"] = "═══════════════  ASSUMPTIONS  ═══════════════"
    ws["A16"].font = Font(bold=True)
    ws.merge_cells("A16:D16")
    
    ws["F16"] = "═══════════════  CASHFLOW SCENARIOS  ═══════════════"
    ws["F16"].font = Font(bold=True)
    ws.merge_cells("F16:I16")
    
    # ═══════════════════════════════════════════════════════════════════
    # ROW 18: PROPERTY INFO + CASHFLOW HEADERS
    # ═══════════════════════════════════════════════════════════════════
    ws["A18"] = "PROPERTY INFORMATION"
    ws["A18"].font = HEADER_FONT
    ws["A18"].fill = HEADER_FILL
    
    ws["F18"] = "CASHFLOW SUMMARY"
    ws["F18"].font = HEADER_FONT
    ws["F18"].fill = HEADER_FILL
    
    ws["G18"] = "Scenario 1\nCurrent"
    ws["G18"].alignment = Alignment(wrap_text=True, horizontal="center")
    ws["H18"] = "Scenario 2\nRUBS + Trash"
    ws["H18"].alignment = Alignment(wrap_text=True, horizontal="center")
    ws["I18"] = "Scenario 3\nFull Value-Add"
    ws["I18"].alignment = Alignment(wrap_text=True, horizontal="center")
    
    # ═══════════════════════════════════════════════════════════════════
    # ROWS 19-30: PROPERTY INFO + CASHFLOW SUMMARY
    # ═══════════════════════════════════════════════════════════════════
    property_info = [
        ("Property Name", deal_name),
        ("Address", f"{address}, {city}, {state} {zip_code}"),
        ("Asset Type", prop.get("property_type", "Multifamily")),
        ("Year Built", year_built),
        ("APN", prop.get("apn", "")),
        ("Gross Building SF", building_sf),
        ("Lot Size (Acres)", prop.get("lot_size", "")),
        ("Total Units", units),
        ("Unit Mix", "See Rent Roll"),
        ("Current Occupancy", current_occupancy),
        ("Average SF Per Unit", building_sf / units if units > 0 else 0),
        ("Average Current Rent", gross_income / 12 / units if units > 0 else 0),
    ]
    
    for i, (label, val) in enumerate(property_info):
        row = 19 + i
        ws[f"A{row}"] = label
        if label == "Current Occupancy":
            ws[f"B{row}"] = val
            ws[f"B{row}"].number_format = PERCENT_FORMAT
        elif label in ["Gross Building SF", "Total Units"]:
            ws[f"B{row}"] = val
            ws[f"B{row}"].number_format = NUMBER_FORMAT
        elif label == "Average Current Rent":
            ws[f"B{row}"] = val
            ws[f"B{row}"].number_format = CURRENCY_FORMAT
        else:
            ws[f"B{row}"] = val
        ws[f"B{row}"].font = INPUT_FONT
    
    # Cashflow summary (columns F-I)
    cashflow_rows = [
        ("Total NOI", noi, noi, noi),
        ("Annual Debt Service", annual_debt_service, annual_debt_service, annual_debt_service),
        ("Total Cashflow Available", annual_cashflow, annual_cashflow, annual_cashflow),
    ]
    
    for i, (label, s1, s2, s3) in enumerate(cashflow_rows):
        row = 19 + i
        ws[f"F{row}"] = label
        ws[f"G{row}"] = s1
        ws[f"G{row}"].number_format = CURRENCY_FORMAT
        ws[f"H{row}"] = s2
        ws[f"H{row}"].number_format = CURRENCY_FORMAT
        ws[f"I{row}"] = s3
        ws[f"I{row}"].number_format = CURRENCY_FORMAT
    
    # ═══════════════════════════════════════════════════════════════════
    # ROW 32: PURCHASE & FINANCING
    # ═══════════════════════════════════════════════════════════════════
    ws["A32"] = "PURCHASE & FINANCING"
    ws["A32"].font = HEADER_FONT
    ws["A32"].fill = HEADER_FILL
    
    financing_rows = [
        ("Purchase Price", purchase_price, CURRENCY_FORMAT, True),
        ("Down Payment %", down_payment_pct, PERCENT_FORMAT, True),
        ("Interest Rate", interest_rate, PERCENT_FORMAT, True),
        ("Amortization (Years)", amort_years, "0", True),
        ("Closing Costs %", closing_costs_pct, PERCENT_FORMAT, True),
        ("Working Capital ($)", 0, CURRENCY_FORMAT, True),
        ("Down Payment ($)", f"=B33*B34", CURRENCY_FORMAT, False),
        ("CALCULATED FINANCING", "", "", False),
        ("Loan Amount", f"=B33-B39", CURRENCY_FORMAT, False),
        ("Closing Costs ($)", f"=B33*B37", CURRENCY_FORMAT, False),
        ("Total Cash to Close", f"=B39+B42+B38", CURRENCY_FORMAT, False),
        ("Monthly Debt Service", f"=IFERROR(B41*(B35/12)/(1-(1+B35/12)^(-B36*12)),0)", CURRENCY_FORMAT, False),
        ("Annual Debt Service", f"=B44*12", CURRENCY_FORMAT, False),
    ]
    
    for i, (label, val, fmt, is_input) in enumerate(financing_rows):
        row = 33 + i
        ws[f"A{row}"] = label
        if label == "CALCULATED FINANCING":
            ws[f"A{row}"].font = Font(bold=True, size=9)
        elif isinstance(val, str) and val.startswith("="):
            ws[f"B{row}"] = val
        else:
            ws[f"B{row}"] = val
        if fmt:
            ws[f"B{row}"].number_format = fmt
        ws[f"B{row}"].font = INPUT_FONT if is_input else CALC_FONT
    
    # ═══════════════════════════════════════════════════════════════════
    # ROW 48: EQUITY PARTNER STRUCTURE
    # ═══════════════════════════════════════════════════════════════════
    ws["A48"] = "EQUITY PARTNER STRUCTURE"
    ws["A48"].font = HEADER_FONT
    ws["A48"].fill = HEADER_FILL
    
    partner_rows = [
        ("Investor Contribution ($)", 0, CURRENCY_FORMAT, True),
        ("Your Contribution ($)", total_cash_to_close, CURRENCY_FORMAT, True),
        ("Preferred Return %", 0.08, PERCENT_FORMAT, True),
        ("Partner Equity %", 0.20, PERCENT_FORMAT, True),
        ("Your Equity %", 0.80, PERCENT_FORMAT, True),
    ]
    
    for i, (label, val, fmt, is_input) in enumerate(partner_rows):
        row = 49 + i
        ws[f"A{row}"] = label
        ws[f"B{row}"] = val
        ws[f"B{row}"].number_format = fmt
        ws[f"B{row}"].font = INPUT_FONT if is_input else CALC_FONT
    
    # ═══════════════════════════════════════════════════════════════════
    # ROW 55: VALUE-ADD ASSUMPTIONS
    # ═══════════════════════════════════════════════════════════════════
    ws["A55"] = "VALUE-ADD ASSUMPTIONS"
    ws["A55"].font = HEADER_FONT
    ws["A55"].fill = HEADER_FILL
    
    value_add_rows = [
        ("Rent Bump / Unit / Mo ($)", 0, CURRENCY_FORMAT),
        ("RUBS Recovery / Year ($)", 0, CURRENCY_FORMAT),
        ("Trash Fee / Unit / Mo ($)", 0, CURRENCY_FORMAT),
        ("Annual Appreciation Rate", 0.03, PERCENT_FORMAT),
        ("Exit Cap — Conservative", 0.085, PERCENT_FORMAT),
        ("Exit Cap — Market", 0.075, PERCENT_FORMAT),
        ("Hold Period (Years)", 5, "0"),
    ]
    
    for i, (label, val, fmt) in enumerate(value_add_rows):
        row = 56 + i
        ws[f"A{row}"] = label
        ws[f"B{row}"] = val
        ws[f"B{row}"].number_format = fmt
        ws[f"B{row}"].font = INPUT_FONT
    
    # ═══════════════════════════════════════════════════════════════════
    # ROW 65: INCOME SECTION (3 SCENARIOS)
    # ═══════════════════════════════════════════════════════════════════
    ws["A65"] = "═══════════════  OPERATING STATEMENT  ═══════════════"
    ws["A65"].font = Font(bold=True)
    ws.merge_cells("A65:D65")
    
    ws["B67"] = "Scenario 1\nCurrent"
    ws["C67"] = "Scenario 2\nRUBS + Trash"
    ws["D67"] = "Scenario 3\nFull Value-Add"
    
    ws["A68"] = "Scheduled Gross Income"
    ws["B68"] = gross_income
    ws["B68"].number_format = CURRENCY_FORMAT
    ws["B68"].font = INPUT_FONT
    ws["C68"] = gross_income
    ws["C68"].number_format = CURRENCY_FORMAT
    ws["D68"] = f"=B68+B56*12*B28*B30"  # Add rent bumps
    ws["D68"].number_format = CURRENCY_FORMAT
    
    ws["A69"] = "Less: Vacancy (10%)"
    ws["B69"] = f"=-B68*0.1"
    ws["C69"] = f"=-C68*0.1"
    ws["D69"] = f"=-D68*0.1"
    for col in ["B", "C", "D"]:
        ws[f"{col}69"].number_format = CURRENCY_FORMAT
    
    ws["A70"] = "Other Income"
    ws["B70"] = other_income
    ws["B70"].number_format = CURRENCY_FORMAT
    ws["B70"].font = INPUT_FONT
    ws["C70"] = other_income
    ws["C70"].number_format = CURRENCY_FORMAT
    ws["D70"] = other_income
    ws["D70"].number_format = CURRENCY_FORMAT
    
    ws["A71"] = "RUBS Recovery"
    ws["B71"] = 0
    ws["C71"] = f"=B57"
    ws["D71"] = f"=B57"
    for col in ["B", "C", "D"]:
        ws[f"{col}71"].number_format = CURRENCY_FORMAT
    
    ws["A72"] = "Trash Fee Income"
    ws["B72"] = 0
    ws["C72"] = f"=B58*12*B28*B30"
    ws["D72"] = f"=B58*12*B28*B30"
    for col in ["B", "C", "D"]:
        ws[f"{col}72"].number_format = CURRENCY_FORMAT
    
    ws["A73"] = "GROSS OPERATING INCOME"
    ws["A73"].font = Font(bold=True)
    ws["B73"] = f"=SUM(B68:B72)"
    ws["C73"] = f"=SUM(C68:C72)"
    ws["D73"] = f"=SUM(D68:D72)"
    for col in ["B", "C", "D"]:
        ws[f"{col}73"].number_format = CURRENCY_FORMAT
        ws[f"{col}73"].font = Font(bold=True)
    
    # ═══════════════════════════════════════════════════════════════════
    # EXPENSES
    # ═══════════════════════════════════════════════════════════════════
    ws["A75"] = "EXPENSES"
    ws["A75"].font = HEADER_FONT
    ws["A75"].fill = HEADER_FILL
    
    expense_items = [
        ("Property Tax", property_tax),
        ("Insurance", insurance),
        ("Management (6% of GOI)", f"=B73*0.06"),
        ("Repairs & Maintenance", repairs),
        ("Turnover / Cleaning", 0),
        ("Landscaping", landscaping),
        ("Payroll", payroll),
        ("Water / Sewer", water_sewer),
        ("Electricity (Common)", utilities),
        ("Pest Control", 0),
        ("Trash Service", trash),
        ("General Admin", admin),
        ("Marketing", marketing),
        ("Reserves", reserves),
        ("Licenses & Permits", 0),
    ]
    
    for i, (label, val) in enumerate(expense_items):
        row = 76 + i
        ws[f"A{row}"] = label
        for col in ["B", "C", "D"]:
            if isinstance(val, str) and val.startswith("="):
                ws[f"{col}{row}"] = val.replace("B73", f"{col}73")
            else:
                ws[f"{col}{row}"] = val
            ws[f"{col}{row}"].number_format = CURRENCY_FORMAT
            if not (isinstance(val, str) and val.startswith("=")):
                ws[f"B{row}"].font = INPUT_FONT
    
    ws["A91"] = "TOTAL EXPENSES"
    ws["A91"].font = Font(bold=True)
    ws["B91"] = f"=SUM(B76:B90)"
    ws["C91"] = f"=SUM(C76:C90)"
    ws["D91"] = f"=SUM(D76:D90)"
    for col in ["B", "C", "D"]:
        ws[f"{col}91"].number_format = CURRENCY_FORMAT
        ws[f"{col}91"].font = Font(bold=True)
    
    # ═══════════════════════════════════════════════════════════════════
    # NOI & CASHFLOW
    # ═══════════════════════════════════════════════════════════════════
    ws["A93"] = "NET OPERATING INCOME"
    ws["A93"].font = Font(bold=True, size=11)
    ws["A93"].fill = GREEN_FILL
    ws["B93"] = f"=B73-B91"
    ws["C93"] = f"=C73-C91"
    ws["D93"] = f"=D73-D91"
    for col in ["B", "C", "D"]:
        ws[f"{col}93"].number_format = CURRENCY_FORMAT
        ws[f"{col}93"].font = Font(bold=True)
        ws[f"{col}93"].fill = GREEN_FILL
    
    ws["A95"] = "DEBT SERVICE & CASHFLOW"
    ws["A95"].font = HEADER_FONT
    ws["A95"].fill = HEADER_FILL
    
    ws["A96"] = "Annual Debt Service"
    ws["B96"] = f"=B45"
    ws["C96"] = f"=B45"
    ws["D96"] = f"=B45"
    for col in ["B", "C", "D"]:
        ws[f"{col}96"].number_format = CURRENCY_FORMAT
    
    ws["A97"] = "Annual Cashflow"
    ws["B97"] = f"=B93-B96"
    ws["C97"] = f"=C93-C96"
    ws["D97"] = f"=D93-D96"
    for col in ["B", "C", "D"]:
        ws[f"{col}97"].number_format = CURRENCY_FORMAT
    
    ws["A98"] = "Monthly Cashflow"
    ws["B98"] = f"=B97/12"
    ws["C98"] = f"=C97/12"
    ws["D98"] = f"=D97/12"
    for col in ["B", "C", "D"]:
        ws[f"{col}98"].number_format = CURRENCY_FORMAT
    
    # ═══════════════════════════════════════════════════════════════════
    # RETURNS
    # ═══════════════════════════════════════════════════════════════════
    ws["A100"] = "RETURNS"
    ws["A100"].font = HEADER_FONT
    ws["A100"].fill = HEADER_FILL
    
    ws["A101"] = "Cap Rate"
    ws["B101"] = f"=B93/B33"
    ws["C101"] = f"=C93/B33"
    ws["D101"] = f"=D93/B33"
    for col in ["B", "C", "D"]:
        ws[f"{col}101"].number_format = PERCENT_FORMAT
    
    ws["A102"] = "DSCR"
    ws["B102"] = f"=B93/B96"
    ws["C102"] = f"=C93/B96"
    ws["D102"] = f"=D93/B96"
    for col in ["B", "C", "D"]:
        ws[f"{col}102"].number_format = "0.00x"
    
    ws["A103"] = "Cash-on-Cash (Total)"
    ws["B103"] = f"=B97/B43"
    ws["C103"] = f"=C97/B43"
    ws["D103"] = f"=D97/B43"
    for col in ["B", "C", "D"]:
        ws[f"{col}103"].number_format = PERCENT_FORMAT
    
    ws["A104"] = "Price Per Unit"
    ws["B104"] = f"=B33/B28"
    ws["C104"] = f"=B33/B28"
    ws["D104"] = f"=B33/B28"
    for col in ["B", "C", "D"]:
        ws[f"{col}104"].number_format = CURRENCY_FORMAT
    
    # ═══════════════════════════════════════════════════════════════════
    # SENSITIVITY ANALYSIS
    # ═══════════════════════════════════════════════════════════════════
    ws["A107"] = "═══════════════  SENSITIVITY ANALYSIS  ═══════════════"
    ws["A107"].font = Font(bold=True)
    ws.merge_cells("A107:I107")
    
    ws["A109"] = "TABLE 1: PROPERTY VALUE AT EXIT — NOI × Exit Cap Rate"
    ws["A109"].font = Font(bold=True)
    
    # Exit cap rate headers
    exit_caps = [0.08, 0.085, 0.09, 0.095, 0.10, 0.105, 0.11]
    ws["A110"] = "NOI Scenario"
    ws["B110"] = "NOI Value ($)"
    for i, cap in enumerate(exit_caps):
        col = get_column_letter(3 + i)
        ws[f"{col}110"] = cap
        ws[f"{col}110"].number_format = PERCENT_FORMAT
    
    # NOI scenarios
    noi_scenarios = [
        ("Current Expenses NOI", "=B93"),
        ("RUBS + Trash NOI", "=C93"),
        ("Full Value-Add NOI", "=D93"),
        ("Value-Add +10%", "=D93*1.1"),
        ("Value-Add -10%", "=D93*0.9"),
    ]
    
    for i, (label, noi_formula) in enumerate(noi_scenarios):
        row = 111 + i
        ws[f"A{row}"] = label
        ws[f"B{row}"] = noi_formula
        ws[f"B{row}"].number_format = CURRENCY_FORMAT
        for j, _ in enumerate(exit_caps):
            col = get_column_letter(3 + j)
            cap_col = get_column_letter(3 + j)
            ws[f"{col}{row}"] = f"=$B{row}/{cap_col}110"
            ws[f"{col}{row}"].number_format = CURRENCY_FORMAT
    
    # ═══════════════════════════════════════════════════════════════════
    # RENT ROLL
    # ═══════════════════════════════════════════════════════════════════
    ws["A120"] = "═══════════════  RENT ROLL  ═══════════════"
    ws["A120"].font = Font(bold=True)
    ws.merge_cells("A120:H120")
    
    ws["A122"] = "Unit #"
    ws["B122"] = "Mix"
    ws["C122"] = "Current Rent"
    ws["D122"] = "Market Rent"
    ws["E122"] = "Loss to Lease"
    ws["F122"] = "SF"
    ws["G122"] = "Rent/SF"
    ws["H122"] = "Market Rent/SF"
    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws[f"{col}122"].font = HEADER_FONT
        ws[f"{col}122"].fill = HEADER_FILL
    
    # Add unit rows (up to 50 units)
    max_units = min(int(units) if units > 0 else 20, 50)
    for i in range(max_units):
        row = 123 + i
        ws[f"A{row}"] = i + 1
        ws[f"E{row}"] = f"=D{row}-C{row}"
        ws[f"E{row}"].number_format = CURRENCY_FORMAT
        ws[f"G{row}"] = f"=IFERROR(C{row}/F{row},0)"
        ws[f"G{row}"].number_format = CURRENCY_FORMAT
        ws[f"H{row}"] = f"=IFERROR(D{row}/F{row},0)"
        ws[f"H{row}"].number_format = CURRENCY_FORMAT
    
    # ═══════════════════════════════════════════════════════════════════
    # COLUMN WIDTHS
    # ═══════════════════════════════════════════════════════════════════
    col_widths = {
        "A": 28, "B": 14, "C": 14, "D": 14, "E": 14,
        "F": 22, "G": 14, "H": 14, "I": 14, "J": 20, "K": 14
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    log.info("[TemplateGen] Model built successfully")
    return wb
