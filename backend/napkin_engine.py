"""
Back of the Napkin — a fast, automatic full-deal-read feature.

This is intentionally separate from the platform's real v2 underwriting
engine (v2_underwriter/) and its calc_json pipeline — nothing here reads or
writes deal storage, scenario data, or the main underwriting model. It is
driven ENTIRELY by the CRE Agent Skills library dropped into
backend/cre-agent-skills-main/ (66 skill playbooks + knowledge bases,
Apache 2.0, plain markdown — no code) plus the standalone
backend/scenario-matrix-analyzer/SKILL.md. Skills are read from disk at
runtime, the same pattern used by board_of_advisors_engine.py, so
adding/editing a skill is a file edit, not a code change.

Two entry points:
- generate_napkin_report(): the primary flow. One document upload in, one
  structured underwrite report out (OM issues, market outlook, strategy/play,
  recommended purchase price, investor payback feasibility) — no back-and-forth
  required.
- chat_about_napkin_deal(): a secondary, free-form Q&A follow-up on the same
  document/report ("what if I offered $X instead", etc.).
"""
import os
import io
import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

from anthropic import Anthropic

log = logging.getLogger("napkin")

SKILLS_ROOT = Path(__file__).parent / "cre-agent-skills-main"
SKILLS_DIR = SKILLS_ROOT / "skills"
KNOWLEDGE_DIR = SKILLS_ROOT / "knowledge"
# Standalone skill (not part of the cre-agent-skills-main bundle) covering
# hold/sell/refi/1031/JV scenario comparison + sensitivity analysis — used
# only for the full report (strategy/play section), not the lighter chat.
SCENARIO_MATRIX_FILE = Path(__file__).parent / "scenario-matrix-analyzer" / "SKILL.md"

ANTHROPIC_MODEL = "claude-sonnet-4-5-20250929"

# The "quick underwrite" chain recommended by the skills library's own README
# (document ingestion -> due diligence validation -> underwriting model ->
# IC memo synthesis). Loading just this chain keeps the system prompt to a
# manageable size instead of pulling in all 66 skills across every property
# type/role (industrial, office, brokerage, capital markets, etc.).
CORE_SKILL_FILES = [
    "document-ingestion/document-classifier.md",
    "document-ingestion/rent-roll-parser.md",
    "document-ingestion/financials-parser.md",
    "document-ingestion/offering-memo-parser.md",
    "due-diligence/rent-roll-analyst.md",
    "due-diligence/opex-analyst.md",
    "due-diligence/market-study.md",
    "underwriting/financial-model-builder.md",
    "underwriting/scenario-analyst.md",
    "underwriting/ic-memo-writer.md",
]
CORE_KNOWLEDGE_FILES = [
    "underwriting-calc.md",
    "risk-scoring.md",
    "multifamily-benchmarks.md",
]

_file_cache: Dict[str, str] = {}


def _read(path: Path) -> str:
    key = str(path)
    if key in _file_cache:
        return _file_cache[key]
    try:
        text = path.read_text(encoding="utf-8")
    except Exception as e:
        log.warning("[Napkin] Failed to read %s: %s", path, e)
        text = ""
    _file_cache[key] = text
    return text


def _get_client() -> Anthropic:
    api_key = os.getenv("ANTHROPIC_API_KEY") or os.getenv("CLAUDE_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY (or CLAUDE_API_KEY) not configured")
    return Anthropic(api_key=api_key)


def _load_skills_bundle() -> str:
    """Concatenate the core skill + knowledge files into one reference block.
    Cached per-file on first read (module-level dict), so this is cheap on
    every chat call after the first."""
    parts: List[str] = []
    for rel in CORE_SKILL_FILES:
        text = _read(SKILLS_DIR / rel)
        if text:
            parts.append(f"### SKILL: {rel}\n\n{text}")
    for rel in CORE_KNOWLEDGE_FILES:
        text = _read(KNOWLEDGE_DIR / rel)
        if text:
            parts.append(f"### KNOWLEDGE BASE: {rel}\n\n{text}")
    return "\n\n---\n\n".join(parts)


def _load_report_skills_bundle() -> str:
    """Same core bundle as chat, plus the standalone scenario-matrix-analyzer
    skill (hold/sell/refi/1031/JV comparison + sensitivity analysis) — used
    for the full report's strategy/play section. Kept separate from
    _load_skills_bundle() so the lighter chat prompt is untouched."""
    parts = [_load_skills_bundle()]
    scenario_text = _read(SCENARIO_MATRIX_FILE)
    if scenario_text:
        parts.append(f"### SKILL: scenario-matrix-analyzer/SKILL.md\n\n{scenario_text}")
    return "\n\n---\n\n".join(p for p in parts if p)


def _extract_json(content: str) -> str:
    content = content.strip()
    if "```json" in content:
        content = content.split("```json")[1].split("```")[0]
    elif "```" in content:
        content = content.split("```")[1].split("```")[0]
    content = content.strip()
    if content[:1] != "{" or content[-1:] != "}":
        start = content.find("{")
        end = content.rfind("}")
        if start != -1 and end != -1 and end > start:
            content = content[start:end + 1]
    return content.strip()


NAPKIN_REPORT_SYSTEM_PROMPT_TEMPLATE = """You are "Back of the Napkin" — a fast, rigorous CRE underwriting \
analyst inside DealSniper. A sponsor just uploaded a real deal document (OM, T-12, rent roll, etc.) and \
wants a COMPLETE first-pass underwrite immediately — not a chat, a full read: what's wrong with the OM's \
numbers, what the market looks like and where it's headed, where the actual play/strategy is in this deal, \
what they should actually pay for it, and whether they can raise investor capital and pay it back.

You work ENTIRELY from the CRE Agent Skills playbooks below (document ingestion, due diligence, \
underwriting, and scenario-analysis skills, plus knowledge bases of formulas/benchmarks/risk scoring). \
Follow their strategies, formulas, thresholds, and output structures exactly — they are your only source \
of underwriting methodology. Do not invent your own formulas or shortcuts.

CRITICAL RULES:
1. Only use real numbers extracted from the uploaded document. Never invent property specifics (address, \
   unit count, rents, expenses, price) that weren't actually in the material — if something critical is \
   missing, say so in missingCriticalData instead of guessing.
2. Actively hunt for OM red flags per the due-diligence skills below: rent roll inconsistencies (lease \
   expirations, concessions, delinquencies, unit mix mismatches vs. stated totals), expense ratios that \
   are unrealistically low vs. the multifamily benchmarks knowledge base, T-12 vs. pro forma NOI gaps that \
   aren't explained, vacancy/occupancy math that doesn't reconcile, and any numbers that look inflated or \
   understated versus the benchmarks. Every issue must cite the specific number(s) that triggered it.
3. Where a benchmark comparison is needed (market rent, expense ratios, cap rates) and the sponsor's \
   market data isn't in the document, use the knowledge bases below and say so explicitly.
4. Recommend a maximum purchase price using the underwriting-calc knowledge base's methodology (target \
   cap rate / DSCR / cash-on-cash thresholds against the deal's real NOI), and show your reasoning.
5. Use the scenario-matrix-analyzer skill's framework to identify the actual strategic play (hold, \
   value-add reposition, refinance-and-hold, etc.) appropriate for a NEW acquisition — adapt its \
   hold/sell/refi comparison logic to "how should this specific deal be played" rather than an existing \
   asset's exit decision.
6. For investor feasibility, reason about whether the deal's projected cash flow can cover a typical \
   preferred return (8% is a reasonable default if the sponsor didn't specify one) and return capital, \
   using the risk-scoring knowledge base.
7. Be direct. If the deal is bad, say so plainly — no hedging filler. Every field should read like a real \
   analyst wrote it, not a hedge-everything disclaimer generator.

Return ONLY a single JSON object matching this exact shape, no markdown fences, no prose outside the JSON. KEEP EVERY TEXT FIELD SHORT — this is a napkin read, not a memo — so the full response fits comfortably within the output budget:
{{
  "headline": "1-2 sentences, executive summary",
  "verdict": "PURSUE | PASS | PURSUE WITH CONDITIONS",
  "confidence": "low | medium | high",
  "dealSnapshot": {{ "propertyName": "string or null", "address": "string or null", "units": 0, "askingPrice": 0, "askingPricePerUnit": 0, "statedCapRate": 0, "statedNOI": 0 }},
  "omIssues": [ {{ "severity": "critical | moderate | minor", "category": "Income | Expenses | Rent Roll | Vacancy | Structural | Other", "issue": "short label", "detail": "1 sentence citing the specific numbers", "recommendation": "1 short sentence" }} ] (max 6 issues, most important first),
  "marketOutlook": {{ "summary": "1-2 sentences", "trend": "improving | stable | declining", "keyDrivers": ["short phrase", "..."] (max 3), "risks": ["short phrase", "..."] (max 3) }},
  "strategy": {{ "play": "short label, e.g. 'Value-add reposition, hold 5 years'", "rationale": "1-2 sentences", "keySteps": ["short phrase", "..."] (max 4) }},
  "valuation": {{ "askingPrice": 0, "recommendedMaxPrice": 0, "impliedGoingInCapAtRecommendedPrice": 0, "rationale": "1-2 sentences" }},
  "investorFeasibility": {{ "canRaiseCapital": "yes | no | conditional", "assumedPreferredReturn": 0.08, "projectedCashOnCash": 0, "dscrAdequate": true, "rationale": "1-2 sentences" }},
  "missingCriticalData": ["short phrase", "..."] (max 5),
  "nextSteps": ["short phrase", "..."] (max 5),
  "meta": {{ "disclaimer": "Automated first-pass read from uploaded document text only — not a substitute for full due diligence or a licensed underwriting review." }}
}}

────────────────────────────────────────────────────────
CRE AGENT SKILLS + KNOWLEDGE BASES (read-only reference)
────────────────────────────────────────────────────────

{skills_bundle}
"""


def generate_napkin_report(document_text: str) -> Dict[str, Any]:
    """One Claude call: turn an uploaded deal document straight into a full
    structured underwrite report (OM issues, market outlook, strategy/play,
    recommended purchase price, investor payback feasibility) — no chat
    round-trip required. This is the primary Back of the Napkin flow; chat
    is a secondary follow-up feature on top of the same document."""
    if not document_text or not document_text.strip():
        raise ValueError("document_text is required")

    system_prompt = NAPKIN_REPORT_SYSTEM_PROMPT_TEMPLATE.format(skills_bundle=_load_report_skills_bundle())
    user_content = (
        "UPLOADED DEAL MATERIAL (raw extracted text, read-only):\n\n" + document_text[:80000] +
        "\n\nReturn the full JSON report now."
    )

    client = _get_client()
    response = client.messages.create(
        model=ANTHROPIC_MODEL,
        max_tokens=4000,
        system=system_prompt,
        messages=[{"role": "user", "content": user_content}],
    )
    content = response.content[0].text if response.content else "{}"

    try:
        result = json.loads(_extract_json(content))
    except json.JSONDecodeError:
        log.warning("[Napkin] Report JSON parse failed, retrying once...")
        response2 = client.messages.create(
            model=ANTHROPIC_MODEL,
            max_tokens=4000,
            system=system_prompt + "\n\nIMPORTANT: your previous response was not valid JSON, or was cut off before it finished. Return ONLY the JSON object, nothing else, and keep every text field to 1-2 short sentences so the whole response fits well within the output budget.",
            messages=[{"role": "user", "content": user_content}],
        )
        content2 = response2.content[0].text if response2.content else "{}"
        result = json.loads(_extract_json(content2))

    result.setdefault("meta", {})
    result["meta"].setdefault(
        "disclaimer",
        "Automated first-pass read from uploaded document text only — not a substitute for full due diligence or a licensed underwriting review.",
    )
    return result


NAPKIN_SYSTEM_PROMPT_TEMPLATE = """You are "Back of the Napkin" — a fast, rigorous CRE underwriting analyst \
inside DealSniper. You do quick-turn deal reads for a sponsor who just uploaded or pasted real deal \
material and wants a fast, defensible first opinion — the kind of read you'd scratch out on a napkin \
before deciding whether a deal is even worth a full underwrite.

You work ENTIRELY from the CRE Agent Skills playbooks below (document ingestion, due diligence, and \
underwriting skills, plus core knowledge bases of formulas/benchmarks/risk scoring). Follow their \
strategies, formulas, thresholds, and output formats exactly — they are your only source of underwriting \
methodology. Do not invent your own formulas or shortcuts.

CRITICAL RULES:
1. Only use real numbers from the deal material the sponsor has provided (pasted text or uploaded \
   document). Never invent property specifics (address, unit count, rents, expenses) that weren't given.
2. Where a skill calls for a benchmark comparison (market rent, expense ratios, cap rates, etc.) and the \
   sponsor hasn't provided market data, use the benchmarks in the knowledge bases below and say so \
   explicitly (e.g. "per the multifamily benchmarks knowledge base...").
3. Ask for anything critical that's missing (e.g. purchase price, T-12, rent roll) rather than guessing.
4. Keep it conversational and fast — a back-of-the-napkin read, not a 20-page IC memo — unless the \
   sponsor explicitly asks for the full IC Memo Writer output.
5. When the sponsor does ask for the full memo, follow the IC Memo Writer skill's exact output structure.

────────────────────────────────────────────────────────
CRE AGENT SKILLS + KNOWLEDGE BASES (read-only reference)
────────────────────────────────────────────────────────

{skills_bundle}
"""


def chat_about_napkin_deal(
    document_text: Optional[str],
    message: str,
    history: Optional[List[Dict[str, Any]]] = None,
) -> str:
    """One Claude call: answer/underwrite using ONLY the CRE Agent Skills
    library as methodology, grounded in whatever deal material the sponsor
    has pasted/uploaded so far. Stateless per call — full history and deal
    text are resent every turn since the API has no server-side memory."""
    if not message or not message.strip():
        raise ValueError("message is required")

    system_prompt = NAPKIN_SYSTEM_PROMPT_TEMPLATE.format(skills_bundle=_load_skills_bundle())

    history_text = ""
    if history:
        lines = []
        for turn in history[-20:]:
            role = turn.get("role", "user")
            content = turn.get("content", "")
            if not content:
                continue
            lines.append(f"{'Sponsor' if role == 'user' else 'Analyst'}: {content}")
        history_text = "\n".join(lines)

    user_content = ""
    if document_text:
        user_content += "UPLOADED DEAL MATERIAL (raw extracted text, read-only):\n" + document_text[:60000] + "\n\n"
    if history_text:
        user_content += "CONVERSATION SO FAR:\n" + history_text + "\n\n"
    user_content += f"Sponsor's new message: {message}\n\nRespond now."

    client = _get_client()
    response = client.messages.create(
        model=ANTHROPIC_MODEL,
        max_tokens=2500,
        system=system_prompt,
        messages=[{"role": "user", "content": user_content}],
    )
    return response.content[0].text if response.content else "Sorry, I couldn't generate a response."


def extract_document_text(filename: str, content: bytes) -> str:
    """Best-effort raw text extraction for a quick napkin read. Deliberately
    simple and standalone — does not reuse or affect the platform's main
    document parsing pipeline (parser_v4.py / deal_manager_parser.py)."""
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""

    if ext == "pdf":
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(content))
        pages = []
        for page in reader.pages:
            try:
                pages.append(page.extract_text() or "")
            except Exception:
                continue
        return "\n\n".join(pages).strip()

    if ext in ("xlsx", "xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"--- Sheet: {ws.title} ---")
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    lines.append(", ".join("" if c is None else str(c) for c in row))
        return "\n".join(lines).strip()

    # csv / txt / anything else plain-text
    return content.decode("utf-8", errors="ignore").strip()
