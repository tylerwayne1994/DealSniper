"""
LLM-Powered Excel Export for DealSniper
Uses Claude to intelligently map deal data to the underwriting template.
Matches the exact layout of client/public/underwriting.xlsx
Charges 1 token per export.
"""

import io
import os
import json
import logging
import copy
from typing import Dict, Any, Optional, List
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from anthropic import Anthropic

log = logging.getLogger(__name__)

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")

# Template cell mappings - maps template cells to data paths
TEMPLATE_PATH = Path(__file__).parent.parent / "client" / "public" / "underwriting.xlsx"


def get_anthropic_client():
    """Get Anthropic client."""
    if not ANTHROPIC_API_KEY:
        raise ValueError("ANTHROPIC_API_KEY not configured")
    return Anthropic(api_key=ANTHROPIC_API_KEY)


EXCEL_SYSTEM_PROMPT = """You are an expert commercial real estate underwriter.
Your task is to map deal data to specific cells in an existing Excel template.

The template has a single sheet "Underwriting Model" with this layout:
- Row 1: Title "DEAL NAME - INVESTMENT UNDERWRITING MODEL"
- Row 2: Address, City, ST ZIP, Units, Year Built, etc.
- Rows 4-12: Deal Snapshot (left Col A-B) + Financial Performance (right Col G-H)
- Row 16+: ASSUMPTIONS section
- Row 18-29: Property Information (left) + Cashflow Summary (right)
- Row 31-58: Purchase & Financing, Equity Partner Structure, Value-Add Assumptions
- Row 62+: CASHFLOW SCENARIOS with 3 columns (Current, RUBS+Trash, Full Value-Add)
- Row 69-85: EXPENSES breakdown
- Row 89-99: DEBT SERVICE & CASHFLOW, RETURNS

You will receive JSON deal data. Map it to template cells by returning a JSON object like:
{
  "cells": {
    "B5": 850000,  // Purchase Price
    "B26": 12,     // Total Units
    "B63": 110775, // Scheduled Gross Income
    ...etc
  },
  "title": "Property Name - INVESTMENT UNDERWRITING MODEL",
  "address_line": "521 SW 24th St, Oklahoma City, OK | 12 Units | 1962"
}

Important cell mappings to fill:
- B5: Purchase Price
- B11: Total Units (for Price Per Unit formula =B5/B11)
- B19: Property Name
- B20: Address
- B26: Total Units
- B28: Current Occupancy (decimal like 0.95)
- B32: Purchase Price (again, for formulas)
- B33: Down Payment % (decimal)
- B34: Interest Rate (decimal)
- B35: Amortization Years
- B36: Closing Costs %
- B46: Investor Contribution $
- B48: Preferred Return %
- B49: Partner Equity %
- B50: Your Equity %
- B52: Rent Bump Per Unit Per Month $
- B53: RUBS Recovery Per Year $
- B54: Trash Fee Per Unit Per Month $
- B55: Annual Appreciation Rate
- B56: Exit Cap Conservative
- B57: Exit Cap Market
- B63: Scheduled Gross Income
- B70: Property Tax
- B71: Insurance
- B73: Repairs & Maintenance
- B75: Landscaping
- B77: Water/Sewer
- B78: Electricity
- B80: Trash Service
- B81: General Admin

The template has formulas that auto-calculate:
- B6 (Price Per Unit), B7 (Cash to Close), B8 (Loan Amount), B9 (Monthly Debt Service)
- B38 (Down Payment $), B40 (Loan Amount), B41 (Closing Costs), B42 (Total Cash to Close)
- B43 (Monthly Debt Service), B44 (Annual Debt Service)
- All NOI, cashflow, returns are calculated from inputs

Return ONLY valid JSON. Map as many cells as you can from the provided data."""


def load_template():
    """Load the underwriting template from client/public."""
    # Try multiple paths to find the template
    possible_paths = [
        TEMPLATE_PATH,
        Path(__file__).parent / "underwriting_template.xlsx",
        Path("client/public/underwriting.xlsx"),
    ]
    
    for path in possible_paths:
        if path.exists():
            return openpyxl.load_workbook(str(path))
    
    # If template not found, create a basic structure
    log.warning("Template not found, creating basic structure")
    return None


def fill_template_from_llm(wb: openpyxl.Workbook, llm_response: Dict[str, Any], deal_data: Dict[str, Any]) -> openpyxl.Workbook:
    """Fill template cells based on LLM mapping."""
    if "Underwriting Model" not in wb.sheetnames:
        log.error("Template missing 'Underwriting Model' sheet")
        return wb
    
    ws = wb["Underwriting Model"]
    
    # Update title
    if llm_response.get("title"):
        ws.cell(row=1, column=1).value = llm_response["title"]
    
    # Update address line
    if llm_response.get("address_line"):
        ws.cell(row=2, column=1).value = llm_response["address_line"]
    
    # Fill cells from LLM mapping
    cells = llm_response.get("cells", {})
    for cell_ref, value in cells.items():
        try:
            # Parse cell reference (e.g., "B5" -> row 5, col 2)
            col_letter = ''.join(c for c in cell_ref if c.isalpha())
            row_num = int(''.join(c for c in cell_ref if c.isdigit()))
            
            if value is not None:
                ws[cell_ref] = value
                log.debug(f"Set {cell_ref} = {value}")
        except Exception as e:
            log.warning(f"Failed to set cell {cell_ref}: {e}")
    
    return wb


def create_workbook_from_scratch(llm_response: Dict[str, Any], deal_data: Dict[str, Any]) -> openpyxl.Workbook:
    """Create workbook matching template layout when template file not available."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Underwriting Model"
    
    # Styling
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    section_font = Font(bold=True, size=10)
    input_font = Font(color="0000FF")  # Blue for inputs
    
    cells = llm_response.get("cells", {})
    title = llm_response.get("title", "DEAL NAME — INVESTMENT UNDERWRITING MODEL")
    address_line = llm_response.get("address_line", "Address | City, ST ZIP")
    
    # Row 1: Title
    ws.cell(row=1, column=1, value=title).font = title_font
    
    # Row 2: Address line
    ws.cell(row=2, column=1, value=address_line)
    
    # Row 4: Headers
    ws.cell(row=4, column=1, value="📋  DEAL SNAPSHOT").font = header_font
    ws.cell(row=4, column=7, value="💰  FINANCIAL PERFORMANCE").font = header_font
    
    # Deal Snapshot (Col A-B, Rows 5-12)
    snapshot_labels = [
        ("A5", "Purchase Price"), ("A6", "Price Per Unit"), ("A7", "Total Cash to Close"),
        ("A8", "Loan Amount"), ("A9", "Monthly Debt Service"), ("A10", "Occupancy"),
        ("A11", "Total Units"), ("A12", "Year Built")
    ]
    for cell_ref, label in snapshot_labels:
        ws[cell_ref] = label
    
    # Set values and formulas for Deal Snapshot
    ws["B5"] = cells.get("B5", 0)  # Purchase Price
    ws["B6"] = "=B5/B11"  # Price Per Unit
    ws["B7"] = "=B42"  # Total Cash to Close
    ws["B8"] = "=B40"  # Loan Amount
    ws["B9"] = "=B43"  # Monthly Debt Service
    ws["B10"] = "=B28"  # Occupancy
    ws["B11"] = cells.get("B11", cells.get("B26", 0))  # Total Units
    ws["B12"] = cells.get("B12", "")  # Year Built
    
    # Financial Performance (Col G-H, Rows 5-12)
    fin_labels = [
        ("G5", "NOI — Current Expenses"), ("G6", "NOI — Full Value-Add"),
        ("G7", "Cap Rate — Current"), ("G8", "Cap Rate — Value-Add"),
        ("G9", "DSCR"), ("G10", "Annual Debt Service"),
        ("G11", "Interest Rate"), ("G12", "Amortization")
    ]
    for cell_ref, label in fin_labels:
        ws[cell_ref] = label
    
    ws["H5"] = "=B87"  # NOI Current
    ws["H6"] = "=D87"  # NOI Full Value-Add
    ws["H7"] = "=B95"  # Cap Rate Current
    ws["H8"] = "=D95"  # Cap Rate Value-Add
    ws["H9"] = "=B96"  # DSCR
    ws["H10"] = "=B44"  # Annual Debt Service
    ws["H11"] = cells.get("B34", 0.07)  # Interest Rate
    ws["H12"] = cells.get("B35", 30)  # Amortization
    
    # Row 16: ASSUMPTIONS header
    ws.cell(row=16, column=1, value="━━━━━━━━━━━━━━━━━━━━━━━  ASSUMPTIONS  ━━━━━━━━━━━━━━━━━━━━━━━").font = section_font
    
    # Property Information (Rows 18-29)
    ws.cell(row=18, column=1, value="PROPERTY INFORMATION").font = header_font
    ws["A19"] = "Property Name"
    ws["B19"] = cells.get("B19", "")
    ws["A20"] = "Address"
    ws["B20"] = cells.get("B20", "")
    ws["A21"] = "Asset Type"
    ws["A22"] = "Year Built / Renovated"
    ws["A23"] = "APN"
    ws["A24"] = "Gross Building SF"
    ws["A25"] = "Lot Size (Acres)"
    ws["A26"] = "Total Units"
    ws["B26"] = cells.get("B26", 0)
    ws["A27"] = "Unit Mix"
    ws["A28"] = "Current Occupancy"
    ws["B28"] = cells.get("B28", 0.95)
    ws["A29"] = "Average SF Per Unit"
    ws["A30"] = "Average Current Rent/Unit"
    
    # Purchase & Financing (Rows 31-44)
    ws.cell(row=31, column=1, value="PURCHASE & FINANCING").font = header_font
    ws["A32"] = "Purchase Price"
    ws["B32"] = cells.get("B32", cells.get("B5", 0))
    ws["A33"] = "Down Payment %"
    ws["B33"] = cells.get("B33", 0.2)
    ws["A34"] = "Interest Rate"
    ws["B34"] = cells.get("B34", 0.07)
    ws["A35"] = "Amortization (Years)"
    ws["B35"] = cells.get("B35", 30)
    ws["A36"] = "Closing Costs %"
    ws["B36"] = cells.get("B36", 0.01)
    ws["A37"] = "Working Capital ($)"
    ws["B37"] = cells.get("B37", 0)
    ws["A38"] = "Down Payment ($)"
    ws["B38"] = "=B32*B33"
    ws["A39"] = "CALCULATED FINANCING"
    ws["A40"] = "Loan Amount"
    ws["B40"] = "=B32-B38"
    ws["A41"] = "Closing Costs ($)"
    ws["B41"] = "=B32*B36"
    ws["A42"] = "Total Cash to Close"
    ws["B42"] = "=B38+B41+B37"
    ws["A43"] = "Monthly Debt Service"
    ws["B43"] = "=IFERROR(B40*(B34/12)/(1-(1+B34/12)^(-B35*12)),0)"
    ws["A44"] = "Annual Debt Service"
    ws["B44"] = "=B43*12"
    
    # Equity Partner Structure (Rows 45-50)
    ws.cell(row=45, column=1, value="EQUITY PARTNER STRUCTURE").font = header_font
    ws["A46"] = "Investor Contribution ($)"
    ws["B46"] = cells.get("B46", 0)
    ws["A47"] = "Your Contribution ($)"
    ws["B47"] = cells.get("B47", 0)
    ws["A48"] = "Preferred Return %"
    ws["B48"] = cells.get("B48", 0.08)
    ws["A49"] = "Partner Equity %"
    ws["B49"] = cells.get("B49", 0.2)
    ws["A50"] = "Your Equity %"
    ws["B50"] = cells.get("B50", 0.8)
    
    # Value-Add Assumptions (Rows 51-58)
    ws.cell(row=51, column=1, value="VALUE-ADD ASSUMPTIONS").font = header_font
    ws["A52"] = "Rent Bump / Unit / Mo ($)"
    ws["B52"] = cells.get("B52", 0)
    ws["A53"] = "RUBS Recovery / Year ($)"
    ws["B53"] = cells.get("B53", 0)
    ws["A54"] = "Trash Fee / Unit / Mo ($)"
    ws["B54"] = cells.get("B54", 0)
    ws["A55"] = "Annual Appreciation Rate"
    ws["B55"] = cells.get("B55", 0.03)
    ws["A56"] = "Exit Cap — Conservative"
    ws["B56"] = cells.get("B56", 0.075)
    ws["A57"] = "Exit Cap — Market"
    ws["B57"] = cells.get("B57", 0.065)
    ws["A58"] = "Hold Period (Years)"
    ws["B58"] = cells.get("B58", 5)
    
    # Cashflow Scenarios Header (Row 60)
    ws.cell(row=60, column=1, value="━━━━━━━━━━━━  CASHFLOW SCENARIOS  ━━━━━━━━━━━━").font = section_font
    
    # Scenario headers (Row 62)
    ws["B62"] = "Scenario 1\nCurrent Expenses"
    ws["C62"] = "Scenario 2\nRUBS + Trash"
    ws["D62"] = "Scenario 3\nFull Value-Add\n(+Rent Bump)"
    
    # Income section (Rows 63-68)
    ws["A63"] = "Scheduled Gross Income"
    ws["B63"] = cells.get("B63", 0)
    ws["C63"] = cells.get("C63", cells.get("B63", 0))
    ws["D63"] = f"=B63+B52*12*B26*B28"
    
    ws["A64"] = "Less: Vacancy (10%)"
    ws["B64"] = "=-B63*0.1"
    ws["C64"] = "=-C63*0.1"
    ws["D64"] = "=-D63*0.1"
    
    ws["A65"] = "Other Income"
    ws["A66"] = "RUBS Recovery"
    ws["C66"] = "=B53"
    ws["D66"] = "=B53"
    
    ws["A67"] = "Trash Fee Income"
    ws["C67"] = "=B54*12*B26*B28"
    ws["D67"] = "=B54*12*B26*B28"
    
    ws["A68"] = "GROSS OPERATING INCOME"
    ws["B68"] = "=SUM(B63:B67)"
    ws["C68"] = "=SUM(C63:C67)"
    ws["D68"] = "=SUM(D63:D67)"
    
    # Expenses section (Rows 69-85)
    ws.cell(row=69, column=1, value="EXPENSES").font = header_font
    expense_rows = [
        ("A70", "Property Tax", "B70"),
        ("A71", "Insurance", "B71"),
        ("A72", "Management (6% of GOI)", "B72"),
        ("A73", "Repairs & Maintenance", "B73"),
        ("A74", "Turnover / Cleaning", "B74"),
        ("A75", "Landscaping", "B75"),
        ("A76", "Payroll", "B76"),
        ("A77", "Water / Sewer", "B77"),
        ("A78", "Electricity", "B78"),
        ("A79", "Pest Control", "B79"),
        ("A80", "Trash Service", "B80"),
        ("A81", "General Admin", "B81"),
        ("A82", "Marketing", "B82"),
        ("A83", "Reserves", "B83"),
        ("A84", "Licenses & Permits", "B84"),
    ]
    for label_cell, label, value_cell in expense_rows:
        ws[label_cell] = label
        val = cells.get(value_cell, 0)
        ws[value_cell] = val if val else 0
        # Copy to other scenarios
        col_b = value_cell[1:]
        ws[f"C{col_b}"] = val if val else 0
        ws[f"D{col_b}"] = val if val else 0
    
    ws["A85"] = "TOTAL EXPENSES"
    ws["B85"] = "=SUM(B70:B84)"
    ws["C85"] = "=SUM(C70:C84)"
    ws["D85"] = "=SUM(D70:D84)"
    
    # NOI (Row 87)
    ws["A87"] = "NET OPERATING INCOME"
    ws["B87"] = "=B68-B85"
    ws["C87"] = "=C68-C85"
    ws["D87"] = "=D68-D85"
    
    # Debt Service & Cashflow (Rows 89-92)
    ws.cell(row=89, column=1, value="DEBT SERVICE & CASHFLOW").font = header_font
    ws["A90"] = "Annual Debt Service"
    ws["B90"] = "=B44"
    ws["C90"] = "=B44"
    ws["D90"] = "=B44"
    ws["A91"] = "Annual Cashflow"
    ws["B91"] = "=B87-B90"
    ws["C91"] = "=C87-C90"
    ws["D91"] = "=D87-D90"
    ws["A92"] = "Monthly Cashflow"
    ws["B92"] = "=B91/12"
    ws["C92"] = "=C91/12"
    ws["D92"] = "=D91/12"
    
    # Returns (Rows 94-99)
    ws.cell(row=94, column=1, value="RETURNS").font = header_font
    ws["A95"] = "Cap Rate"
    ws["B95"] = "=B87/B32"
    ws["C95"] = "=C87/B32"
    ws["D95"] = "=D87/B32"
    ws["A96"] = "DSCR"
    ws["B96"] = "=B87/B44"
    ws["C96"] = "=C87/B44"
    ws["D96"] = "=D87/B44"
    ws["A97"] = "Cash-on-Cash (Your Investment)"
    ws["B97"] = "=B91/B47"
    ws["C97"] = "=C91/B47"
    ws["D97"] = "=D91/B47"
    ws["A98"] = "Cash-on-Cash (Partner)"
    ws["B98"] = "=B91/B46"
    ws["C98"] = "=C91/B46"
    ws["D98"] = "=D91/B46"
    ws["A99"] = "Price Per Unit"
    ws["B99"] = "=B32/B26"
    ws["C99"] = "=B32/B26"
    ws["D99"] = "=B32/B26"
    
    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 18
    
    return wb


def llm_export_to_excel(
    scenario_data: Dict[str, Any],
    full_calcs: Dict[str, Any] = None,
    sensitivity_data: Dict[str, Any] = None,
    waterfall_data: Dict[str, Any] = None
) -> io.BytesIO:
    """
    Use Claude to map deal data to template cells, then fill the template.
    
    Args:
        scenario_data: Full scenarioData from frontend
        full_calcs: Calculated values from frontend
        sensitivity_data: Stress test data (revaluation, refi, post-refi cashflow)
        waterfall_data: Equity waterfall structure
    
    Returns:
        BytesIO buffer containing Excel file
    """
    client = get_anthropic_client()
    
    # Extract relevant sections
    property_info = scenario_data.get('property', {})
    pricing = scenario_data.get('pricing_financing', {})
    pnl = scenario_data.get('pnl', {})
    expenses = scenario_data.get('expenses', {})
    unit_mix = scenario_data.get('unit_mix', [])
    value_add = scenario_data.get('value_add', {})
    financing = scenario_data.get('financing', {})
    
    # Build the data payload for Claude
    deal_data = {
        "property": property_info,
        "pricing": {
            "purchase_price": pricing.get('purchase_price') or pricing.get('price'),
            "loan_amount": pricing.get('loan_amount'),
            "down_payment_pct": pricing.get('down_payment_pct') or pricing.get('down_payment'),
            "interest_rate": pricing.get('interest_rate'),
            "amortization_years": pricing.get('amortization_years'),
            "term_years": pricing.get('term_years'),
            "closing_costs_pct": pricing.get('closing_costs_pct'),
        },
        "income": {
            "gross_scheduled_income": pnl.get('gross_scheduled_income') or pnl.get('gsi'),
            "effective_gross_income": pnl.get('effective_gross_income') or pnl.get('egi'),
            "noi": pnl.get('noi') or pnl.get('noi_t12'),
            "vacancy_rate": pnl.get('vacancy_rate'),
        },
        "expenses": expenses,
        "value_add": value_add,
        "unit_mix": unit_mix,
        "equity_structure": financing.get('equity_structure', {}),
        "waterfall": waterfall_data,
    }
    
    if full_calcs:
        deal_data["calculations"] = {
            "cap_rate": full_calcs.get('capRate'),
            "cash_on_cash": full_calcs.get('cashOnCash'),
            "dscr": full_calcs.get('dscr'),
            "irr": full_calcs.get('irr'),
            "equity_multiple": full_calcs.get('equityMultiple'),
        }
    
    user_prompt = f"""Map this commercial real estate deal data to template cells.

DEAL DATA:
{json.dumps(deal_data, indent=2, default=str)}

Return a JSON object mapping cell references to values. Example:
{{
  "title": "Walker Apartments — INVESTMENT UNDERWRITING MODEL",
  "address_line": "521 SW 24th St, Oklahoma City, OK | 12 Units | 1962",
  "cells": {{
    "B5": 850000,
    "B11": 12,
    "B19": "Walker Apartments",
    "B20": "521 SW 24th St, Oklahoma City, OK 73109",
    "B26": 12,
    "B28": 0.95,
    "B32": 850000,
    "B33": 0.2,
    "B34": 0.07,
    "B35": 30,
    "B36": 0.01,
    "B46": 178000,
    "B47": 50000,
    "B48": 0.08,
    "B49": 0.2,
    "B50": 0.8,
    "B52": 50,
    "B53": 2400,
    "B54": 25,
    "B56": 0.075,
    "B57": 0.065,
    "B63": 110775,
    "B70": 7560,
    "B71": 5544,
    "B73": 11906,
    "B75": 900,
    "B77": 5209,
    "B78": 11207,
    "B80": 2794,
    "B81": 1625
  }}
}}

Key cell mappings:
- B5/B32: Purchase Price
- B11/B26: Total Units
- B19: Property Name
- B20: Full Address
- B28: Occupancy (decimal 0-1)
- B33: Down Payment % (decimal)
- B34: Interest Rate (decimal)
- B35: Amortization years
- B46: Investor contribution $
- B47: Your contribution $
- B48: Preferred return %
- B49/B50: Partner/Your equity %
- B52: Rent bump per unit per month
- B53: RUBS annual recovery
- B54: Trash fee per unit per month
- B56/B57: Exit cap rates
- B63: Scheduled Gross Income
- B70-B84: Individual expense line items

Use the actual numbers from the data. Return ONLY valid JSON."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2048,
            messages=[
                {"role": "user", "content": user_prompt}
            ],
            system=EXCEL_SYSTEM_PROMPT
        )
        
        content = response.content[0].text
        
        # Extract JSON from response
        try:
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0]
            elif "```" in content:
                content = content.split("```")[1].split("```")[0]
            
            llm_response = json.loads(content.strip())
        except json.JSONDecodeError as e:
            log.error(f"Failed to parse LLM response: {e}")
            log.error(f"Content: {content[:500]}")
            llm_response = {"cells": {}, "title": "Underwriting Model"}
        
        # Try to load template, or create from scratch
        template_wb = load_template()
        
        if template_wb:
            # Fill the template
            wb = fill_template_from_llm(template_wb, llm_response, deal_data)
        else:
            # Create workbook matching template layout
            wb = create_workbook_from_scratch(llm_response, deal_data)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
        
    except Exception as e:
        log.error(f"LLM export failed: {e}")
        import traceback
        traceback.print_exc()
        
        # Fallback: create basic workbook
        wb = create_workbook_from_scratch({"cells": {}, "title": "Underwriting Model"}, deal_data)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
