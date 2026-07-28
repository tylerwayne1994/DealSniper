"""
AI-Powered Spreadsheet Builder for DealSniper
--------------------------------------------
Uses Claude to design ONE complete, well-organized "workbook spec" from all
available deal data (parsed OM data + the platform's own calculated returns,
sensitivity/stress tests, and equity waterfall). The same spec is then
rendered to either an Excel file (openpyxl) or a Google Sheet (via the
existing service-account Sheets API client), so the "Sheets" and "Excel"
export buttons produce the same curated, formatted spreadsheet.
"""
import io
import os
import re
import json
import logging
from typing import Dict, Any, List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from anthropic import Anthropic

log = logging.getLogger(__name__)

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY") or os.getenv("CLAUDE_API_KEY")
ANTHROPIC_MODEL = os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-5-20250929")


def _get_client():
    if not ANTHROPIC_API_KEY:
        raise ValueError("ANTHROPIC_API_KEY not configured")
    return Anthropic(api_key=ANTHROPIC_API_KEY)


SPEC_SYSTEM_PROMPT = """You are an expert commercial real estate underwriter and financial analyst \
who builds clean, professional investment underwriting spreadsheets.

You will be given a JSON blob containing everything known about a real estate deal: the parsed \
property/financing/income/expense data extracted from an offering memorandum, plus the platform's \
own calculated returns (cap rate, cash-on-cash, DSCR, IRR, equity multiple), sensitivity/stress-test \
results, equity waterfall structure, rent comps and market data.

Your job is to design a complete, multi-sheet workbook that presents ALL of this information in a \
clear, well-organized, professional layout — the kind of spreadsheet a real GP would hand to an LP \
investor. Use your judgment: only include sheets/sections for data that is actually present, order \
sheets logically (Summary first), and put units ($ / % / years) in the labels, not in the values.

Return ONLY a JSON object with this exact shape, nothing else:
{
  "workbookTitle": "Property Name — Investment Underwriting Model",
  "sheets": [
    {
      "title": "Summary",
      "rows": [
        ["DEAL SUMMARY", ""],
        ["Property", "123 Main St, City, ST"],
        ["Purchase Price", 1250000],
        ["", ""],
        ["KEY RETURNS", ""],
        ["Cap Rate", 0.062]
      ],
      "headerRows": [0, 4],
      "columnFormats": {"B": "currency"}
    }
  ]
}

Rules:
- "rows" is a 2D array of plain strings and numbers only (no formulas, no nested objects).
- Percentages must be plain decimals (e.g. 0.065 for 6.5%), with that column marked "percent" in \
  columnFormats so it renders correctly — never bake a "%" symbol into the value itself.
- CAUTION on percent units in the input data: some source fields are percent-scaled (e.g. \
  vacancy_rate: 5.0 means 5%, ltv: 75 means 75%, expense_ratio: 46.8 means 46.8%) while others are \
  already decimals (e.g. interest_rate: 0.065). Use the field's magnitude and real-estate common \
  sense to tell them apart, and ALWAYS normalize to a decimal in your output (5.0% -> 0.05).
- Fields equal to 0 in the input frequently mean "not provided", not an actual zero (e.g. price: 0, \
  cap_rate: 0). Omit those line items or leave the value blank ("") instead of showing $0 or 0%.
- "headerRows" lists 0-based row indices within that sheet that are section headers/titles (they \
  will be bolded and shaded) — the first row of every sheet must be a title row included here, plus \
  any section dividers and table column-header rows you add.
- Use blank spacer rows (["", ""]) between sections so the sheet breathes.
- For tabular data (unit mix, rent roll, expenses, projections, sensitivity grids), lay it out as a \
  real table: one column-header row (add its index to headerRows) followed by data rows, one column \
  per field.
- "columnFormats" maps a column letter to one of: "currency", "percent", "number", "text". Only set \
  this for columns that are entirely numeric in that sheet. If a sheet mixes currency and percent \
  values in the same value column (typical for a Summary label/value layout), prefer "currency" and \
  express the percent line items as pre-formatted strings like "6.20%" instead.
- In a label/value sheet whose value column is marked "currency", any non-dollar numeric line item \
  (unit count, year built, square footage, acres, DSCR, equity multiple, years) must be written as \
  a short string instead — e.g. "78", "1960", "56,176 SF", "3.3 acres", "1.45x" — so it doesn't \
  render with a dollar sign.
- Build 4-9 sheets depending on what data is available, for example: Summary, Property & Financing, \
  Unit Mix / Rent Roll (current vs market rent with upside), Income & Expenses (T12 vs pro forma), \
  Cash Flow & Projections, Returns & Sensitivity, Equity Waterfall, Market Data & Comps. Skip any \
  sheet where the underlying data wasn't provided.
- For long schedules (e.g. monthly amortization), summarize to at most ~30 rows (annual summaries), \
  never dump hundreds of rows.
- Never invent numbers that weren't provided. If a value is missing, leave it blank ("") rather than \
  guessing.
Return ONLY valid JSON — no markdown fences, no commentary."""


def _extract_json(content: str) -> str:
    content = (content or "").strip()
    if "```json" in content:
        content = content.split("```json")[1].split("```")[0]
    elif "```" in content:
        content = content.split("```")[1].split("```")[0]
    content = content.strip()
    # Last resort: slice from the first "{" to the last "}"
    if not content.startswith("{"):
        start, end = content.find("{"), content.rfind("}")
        if start != -1 and end > start:
            content = content[start:end + 1]
    return content


def build_workbook_spec(deal_payload: Dict[str, Any]) -> Dict[str, Any]:
    """Call Claude once with the full deal payload and get back a workbook spec."""
    client = _get_client()

    user_prompt = (
        "Build a complete underwriting workbook spec from this deal data:\n\n"
        + json.dumps(deal_payload, indent=2, default=str)
    )

    with client.messages.stream(
        model=ANTHROPIC_MODEL,
        max_tokens=16000,
        messages=[{"role": "user", "content": user_prompt}],
        system=SPEC_SYSTEM_PROMPT,
    ) as stream:
        content = "".join(stream.text_stream)

    try:
        spec = json.loads(_extract_json(content))
    except json.JSONDecodeError as e:
        log.error(f"[SpreadsheetAI] Failed to parse workbook spec: {e}")
        log.error(f"[SpreadsheetAI] Content: {content[:800]}")
        spec = {"workbookTitle": "Underwriting Model", "sheets": []}

    if not spec.get("sheets"):
        spec["sheets"] = [{
            "title": "Summary",
            "rows": [["No data available", "The export could not be generated from this deal."]],
            "headerRows": [0],
            "columnFormats": {},
        }]
    return spec


# ---------------------------------------------------------------------------
#  Excel renderer
# ---------------------------------------------------------------------------

NUMBER_FORMATS = {
    "currency": '"$"#,##0',
    "percent": '0.00%',
    "number": '#,##0.00',
    "text": None,
}


def render_spec_to_excel(spec: Dict[str, Any]) -> io.BytesIO:
    """Render a workbook spec (as produced by build_workbook_spec) to an .xlsx file."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    title_font = Font(bold=True, size=13, color="FFFFFF")
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    section_font = Font(bold=True, size=11)
    section_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

    used_titles = set()
    for sheet_spec in spec.get("sheets", []):
        raw_title = (sheet_spec.get("title") or "Sheet")[:31] or "Sheet"
        title = raw_title
        suffix = 2
        while title in used_titles:
            title = f"{raw_title[:28]}-{suffix}"
            suffix += 1
        used_titles.add(title)

        ws = wb.create_sheet(title=title)
        rows = sheet_spec.get("rows") or []
        header_rows = set(sheet_spec.get("headerRows") or [])
        col_formats = sheet_spec.get("columnFormats") or {}

        max_cols = max((len(r) for r in rows), default=1)
        col_widths = [10] * max_cols

        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=value)
                col_letter = get_column_letter(c_idx + 1)
                fmt = col_formats.get(col_letter)
                if fmt and fmt != "text" and isinstance(value, (int, float)):
                    excel_fmt = NUMBER_FORMATS.get(fmt)
                    if excel_fmt:
                        cell.number_format = excel_fmt
                if r_idx in header_rows:
                    if r_idx == 0:
                        cell.font = title_font
                        cell.fill = title_fill
                    else:
                        cell.font = section_font
                        cell.fill = section_fill
                if c_idx > 0 and isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                col_widths[c_idx] = max(col_widths[c_idx], min(len(str(value)) + 2, 45))

        for c_idx, width in enumerate(col_widths):
            ws.column_dimensions[get_column_letter(c_idx + 1)].width = width

        ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------------------
#  Google Sheets renderer
# ---------------------------------------------------------------------------

GOOGLE_NUMBER_FORMATS = {
    "currency": {"type": "CURRENCY", "pattern": '"$"#,##0'},
    "percent": {"type": "PERCENT", "pattern": "0.00%"},
    "number": {"type": "NUMBER", "pattern": "#,##0.00"},
}

_TITLE_BG = {"red": 0x0F / 255, "green": 0x17 / 255, "blue": 0x2A / 255}
_TITLE_FG = {"red": 1.0, "green": 1.0, "blue": 1.0}
_SECTION_BG = {"red": 0xEC / 255, "green": 0xFD / 255, "blue": 0xF5 / 255}

MAX_SHEET_TITLE_LENGTH = 100
INVALID_SHEET_TITLE_CHARS = r'[\\/*?:\[\]]'


def _service_account_email() -> Optional[str]:
    try:
        raw = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
        if raw:
            return json.loads(raw).get("client_email")
        creds_file = os.path.join(os.path.dirname(__file__), ".google_service_account.json")
        with open(creds_file, "r", encoding="utf-8") as f:
            return json.load(f).get("client_email")
    except Exception:
        return None


def _sheet_title(base_tab_name: Optional[str], title: str) -> str:
    full = f"{base_tab_name} - {title}" if base_tab_name else title
    cleaned = re.sub(INVALID_SHEET_TITLE_CHARS, '-', str(full or '').strip())
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned[:MAX_SHEET_TITLE_LENGTH] or 'Results'


def _col_letter_to_index(letter: str) -> int:
    """'A' -> 0, 'B' -> 1, ..., 'AA' -> 26."""
    idx = 0
    for ch in (letter or "").strip().upper():
        if not ch.isalpha():
            return -1
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


def _normalize_rows(rows: List[List[Any]]) -> List[List[Any]]:
    normalized = []
    for row in rows or []:
        normalized.append([
            value if isinstance(value, (int, float)) else '' if value is None else str(value)
            for value in row
        ])
    return normalized or [['No data available']]


def render_spec_to_google_sheets(spec: Dict[str, Any], sheet_id: str,
                                 base_tab_name: Optional[str] = None) -> Dict[str, Any]:
    """Render a workbook spec into tabs of an existing Google Spreadsheet,
    with the same styling as the Excel renderer (bold shaded headers,
    currency/percent number formats, frozen top row, auto-sized columns)."""
    from googleapiclient.discovery import build as g_build
    from googleapiclient.errors import HttpError
    from google_sheets_updater import get_credentials

    try:
        service = g_build('sheets', 'v4', credentials=get_credentials())

        meta = service.spreadsheets().get(
            spreadsheetId=sheet_id,
            fields='sheets.properties(sheetId,title,gridProperties)'
        ).execute()
        existing = {s['properties']['title']: s['properties'] for s in meta.get('sheets', [])}

        # Resolve final tab titles and grid sizes
        tabs = []
        used_titles = set()
        for sheet_spec in spec.get("sheets", []):
            rows = _normalize_rows(sheet_spec.get("rows"))
            title = _sheet_title(base_tab_name, sheet_spec.get("title") or "Sheet")
            suffix = 2
            while title in used_titles:
                title = _sheet_title(base_tab_name, f"{sheet_spec.get('title') or 'Sheet'} ({suffix})")
                suffix += 1
            used_titles.add(title)
            tabs.append({
                "title": title,
                "rows": rows,
                "headerRows": set(sheet_spec.get("headerRows") or []),
                "columnFormats": sheet_spec.get("columnFormats") or {},
                "rowCount": max(len(rows) + 20, 100),
                "colCount": max(max((len(r) for r in rows), default=1) + 2, 12),
            })

        # 1) Create missing tabs / grow undersized existing ones — one batchUpdate
        structural = []
        for tab in tabs:
            props = existing.get(tab["title"])
            if props is None:
                structural.append({"addSheet": {"properties": {
                    "title": tab["title"],
                    "gridProperties": {"rowCount": tab["rowCount"], "columnCount": tab["colCount"]},
                }}})
            else:
                grid = props.get("gridProperties", {})
                if grid.get("rowCount", 0) < tab["rowCount"] or grid.get("columnCount", 0) < tab["colCount"]:
                    structural.append({"updateSheetProperties": {
                        "properties": {
                            "sheetId": props["sheetId"],
                            "gridProperties": {
                                "rowCount": max(grid.get("rowCount", 0), tab["rowCount"]),
                                "columnCount": max(grid.get("columnCount", 0), tab["colCount"]),
                            },
                        },
                        "fields": "gridProperties.rowCount,gridProperties.columnCount",
                    }})
        if structural:
            reply = service.spreadsheets().batchUpdate(
                spreadsheetId=sheet_id, body={"requests": structural}
            ).execute()
            for r in reply.get("replies", []):
                added = r.get("addSheet", {}).get("properties")
                if added:
                    existing[added["title"]] = added
        for tab in tabs:
            tab["sheetId"] = existing[tab["title"]]["sheetId"]

        # 2) Clear old values, then write all values in one call
        service.spreadsheets().values().batchClear(
            spreadsheetId=sheet_id,
            body={"ranges": [f"'{tab['title']}'" for tab in tabs]},
        ).execute()
        service.spreadsheets().values().batchUpdate(
            spreadsheetId=sheet_id,
            body={
                "valueInputOption": "USER_ENTERED",
                "data": [
                    {"range": f"'{tab['title']}'!A1", "values": tab["rows"]}
                    for tab in tabs
                ],
            },
        ).execute()

        # 3) Formatting — one batchUpdate for everything
        fmt_requests = []
        for tab in tabs:
            sid = tab["sheetId"]
            n_rows = len(tab["rows"])
            n_cols = max((len(r) for r in tab["rows"]), default=1)

            # Reset any formatting left over from a previous export
            fmt_requests.append({"repeatCell": {
                "range": {"sheetId": sid},
                "cell": {"userEnteredFormat": {}},
                "fields": "userEnteredFormat",
            }})

            # Column number formats (skip the title row)
            for letter, fmt in tab["columnFormats"].items():
                g_fmt = GOOGLE_NUMBER_FORMATS.get(fmt)
                col = _col_letter_to_index(letter)
                if not g_fmt or col < 0:
                    continue
                fmt_requests.append({"repeatCell": {
                    "range": {"sheetId": sid, "startRowIndex": 1, "endRowIndex": n_rows,
                              "startColumnIndex": col, "endColumnIndex": col + 1},
                    "cell": {"userEnteredFormat": {"numberFormat": g_fmt}},
                    "fields": "userEnteredFormat.numberFormat",
                }})

            # Header rows: dark title row, tinted section rows
            for r_idx in sorted(tab["headerRows"]):
                if not isinstance(r_idx, int) or r_idx < 0 or r_idx >= n_rows:
                    continue
                if r_idx == 0:
                    cell_fmt = {
                        "backgroundColor": _TITLE_BG,
                        "textFormat": {"bold": True, "fontSize": 13, "foregroundColor": _TITLE_FG},
                    }
                else:
                    cell_fmt = {
                        "backgroundColor": _SECTION_BG,
                        "textFormat": {"bold": True, "fontSize": 11},
                    }
                fmt_requests.append({"repeatCell": {
                    "range": {"sheetId": sid, "startRowIndex": r_idx, "endRowIndex": r_idx + 1,
                              "startColumnIndex": 0, "endColumnIndex": n_cols},
                    "cell": {"userEnteredFormat": cell_fmt},
                    "fields": "userEnteredFormat(backgroundColor,textFormat)",
                }})

            # Freeze the title row, auto-size columns
            fmt_requests.append({"updateSheetProperties": {
                "properties": {"sheetId": sid, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount",
            }})
            fmt_requests.append({"autoResizeDimensions": {"dimensions": {
                "sheetId": sid, "dimension": "COLUMNS",
                "startIndex": 0, "endIndex": n_cols,
            }}})

        if fmt_requests:
            service.spreadsheets().batchUpdate(
                spreadsheetId=sheet_id, body={"requests": fmt_requests}
            ).execute()

        return {
            "success": True,
            "message": f"Updated {len(tabs)} tabs in Google Sheets",
            "workbookTitle": spec.get("workbookTitle") or "Underwriting Model",
            "updatedTabs": [tab["title"] for tab in tabs],
            "tabCount": len(tabs),
            "rowCount": sum(len(tab["rows"]) for tab in tabs),
        }
    except HttpError as error:
        status = getattr(getattr(error, "resp", None), "status", None)
        log.error(f"[SpreadsheetAI] Google Sheets API error ({status}): {error}")
        if status in (403, 404):
            email = _service_account_email() or "the DealSniper service account"
            return {
                "success": False,
                "errorCode": status,
                "message": (
                    f"DealSniper can't access that spreadsheet. Open it in Google Sheets, "
                    f"click Share, and add {email} as an Editor — then export again."
                ),
            }
        return {"success": False, "errorCode": status, "message": f"Google Sheets API error: {error}"}
    except Exception as error:
        log.exception("[SpreadsheetAI] Error rendering spec to Google Sheets")
        return {"success": False, "message": f"Error rendering workbook to Google Sheets: {error}"}
