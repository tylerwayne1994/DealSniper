"""
AI-Powered Excel Export for DealSniper

Uses Claude to intelligently map scenarioData + fullCalcs into a proper
underwriting model spreadsheet - handles any data structure.
"""

import io
import os
import json
import logging
from typing import Dict, Any, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)


def safe_num(val, default=0):
    """Safely convert to number."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def build_underwriting_excel(scenario_data: Dict[str, Any], full_calcs: Dict[str, Any] = None) -> io.BytesIO:
    """
    Build a complete underwriting model Excel file from scenario data.
    
    This function directly maps the data without needing a template,
    creating a clean underwriting model from scratch.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Underwriting Model"
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    section_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    money_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'
    pct_format = '0.00%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Extract data from scenarioData
    prop = scenario_data.get('property', {})
    pricing = scenario_data.get('pricing_financing', {})
    pnl = scenario_data.get('pnl', {})
    expenses_raw = scenario_data.get('expenses', {})
    financing = scenario_data.get('financing', {})
    unit_mix = scenario_data.get('unit_mix', [])
    value_add = scenario_data.get('value_add', {})
    underwriting = scenario_data.get('underwriting', {})
    
    # Use fullCalcs if provided for calculated values
    calcs = full_calcs or {}
    
    row = 1
    
    # ═══════════════════════════════════════════════════════════════════
    # PROPERTY INFORMATION
    # ═══════════════════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="PROPERTY INFORMATION").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    
    prop_fields = [
        ("Property Name", prop.get('name') or prop.get('address', '')),
        ("Address", prop.get('address', '')),
        ("City", prop.get('city', '')),
        ("State", prop.get('state', '')),
        ("Zip", prop.get('zip', '')),
        ("Units", safe_num(prop.get('units'))),
        ("Year Built", safe_num(prop.get('year_built'))),
        ("Property Type", prop.get('property_type', 'Multifamily')),
        ("Total SF", safe_num(prop.get('total_sf') or prop.get('sqft'))),
    ]
    
    for label, value in prop_fields:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    
    row += 1
    
    # ═══════════════════════════════════════════════════════════════════
    # PURCHASE & FINANCING
    # ═══════════════════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="PURCHASE & FINANCING").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    
    purchase_price = safe_num(pricing.get('purchase_price') or pricing.get('price'))
    loan_amount = safe_num(pricing.get('loan_amount') or calcs.get('loanAmount'))
    down_payment = safe_num(pricing.get('down_payment') or calcs.get('downPayment'))
    interest_rate = safe_num(pricing.get('interest_rate'))
    if interest_rate > 1:
        interest_rate = interest_rate / 100
    
    units = safe_num(prop.get('units'), 1)
    price_per_unit = purchase_price / units if units > 0 else 0
    ltv = loan_amount / purchase_price if purchase_price > 0 else 0
    
    financing_fields = [
        ("Purchase Price", purchase_price, money_format),
        ("Price Per Unit", price_per_unit, money_format),
        ("Loan Amount", loan_amount, money_format),
        ("Down Payment", down_payment, money_format),
        ("LTV", ltv, pct_format),
        ("Interest Rate", interest_rate, pct_format),
        ("Amortization (years)", safe_num(pricing.get('amortization_years'), 30), None),
        ("Loan Term (years)", safe_num(pricing.get('term_years'), 10), None),
    ]
    
    for item in financing_fields:
        label, value = item[0], item[1]
        fmt = item[2] if len(item) > 2 else None
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        if fmt:
            cell.number_format = fmt
        row += 1
    
    row += 1
    
    # ═══════════════════════════════════════════════════════════════════
    # UNIT MIX
    # ═══════════════════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="UNIT MIX").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    
    # Headers
    unit_headers = ["Unit Type", "Sq Ft", "# Units", "Occupied", "Vacant", "Current Rent", "Market Rent"]
    for col, header in enumerate(unit_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    row += 1
    
    # Unit mix data
    if unit_mix:
        for unit in unit_mix:
            # Handle different field names
            unit_type = unit.get('type') or unit.get('unit_type') or unit.get('name') or unit.get('Unit Name', '')
            sqft = safe_num(unit.get('sqft') or unit.get('sf') or unit.get('Sq Ft'))
            num_units = safe_num(unit.get('count') or unit.get('num_units') or unit.get('# Units') or unit.get('units', 1))
            occupied = safe_num(unit.get('occupied') or unit.get('Occupied', num_units))
            vacant = safe_num(unit.get('vacant') or unit.get('Vacant', 0))
            current_rent = safe_num(unit.get('rent') or unit.get('current_rent') or unit.get('Rent'))
            market_rent = safe_num(unit.get('market_rent') or unit.get('proforma_rent') or unit.get('Proforma Rent') or current_rent)
            
            ws.cell(row=row, column=1, value=unit_type)
            ws.cell(row=row, column=2, value=sqft)
            ws.cell(row=row, column=3, value=num_units)
            ws.cell(row=row, column=4, value=occupied)
            ws.cell(row=row, column=5, value=vacant)
            ws.cell(row=row, column=6, value=current_rent).number_format = money_format
            ws.cell(row=row, column=7, value=market_rent).number_format = money_format
            row += 1
    
    row += 1
    
    # ═══════════════════════════════════════════════════════════════════
    # INCOME
    # ═══════════════════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="INCOME").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    
    # Try multiple field names for income
    egi = safe_num(
        pnl.get('effective_gross_income') or 
        pnl.get('egi') or 
        calcs.get('effectiveGrossIncome')
    )
    
    vacancy_pct = safe_num(pnl.get('vacancy_percent') or pnl.get('vacancy_pct') or pnl.get('vacancy', 5))
    if vacancy_pct > 1:
        vacancy_pct = vacancy_pct / 100
    
    other_income = safe_num(pnl.get('other_income') or calcs.get('otherIncome'))
    
    # Calculate GPI from EGI if not directly available
    gpi = safe_num(
        pnl.get('gross_potential_income') or 
        pnl.get('gross_potential_rent') or 
        pnl.get('gross_scheduled_income') or
        pnl.get('income_t12') or
        calcs.get('grossPotentialIncome')
    )
    # If GPI is 0 but we have EGI, back-calculate
    if gpi == 0 and egi > 0:
        # GPI = (EGI - other income) / (1 - vacancy)
        gpi = (egi - other_income) / (1 - vacancy_pct) if vacancy_pct < 1 else egi
    
    vacancy_loss = gpi * vacancy_pct
    
    # Recalc EGI if we didn't have it
    if egi == 0:
        egi = gpi - vacancy_loss + other_income
    
    income_fields = [
        ("Gross Potential Income", gpi, money_format),
        ("Vacancy Rate", vacancy_pct, pct_format),
        ("Vacancy Loss", -vacancy_loss, money_format),
        ("Other Income", other_income, money_format),
        ("Effective Gross Income", egi, money_format),
    ]
    
    for item in income_fields:
        ws.cell(row=row, column=1, value=item[0])
        cell = ws.cell(row=row, column=2, value=item[1])
        if item[2]:
            cell.number_format = item[2]
        row += 1
    
    row += 1
    
    # ═══════════════════════════════════════════════════════════════════
    # OPERATING EXPENSES
    # ═══════════════════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="OPERATING EXPENSES").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    
    # Build expense list from whatever keys exist in expenses_raw
    # Map common variations to display names
    expense_display_names = {
        'property_tax': 'Property Taxes',
        'property_taxes': 'Property Taxes',
        'taxes': 'Property Taxes',
        'real_estate_taxes': 'Property Taxes',
        'insurance': 'Insurance',
        'management': 'Management',
        'property_management_fees': 'Management',
        'management_fee': 'Management',
        'repairs_maintenance': 'Repairs & Maintenance',
        'repairs': 'Repairs & Maintenance',
        'maintenance': 'Repairs & Maintenance',
        'maintenance_related': 'Repairs & Maintenance',
        'water_sewer': 'Water / Sewer',
        'water': 'Water / Sewer',
        'utilities_water': 'Water / Sewer',
        'electricity': 'Electricity',
        'electric': 'Electricity',
        'utilities_electric': 'Electricity',
        'gas': 'Gas',
        'utilities_gas': 'Gas',
        'trash': 'Trash',
        'trash_removal': 'Trash',
        'trash_service': 'Trash',
        'landscaping': 'Landscaping',
        'grounds': 'Landscaping',
        'payroll': 'Payroll',
        'salaries': 'Payroll',
        'salaries_payroll_related': 'Payroll',
        'admin': 'Administrative',
        'administrative': 'Administrative',
        'general_admin': 'Administrative',
        'marketing': 'Marketing',
        'media_advertising': 'Marketing',
        'advertising': 'Marketing',
        'reserves': 'Reserves',
        'replacement_reserves': 'Reserves',
        'capex_reserves': 'Reserves',
        'pest_control': 'Pest Control',
        'pest': 'Pest Control',
        'turnover': 'Turnover / Cleaning',
        'turnover_cleaning': 'Turnover / Cleaning',
        'cleaning': 'Turnover / Cleaning',
        'licenses': 'Licenses & Permits',
        'licenses_permits': 'Licenses & Permits',
        'permits': 'Licenses & Permits',
        'other': 'Other Expenses',
        'other_expenses': 'Other Expenses',
    }
    
    total_expenses = 0
    expense_rows_written = []
    
    for key, value in expenses_raw.items():
        if key.startswith('_') or key in ['total', 'total_expenses']:
            continue
        amount = safe_num(value)
        if amount != 0:
            display_name = expense_display_names.get(key.lower(), key.replace('_', ' ').title())
            ws.cell(row=row, column=1, value=display_name)
            cell = ws.cell(row=row, column=2, value=amount)
            cell.number_format = money_format
            total_expenses += amount
            expense_rows_written.append(display_name)
            row += 1
    
    # Also check for expenses in calcs
    if calcs and 'expenses' in calcs:
        for key, value in calcs['expenses'].items():
            if key.startswith('_') or key in ['total', 'total_expenses']:
                continue
            display_name = expense_display_names.get(key.lower(), key.replace('_', ' ').title())
            if display_name not in expense_rows_written:
                amount = safe_num(value)
                if amount != 0:
                    ws.cell(row=row, column=1, value=display_name)
                    cell = ws.cell(row=row, column=2, value=amount)
                    cell.number_format = money_format
                    total_expenses += amount
                    row += 1
    
    # Use total from pnl if available
    total_from_pnl = safe_num(
        pnl.get('total_expenses') or 
        pnl.get('operating_expenses') or
        calcs.get('totalExpenses')
    )
    if total_from_pnl > 0:
        total_expenses = total_from_pnl
    
    # Total row
    ws.cell(row=row, column=1, value="TOTAL OPERATING EXPENSES").font = Font(bold=True)
    cell = ws.cell(row=row, column=2, value=total_expenses)
    cell.number_format = money_format
    cell.font = Font(bold=True)
    row += 2
    
    # ═══════════════════════════════════════════════════════════════════
    # NET OPERATING INCOME
    # ═══════════════════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="NET OPERATING INCOME").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    
    noi = safe_num(
        pnl.get('noi') or 
        pnl.get('noi_t12') or 
        pnl.get('net_operating_income') or
        calcs.get('noi') or
        (egi - total_expenses)
    )
    
    ws.cell(row=row, column=1, value="NOI")
    cell = ws.cell(row=row, column=2, value=noi)
    cell.number_format = money_format
    cell.font = Font(bold=True, size=14)
    row += 2
    
    # ═══════════════════════════════════════════════════════════════════
    # DEBT SERVICE & CASHFLOW
    # ═══════════════════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="DEBT SERVICE & CASHFLOW").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    
    annual_debt_service = safe_num(
        pricing.get('annual_debt_service') or
        calcs.get('annualDebtService') or
        financing.get('annual_debt_service')
    )
    
    cashflow = noi - annual_debt_service
    monthly_cashflow = cashflow / 12
    
    debt_fields = [
        ("Annual Debt Service", annual_debt_service, money_format),
        ("Annual Cashflow", cashflow, money_format),
        ("Monthly Cashflow", monthly_cashflow, money_format),
    ]
    
    for item in debt_fields:
        ws.cell(row=row, column=1, value=item[0])
        cell = ws.cell(row=row, column=2, value=item[1])
        if item[2]:
            cell.number_format = item[2]
        row += 1
    
    row += 1
    
    # ═══════════════════════════════════════════════════════════════════
    # RETURNS
    # ═══════════════════════════════════════════════════════════════════
    ws.cell(row=row, column=1, value="RETURNS").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    
    cap_rate = noi / purchase_price if purchase_price > 0 else 0
    dscr = noi / annual_debt_service if annual_debt_service > 0 else 0
    coc = cashflow / down_payment if down_payment > 0 else 0
    
    # Override with calcs if available
    cap_rate = safe_num(calcs.get('capRate') or calcs.get('cap_rate')) or cap_rate
    if cap_rate > 1:
        cap_rate = cap_rate / 100
    dscr = safe_num(calcs.get('dscr')) or dscr
    coc = safe_num(calcs.get('cashOnCash') or calcs.get('cash_on_cash')) or coc
    if coc > 1:
        coc = coc / 100
    
    returns_fields = [
        ("Cap Rate", cap_rate, pct_format),
        ("DSCR", dscr, '0.00x'),
        ("Cash-on-Cash Return", coc, pct_format),
        ("Price Per Unit", price_per_unit, money_format),
    ]
    
    for item in returns_fields:
        ws.cell(row=row, column=1, value=item[0])
        cell = ws.cell(row=row, column=2, value=item[1])
        if item[2]:
            cell.number_format = item[2]
        row += 1
    
    row += 1
    
    # ═══════════════════════════════════════════════════════════════════
    # VALUE-ADD ASSUMPTIONS
    # ═══════════════════════════════════════════════════════════════════
    if value_add:
        ws.cell(row=row, column=1, value="VALUE-ADD ASSUMPTIONS").font = header_font
        ws.cell(row=row, column=1).fill = header_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
        
        va_fields = [
            ("Rent Bump ($/unit/mo)", safe_num(value_add.get('rent_bump') or value_add.get('rent_bump_per_unit'))),
            ("RUBS Recovery ($/yr)", safe_num(value_add.get('rubs_recovery') or value_add.get('rubs_annual'))),
            ("Trash Fee ($/unit/mo)", safe_num(value_add.get('trash_fee') or value_add.get('trash_per_unit'))),
            ("Expense Savings ($/yr)", safe_num(value_add.get('expense_savings'))),
        ]
        
        for label, value in va_fields:
            if value > 0:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=value)
                row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    
    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_to_excel_ai(scenario_data: Dict[str, Any], full_calcs: Dict[str, Any] = None) -> io.BytesIO:
    """
    Export scenario data to an Excel underwriting model.
    
    Args:
        scenario_data: The deal's scenario data from frontend
        full_calcs: Optional calculated values from frontend
        
    Returns:
        BytesIO buffer containing the Excel file
    """
    return build_underwriting_excel(scenario_data, full_calcs)
