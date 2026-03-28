"""
Excel Template Export for DealSniper Underwriting Model

Fills the underwriting.xlsx template with scenarioData and returns the completed file.
Uses openpyxl to preserve formulas while only filling input cells.
"""

import io
import os
import logging
from typing import Dict, Any, Optional, List

import openpyxl
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Path to the template file (in backend folder for deployment)
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'underwriting_template.xlsx')


def safe_num(val, default=0):
    """Safely convert a value to a number."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def fill_underwriting_template(scenario_data: Dict[str, Any]) -> io.BytesIO:
    """
    Fill the underwriting.xlsx template with scenario data.
    
    Args:
        scenario_data: The scenarioData object from the frontend
        
    Returns:
        BytesIO buffer containing the filled Excel file
    """
    # Load the template
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")
    
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active
    
    # Extract nested objects with defaults
    property_info = scenario_data.get('property', {})
    pricing_financing = scenario_data.get('pricing_financing', {})
    pnl = scenario_data.get('pnl', {})
    expenses = scenario_data.get('expenses', {})
    financing = scenario_data.get('financing', {})
    value_add = scenario_data.get('value_add', {})
    underwriting = scenario_data.get('underwriting', {})
    unit_mix = scenario_data.get('unit_mix', [])
    
    # Calculate derived values
    purchase_price = safe_num(pricing_financing.get('purchase_price') or pricing_financing.get('price'))
    units = safe_num(property_info.get('units'), 1)
    
    # Occupancy (default 95% if not specified)
    vacancy_pct = safe_num(pnl.get('vacancy_percent') or pnl.get('vacancy_pct'), 5)
    occupancy = 1 - (vacancy_pct / 100 if vacancy_pct > 1 else vacancy_pct)
    
    # Down payment calculation
    loan_amount = safe_num(pricing_financing.get('loan_amount'))
    down_payment = safe_num(pricing_financing.get('down_payment'))
    if purchase_price > 0:
        down_payment_pct = (purchase_price - loan_amount) / purchase_price if loan_amount > 0 else (down_payment / purchase_price if down_payment > 0 else 0.25)
    else:
        down_payment_pct = 0.25
    
    # Interest rate (convert from percent if > 1)
    interest_rate = safe_num(pricing_financing.get('interest_rate'), 6.5)
    if interest_rate > 1:
        interest_rate = interest_rate / 100
    
    # ═══════════════════════════════════════════════════════════════════════
    # FILL INPUT CELLS
    # ═══════════════════════════════════════════════════════════════════════
    
    # --- Deal Snapshot (Row 5) ---
    ws['B5'] = purchase_price  # Purchase Price
    
    # --- Property Information (Rows 19-30) ---
    ws['B19'] = property_info.get('name') or property_info.get('address', '')  # Property Name
    ws['B20'] = property_info.get('address', '')  # Address
    ws['B21'] = property_info.get('city', '')  # City
    ws['B22'] = property_info.get('state', '')  # State
    ws['B23'] = property_info.get('zip', '')  # Zip
    ws['B24'] = property_info.get('county', '')  # County
    ws['B25'] = property_info.get('property_type', 'Multifamily')  # Property Type
    ws['B26'] = int(units) if units else 0  # Units
    ws['B27'] = safe_num(property_info.get('year_built'))  # Year Built
    ws['B28'] = occupancy  # Occupancy Rate (decimal)
    ws['B29'] = safe_num(property_info.get('total_sf') or property_info.get('sqft'))  # Total SF
    ws['B30'] = safe_num(property_info.get('land_sf') or property_info.get('lot_size'))  # Land SF
    
    # --- Purchase & Financing (Rows 32-37) ---
    ws['B32'] = purchase_price  # Purchase Price (duplicate for formula reference)
    ws['B33'] = down_payment_pct  # Down Payment % (as decimal)
    ws['B34'] = interest_rate  # Interest Rate (as decimal)
    ws['B35'] = safe_num(pricing_financing.get('amortization_years'), 30)  # Amortization
    ws['B36'] = safe_num(pricing_financing.get('term_years'), 10)  # Loan Term
    ws['B37'] = safe_num(pricing_financing.get('io_years') or pricing_financing.get('io_period'), 0)  # IO Years
    
    # --- Equity Partner Structure (Rows 46-50) ---
    # Check for equity partner in financing.loans array
    equity_partner = None
    loans = financing.get('loans', [])
    for loan in loans:
        if loan.get('type') == 'Equity Partner' and loan.get('enabled', True):
            equity_partner = loan
            break
    
    if equity_partner:
        ws['B46'] = safe_num(equity_partner.get('loanDollar'))  # Investor Contribution
        ws['B48'] = safe_num(equity_partner.get('rate'), 8) / 100  # Pref Return (as decimal)
        ws['B49'] = safe_num(equity_partner.get('split'), 20) / 100  # Partner Equity %
    else:
        # Use defaults or values from financing object
        investor_contrib = safe_num(financing.get('investor_contribution'))
        ws['B46'] = investor_contrib
        ws['B47'] = max(0, down_payment - investor_contrib)  # Your Contribution
        ws['B48'] = safe_num(financing.get('pref_return'), 8) / 100
        ws['B49'] = safe_num(financing.get('partner_equity_pct'), 20) / 100
    
    # --- Value-Add Assumptions (Rows 52-58) ---
    rent_bump = safe_num(value_add.get('rent_bump') or value_add.get('rent_bump_per_unit'))
    rubs_recovery = safe_num(value_add.get('rubs_recovery') or value_add.get('rubs_annual'))
    trash_fee = safe_num(value_add.get('trash_fee') or value_add.get('trash_per_unit'))
    exit_cap = safe_num(underwriting.get('exit_cap_rate') or underwriting.get('exit_cap'), 7)
    
    ws['B52'] = rent_bump  # Rent Bump $/unit/month
    ws['B53'] = rubs_recovery  # RUBS Recovery $/year total
    ws['B54'] = trash_fee  # Trash Fee $/unit/month
    ws['B55'] = safe_num(value_add.get('other_income'))  # Other Income boost
    ws['B56'] = safe_num(value_add.get('expense_savings'))  # Expense Savings
    ws['B57'] = exit_cap / 100 if exit_cap > 1 else exit_cap  # Exit Cap Rate (decimal)
    ws['B58'] = safe_num(underwriting.get('hold_period'), 5)  # Hold Period (years)
    
    # --- Cashflow Scenario 1: Current (Rows 63-84) ---
    # SGI (Scheduled Gross Income)
    sgi = safe_num(pnl.get('gross_scheduled_income') or pnl.get('income_t12') or pnl.get('gross_potential_rent'))
    ws['B63'] = sgi
    
    # For Scenario 2 (RUBS + Trash) - use same SGI, RUBS/Trash are additive formulas
    ws['C63'] = sgi
    
    # Other income
    ws['B65'] = safe_num(pnl.get('other_income'))
    
    # --- Expense Line Items (Rows 70-84) ---
    expense_mapping = {
        'B70': ['property_tax', 'taxes', 'real_estate_taxes'],
        'B71': ['insurance'],
        # B72 is Management (formula: 6% of GOI)
        'B73': ['repairs_maintenance', 'repairs', 'maintenance', 'r_m'],
        'B74': ['turnover', 'turnover_cleaning', 'cleaning'],
        'B75': ['landscaping', 'grounds'],
        'B76': ['payroll', 'salaries', 'wages'],
        'B77': ['water_sewer', 'water', 'utilities_water'],
        'B78': ['electricity', 'electric', 'utilities_electric'],
        'B79': ['pest_control', 'pest'],
        'B80': ['trash', 'trash_service', 'garbage'],
        'B81': ['admin', 'general_admin', 'administrative'],
        'B82': ['marketing', 'advertising'],
        'B83': ['reserves', 'replacement_reserves', 'capex_reserves'],
        'B84': ['licenses', 'licenses_permits', 'permits'],
    }
    
    for cell, keys in expense_mapping.items():
        value = 0
        for key in keys:
            if key in expenses:
                value = safe_num(expenses[key])
                break
        if value > 0:
            ws[cell] = value
            # Also set for Scenario 2 and 3 (same expenses)
            col = cell[0]
            row = cell[1:]
            ws[f'C{row}'] = value
            ws[f'D{row}'] = value
    
    # --- Sensitivity Analysis Purchase Price Scenarios (Rows 112-120, Column A) ---
    if purchase_price > 0:
        # Generate price scenarios: -10% to +10% in steps
        price_deltas = [-0.10, -0.075, -0.05, -0.025, 0, 0.025, 0.05, 0.075, 0.10]
        for i, delta in enumerate(price_deltas):
            row = 112 + i
            ws[f'A{row}'] = purchase_price * (1 + delta)
    
    # --- Rent Roll (Rows 125+) ---
    if unit_mix:
        for i, unit in enumerate(unit_mix[:50]):  # Max 50 units
            row = 125 + i
            ws[f'A{row}'] = i + 1  # Unit #
            ws[f'B{row}'] = unit.get('type') or unit.get('unit_type') or unit.get('mix', '')
            ws[f'C{row}'] = safe_num(unit.get('current_rent') or unit.get('rent'))
            ws[f'D{row}'] = safe_num(unit.get('market_rent'))
            # E is formula (loss to lease)
            ws[f'F{row}'] = safe_num(unit.get('sqft') or unit.get('sf'))
    
    # Save to BytesIO buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def export_to_excel(scenario_data: Dict[str, Any], full_calcs: Optional[Dict[str, Any]] = None) -> io.BytesIO:
    """
    Export scenario data to a filled Excel template.
    
    Args:
        scenario_data: The deal's scenario data
        full_calcs: Optional calculated values (not needed as template has formulas)
        
    Returns:
        BytesIO buffer containing the Excel file
    """
    return fill_underwriting_template(scenario_data)
