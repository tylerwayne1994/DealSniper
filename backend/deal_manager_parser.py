# Deal Manager Excel Parser
# Parses Deal Manager Excel files using the cell mapping from deal_manager_mapping.csv
# This is an ADD-ON parser - it does NOT replace existing parsing functionality

import os
import csv
import logging
from typing import Dict, Any, Optional, List
from pathlib import Path

log = logging.getLogger("deal_manager_parser")

# Path to the mapping CSV file
MAPPING_FILE = Path(__file__).parent / "deal_manager_mapping.csv"


def load_cell_mapping() -> List[Dict[str, str]]:
    """Load the cell mapping from CSV file"""
    mappings = []
    
    if not MAPPING_FILE.exists():
        log.warning(f"Mapping file not found: {MAPPING_FILE}")
        return mappings
    
    with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            mappings.append(row)
    
    log.info(f"Loaded {len(mappings)} field mappings from {MAPPING_FILE}")
    return mappings


def parse_cell_reference(cell_ref: str) -> tuple:
    """
    Parse Excel cell reference (e.g., 'B7') into (column, row).
    Returns (column_letter, row_number) tuple.
    """
    import re
    match = re.match(r'^([A-Z]+)(\d+)$', cell_ref.upper().strip())
    if match:
        return (match.group(1), int(match.group(2)))
    return (None, None)


def column_letter_to_index(column_letter: str) -> int:
    """Convert Excel column letter to 0-based index (A=0, B=1, etc.)"""
    result = 0
    for char in column_letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1


def convert_value(raw_value: Any, value_type: str) -> Any:
    """Convert raw cell value to appropriate Python type"""
    if raw_value is None or raw_value == "":
        return None
    
    # Convert to string for processing
    val_str = str(raw_value).strip()
    
    if value_type == "Number":
        # Remove commas, dollar signs
        val_str = val_str.replace(',', '').replace('$', '').strip()
        try:
            if '.' in val_str:
                return float(val_str)
            return int(val_str)
        except ValueError:
            return None
            
    elif value_type == "Currency":
        # Remove currency symbols and commas
        val_str = val_str.replace(',', '').replace('$', '').replace('(', '-').replace(')', '').strip()
        try:
            return float(val_str)
        except ValueError:
            return None
            
    elif value_type == "Percentage":
        # Handle percentage values - remove % and convert
        val_str = val_str.replace('%', '').strip()
        try:
            val = float(val_str)
            # If value is already in decimal form (e.g., 0.06), return as-is
            # If value is in percentage form (e.g., 6), convert to decimal
            if abs(val) > 1:
                return val / 100.0
            return val
        except ValueError:
            return None
            
    else:  # Text or default
        return val_str if val_str else None


def set_nested_value(data: Dict, field_path: str, value: Any):
    """
    Set a nested dictionary value using dot notation.
    e.g., 'characteristics.dealName' -> data['characteristics']['dealName'] = value
    """
    if value is None:
        return
        
    parts = field_path.split('.')
    current = data
    
    for i, part in enumerate(parts[:-1]):
        if part not in current:
            current[part] = {}
        current = current[part]
    
    current[parts[-1]] = value


def parse_deal_manager_excel(file_path: str, worksheet_name: str = None) -> Dict[str, Any]:
    """
    Parse a Deal Manager Excel file using the cell mapping.
    
    Args:
        file_path: Path to the Excel file
        worksheet_name: Optional specific worksheet to parse (defaults to first sheet)
    
    Returns:
        Dictionary with parsed data in our standard JSON format
    """
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl is required for Excel parsing. Install with: pip install openpyxl")
        return {"success": False, "error": "openpyxl not installed"}
    
    # Load mapping
    mappings = load_cell_mapping()
    if not mappings:
        return {"success": False, "error": "No cell mappings found"}
    
    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        log.error(f"Failed to load Excel file: {e}")
        return {"success": False, "error": f"Failed to load Excel file: {str(e)}"}
    
    # Get worksheet
    if worksheet_name:
        if worksheet_name not in wb.sheetnames:
            return {"success": False, "error": f"Worksheet '{worksheet_name}' not found"}
        ws = wb[worksheet_name]
    else:
        ws = wb.active
    
    log.info(f"Parsing Deal Manager Excel: {file_path}, worksheet: {ws.title}")
    
    # Initialize result structure
    result = {
        "characteristics": {},
        "economics": {},
        "financing": {},
        "waterfall": {}
    }
    
    # Track what we found
    fields_found = []
    fields_missing = []
    
    # Process each mapping
    for mapping in mappings:
        cell_ref = mapping.get('cell', '').strip()
        json_field = mapping.get('json_field', '').strip()
        value_type = mapping.get('type', 'Text').strip()
        label = mapping.get('label', json_field)
        required = mapping.get('required', 'No').strip().lower() == 'yes'
        
        if not cell_ref or not json_field:
            continue
        
        # Parse cell reference
        col_letter, row_num = parse_cell_reference(cell_ref)
        if not col_letter or not row_num:
            log.warning(f"Invalid cell reference: {cell_ref} for field {json_field}")
            continue
        
        # Get cell value
        try:
            cell = ws[cell_ref]
            raw_value = cell.value
        except Exception as e:
            log.warning(f"Failed to read cell {cell_ref}: {e}")
            raw_value = None
        
        # Convert value
        converted_value = convert_value(raw_value, value_type)
        
        # Set in result
        if converted_value is not None:
            set_nested_value(result, json_field, converted_value)
            fields_found.append(label)
            log.debug(f"  {label} ({cell_ref}): {converted_value}")
        else:
            if required:
                fields_missing.append(label)
    
    wb.close()
    
    # Convert Deal Manager format to our standard format
    standard_format = convert_dm_to_standard_format(result)
    
    log.info(f"Parsed {len(fields_found)} fields from Deal Manager Excel")
    if fields_missing:
        log.warning(f"Missing required fields: {fields_missing}")
    
    return {
        "success": True,
        "data": standard_format,
        "dm_raw": result,
        "fields_found": fields_found,
        "fields_missing": fields_missing,
        "source": "deal_manager_excel"
    }


def convert_dm_to_standard_format(dm_data: Dict) -> Dict[str, Any]:
    """
    Convert Deal Manager format to our standard parsed JSON format.
    Maps Deal Manager fields to the format expected by the application.
    """
    chars = dm_data.get('characteristics', {})
    econ = dm_data.get('economics', {})
    fin = dm_data.get('financing', {})
    waterfall = dm_data.get('waterfall', {})
    
    # Build standard format
    standard = {
        "property": {
            "address": chars.get('propertyAddress', ''),
            "city": chars.get('city', ''),
            "state": "",  # Not in mapping
            "zip": "",  # Not in mapping
            "units": 0,  # Not in mapping
            "year_built": chars.get('yearBuilt'),
            "rba_sqft": chars.get('totalRSF'),
            "land_area_acres": chars.get('landSizeSF'),  # Note: mapping says SF but might be acres
            "property_type": chars.get('propertyType', ''),
            "property_class": chars.get('subPropertyType', ''),
            "parking_spaces": chars.get('parkingSpaces'),
            "floor_count": chars.get('floorCount'),
            "far": chars.get('far'),
            "legal_interest": chars.get('legalInterest'),
            "year_renovated": chars.get('yearRenovated'),
            "deal_name": chars.get('dealName')
        },
        "pricing_financing": {
            "price": econ.get('purchasePrice'),
            "price_per_unit": 0,
            "price_per_sf": 0,
            "loan_amount": 0,  # Calculate from LTV
            "down_payment": 0,
            "interest_rate": fin.get('interestRate'),
            "ltv": fin.get('ltv'),
            "term_years": fin.get('term'),
            "amortization_years": fin.get('amortPeriod'),
            "io_period": fin.get('IO'),
            "loan_fees": fin.get('loanFees'),
            "acquisition_price": econ.get('acquisitionPrice'),
            "acquisition_fees": econ.get('acquisitionFees'),
            "due_diligence_costs": econ.get('dueDiligenceCosts'),
            "construction_costs": econ.get('constructionCosts'),
            "cap_rate": econ.get('capRate')
        },
        "pnl": {
            "gross_potential_rent": 0,  # Calculate from base rent
            "other_income": econ.get('otherIncome'),
            "vacancy_rate": econ.get('vacancy'),
            "vacancy_amount": 0,
            "effective_gross_income": 0,
            "operating_expenses": 0,
            "noi": econ.get('noi'),
            "cap_rate": econ.get('capRate'),
            "base_rent_psf": econ.get('baseRent'),
            "rent_escalations": econ.get('rentEscalations'),
            "other_income_escalations": econ.get('otherIncomeEscalations'),
            "recoveries": econ.get('recoveries')
        },
        "expenses": {
            "taxes": 0,
            "insurance": 0,
            "utilities": 0,
            "repairs_maintenance": 0,
            "management": 0,
            "payroll": 0,
            "admin": 0,
            "marketing": 0,
            "other": 0,
            "total": 0,
            "property_tax_growth": econ.get('propertyTaxGrowth'),
            "immediate_capex": econ.get('immediateCapEx'),
            "ongoing_capex_psf": econ.get('ongoingCapEx'),
            "capex_growth": econ.get('ongoingCapExEscalations')
        },
        "underwriting": {
            "holding_period": chars.get('holdingPeriod'),
            "exit_cap_rate": chars.get('exitCapRate'),
            "discount_rate": chars.get('discountRate'),
            "incoming_cap_rate": chars.get('incomingCapRate'),
            "selling_costs": chars.get('sellingCosts'),
            "replacement_cost": chars.get('replacementCost'),
            "irr": econ.get('irr'),
            "cash_on_cash": econ.get('cashOnCashReturn'),
            "dscr": 0
        },
        "waterfall": {
            "lp_equity_pct": waterfall.get('LPEquity'),
            "preferred_return_type": waterfall.get('preferredReturnType'),
            "first_hurdle_return": waterfall.get('firstHurdle'),
            "first_hurdle_lp_split": waterfall.get('firstHurdleLP'),
            "first_hurdle_gp_split": waterfall.get('firstHurdleGP'),
            "second_hurdle_return": waterfall.get('secondHurdle'),
            "second_hurdle_lp_split": waterfall.get('secondHurdleLP'),
            "second_hurdle_gp_split": waterfall.get('secondHurdleGP'),
            "third_hurdle_return": waterfall.get('thirdHurdle'),
            "third_hurdle_lp_split": waterfall.get('thirdHurdleLP'),
            "third_hurdle_gp_split": waterfall.get('thirdHurdleGP')
        },
        "unit_mix": [],  # Not in Deal Manager mapping
        "data_quality": {
            "confidence": 0.95,  # High confidence since we're reading exact cells
            "source": "deal_manager_excel",
            "missing_fields": [],
            "assumptions": []
        }
    }
    
    # Calculate derived values
    purchase_price = standard["pricing_financing"].get("price") or 0
    ltv = standard["pricing_financing"].get("ltv") or 0
    
    if purchase_price and ltv:
        standard["pricing_financing"]["loan_amount"] = purchase_price * ltv
        standard["pricing_financing"]["down_payment"] = purchase_price * (1 - ltv)
    
    # Calculate GPR from base rent if we have sqft
    base_rent_psf = standard["pnl"].get("base_rent_psf") or 0
    sqft = standard["property"].get("rba_sqft") or 0
    if base_rent_psf and sqft:
        standard["pnl"]["gross_potential_rent"] = base_rent_psf * sqft
    
    return standard


def parse_deal_manager_bytes(file_bytes: bytes, filename: str = "deal_manager.xlsx") -> Dict[str, Any]:
    """
    Parse Deal Manager Excel from bytes (for API uploads).
    
    Args:
        file_bytes: Raw file bytes
        filename: Original filename (used for temp file extension)
    
    Returns:
        Dictionary with parsed data
    """
    import tempfile
    import os
    
    # Determine extension from filename
    ext = Path(filename).suffix.lower() or '.xlsx'
    
    # Write to temp file for openpyxl
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    
    try:
        result = parse_deal_manager_excel(tmp_path)
    finally:
        # Clean up temp file
        try:
            os.unlink(tmp_path)
        except:
            pass
    
    return result


def is_deal_manager_file(file_bytes: bytes, filename: str = "") -> bool:
    """
    Detect if a file is a Deal Manager Excel file.
    Uses heuristics to check for Deal Manager structure.
    """
    try:
        import openpyxl
        import tempfile
        from io import BytesIO
    except ImportError:
        return False
    
    # Check filename hint
    filename_lower = filename.lower()
    if 'deal_manager' in filename_lower or 'dealmanager' in filename_lower:
        return True
    
    # Check file extension
    if not filename_lower.endswith('.xlsx') and not filename_lower.endswith('.xls'):
        return False
    
    # Try to open and check for Deal Manager structure
    try:
        # Write to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
            ws = wb.active
            
            # Check for Deal Manager indicators
            # Look for specific cells that would indicate Deal Manager format
            dm_indicators = [
                ('B7', 'deal name'),
                ('B8', 'property address'),
                ('B10', 'property type'),
                ('B15', 'purchase price'),
            ]
            
            matches = 0
            for cell_ref, expected_hint in dm_indicators:
                try:
                    # Check if there's a label nearby (typically column A)
                    label_cell = 'A' + cell_ref[1:]
                    label_val = str(ws[label_cell].value or '').lower()
                    if expected_hint in label_val or any(word in label_val for word in expected_hint.split()):
                        matches += 1
                except:
                    pass
            
            wb.close()
            
            # If we match 2+ indicators, likely Deal Manager
            return matches >= 2
            
        finally:
            try:
                import os
                os.unlink(tmp_path)
            except:
                pass
                
    except Exception as e:
        log.debug(f"Error checking for Deal Manager format: {e}")
        return False


# CLI for testing
if __name__ == "__main__":
    import argparse
    import json
    
    logging.basicConfig(level=logging.DEBUG)
    
    parser = argparse.ArgumentParser(description="Parse Deal Manager Excel files")
    parser.add_argument("file", help="Path to Deal Manager Excel file")
    parser.add_argument("--worksheet", "-w", help="Worksheet name (default: active sheet)")
    parser.add_argument("--output", "-o", help="Output JSON file path")
    
    args = parser.parse_args()
    
    result = parse_deal_manager_excel(args.file, args.worksheet)
    
    if result.get("success"):
        print(f"\n✅ Parsed {len(result.get('fields_found', []))} fields successfully")
        print(f"\nStandard Format Data:")
        print(json.dumps(result['data'], indent=2, default=str))
        
        if args.output:
            with open(args.output, 'w') as f:
                json.dump(result, f, indent=2, default=str)
            print(f"\nSaved to: {args.output}")
    else:
        print(f"\n❌ Parse failed: {result.get('error')}")
